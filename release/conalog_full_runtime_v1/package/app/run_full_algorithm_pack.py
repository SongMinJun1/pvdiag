#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

PACKAGE_ROOT = Path(__file__).resolve().parents[1]
if str(PACKAGE_ROOT) not in sys.path:
    sys.path.insert(0, str(PACKAGE_ROOT))
from research.prognostics.heuristic_display_registry_v1 import (
    DISPLAY_HEURISTIC_NAME_MAP,
    HEURISTIC_DISPLAY_NOTE_MAP,
    display_heuristic_name as shared_display_heuristic_name,
    display_heuristic_note as shared_display_heuristic_note,
)

DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")
DEFAULT_SITES = ["conalog", "gangui", "ktc_ess"]
CORE_DIGEST_COLUMNS = [
    "date",
    "panel_id",
    "confirmed_fault",
    "critical_fault",
    "critical_source",
    "final_fault",
    "anom_level",
    "anom_subtype",
]
LIVE_FAULT_COMPARE_COLS = [
    "site",
    "panel_id",
    "패널고장여부_ko",
    "사건유형_ko",
    "최종고장양상_ko",
    "커널로그_원인군_ko",
    "1순위_의심원인_ko",
    "2순위_의심원인_ko",
    "3순위_의심원인_ko",
]
LIVE_FAULT_OUTPUT_COLS = [
    *LIVE_FAULT_COMPARE_COLS,
    "전조날짜",
    "고장날짜",
]
LIVE_PREVIEW_OUTPUT_COLS = [
    "site",
    "panel_id",
    "패널고장여부_ko",
    "사건유형_ko",
    "최종고장양상_ko",
    "전조날짜",
    "고장날짜",
    "라벨된 fault",
    "1순위_의심원인_ko",
    "2순위_의심원인_ko",
    "3순위_의심원인_ko",
    "커널로그 기존 알고리즘",
]
USER_PREVIEW_OUTPUT_COLS = [
    "site",
    "panel_id",
    "전조날짜",
    "고장 기준일",
    "운영 판정",
    "급락 종결 관측",
    "점진 저하 누적",
    "사건 종결 요약",
    "상위 해석 후보",
    "기존 알고리즘 source",
]
SIGNAL_PREVIEW_OUTPUT_COLS = [
    "site",
    "panel_id",
    "전조날짜",
    "신호 기준일",
    "운영 판정",
    "급락 종결 관측",
    "점진 저하 누적",
    "사건 종결 요약",
    "상위 해석 후보",
    "기존 알고리즘 source",
]
PRECURSOR_REPORT_OUTPUT_COLS = [
    "site",
    "panel_id",
    "운영 판정",
    "판정 근거",
    "전조날짜",
    "전조 축",
    "대표 전조 신호",
    "전조 요약",
    "상위 해석 후보",
    "기존 알고리즘 source",
    "패턴 설명",
    "모니터링 권고",
    "공통원인 위험",
    "권고 검토 레인",
    "EWS 전조 일수",
    "pre_alarm 일수",
    "pre_ews 일수",
    "Option B 유효 일수",
    "공통원인 겹침 일수",
    "AE 전조 조건 일수",
    "DTW 전조 조건 일수",
]
FAULT_SIGNAL_REPORT_OUTPUT_COLS = [
    "site",
    "group root",
    "subgroup base",
    "subgroup cluster",
    "panel_id",
    "동일 subgroup row 수",
    "동일 cluster row 수",
    "운영 판정",
    "확정 경로",
    "고장 신호 요약",
    "전조 시작일",
    "신호 기준일",
    "사건유형",
    "사건 종결 요약",
    "근접 공통원인",
    "상위 해석 후보",
    "기존 알고리즘 source",
    "패턴 설명",
    "현장 점검 권고",
]
ROOT_LIVE_FAULT_NAME = "fault_panel_result_current_v1.csv"
ROOT_LIVE_PREVIEW_NAME = "fault_panel_result_current_preview_v1.csv"
ROOT_LIVE_SUMMARY_NAME = "live_chain_summary_v1.json"
ROOT_LIVE_REPORT_NAME = "fault_panel_result_current_report_v1.md"
ROOT_RAWONLY_FAULT_NAME = "fault_panel_result_raw_only_current_v1.csv"
ROOT_RAWONLY_PREVIEW_NAME = "fault_panel_result_raw_only_current_preview_v1.csv"
ROOT_RAWONLY_SUMMARY_NAME = "raw_only_chain_summary_v1.json"
ROOT_RAWONLY_REPORT_NAME = "fault_panel_result_raw_only_current_report_v1.md"
ROOT_MASTER_REPORT_NAME = "fault_panel_result_master_report_v1.md"
ROOT_DETAILED_REPORT_NAME = "fault_panel_result_detailed_report_v1.xlsx"
ROOT_PRECURSOR_REPORT_NAME = "fault_panel_result_precursor_report_v1.csv"
ROOT_FAULT_SIGNAL_REPORT_NAME = "fault_panel_result_raw_only_fault_signal_report_v1.csv"
RAW_ONLY_STRICT_CURRENT_GRADES = {"확정"}
FAULT_SIGNAL_CLUSTER_GAP_DAYS = 3
MAIL_BUCKET_ALGORITHM_MAP = {
    ("conalog", "7f7dd654-2760-4eb2-a197-3ebb72b85cda.2.0"): "panel-bypass",
    ("conalog", "c42997a6-5881-47e7-9035-7de8a2673b54.1.1"): "disconnection",
    ("gangui", "bf1a912f-6cf0-4f12-8e97-9d9d86576511.0.7"): "panel-bypass",
    ("gangui", "bf1a912f-6cf0-4f12-8e97-9d9d86576511.2.16"): "panel-bypass",
    ("ktc_ess", "10305b40-b67e-40d1-9cd1-271b6642a3d9.2.12"): "panel-bypass",
    ("ktc_ess", "70ad2d87-cdb6-4842-81b7-71c7599bbf05.1.4"): "panel-bypass",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Run the real panel_day_engine.py for the packaged baseline sites under a data root, "
            "export the fixed fault result artifacts, and write a shadow-compare report for engine core outputs."
        )
    )
    parser.add_argument(
        "--data-root",
        type=Path,
        required=True,
        help="Folder containing site/raw subdirectories such as data-root/conalog/raw.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        required=True,
        help="Folder where site-wise engine outputs and fixed result tables will be written.",
    )
    parser.add_argument(
        "--sites",
        default=",".join(DEFAULT_SITES),
        help="Comma-separated site list. Defaults to conalog,gangui,ktc_ess.",
    )
    parser.add_argument(
        "--train-days",
        type=int,
        default=60,
        help="Maximum number of early days to reserve for training window proposal.",
    )
    parser.add_argument("--pattern", default="*.csv", help="Filename pattern for raw daily CSVs.")
    parser.add_argument("--epochs", type=int, default=40, help="Engine epochs. Defaults to panel_day_engine.py default.")
    parser.add_argument("--latent", type=int, default=16, help="Engine latent size. Defaults to panel_day_engine.py default.")
    parser.add_argument("--device", default="cpu", help="Torch device to pass through to panel_day_engine.py.")
    parser.add_argument(
        "--prefer-existing-site-outs",
        choices=["auto", "on", "off"],
        default="auto",
        help=(
            "Whether to automatically reuse data-root/<site>/out when available. "
            "Defaults to auto."
        ),
    )
    parser.add_argument(
        "--reuse-existing-site-outs-root",
        type=Path,
        default=None,
        help=(
            "Optional root containing precomputed data/<site>/out trees. "
            "When provided, the runner copies those outputs into the runtime workspace and skips engine execution."
        ),
    )
    parser.add_argument(
        "--run-live-chain",
        choices=["on", "off"],
        default="on",
        help="After engine execution, run the packaged bootstrap verdict -> audit -> final verdict live chain. Defaults to on.",
    )
    parser.add_argument(
        "--run-raw-only-chain",
        choices=["on", "off"],
        default="on",
        help="After engine execution, run the packaged raw-only audit -> verdict -> heuristic chain. Defaults to on.",
    )
    parser.add_argument(
        "--workspace-retention",
        choices=["full", "result-only"],
        default="full",
        help=(
            "Controls post-run retention for large intermediate workspaces. "
            "full keeps the historical behavior. result-only keeps result artifacts and share outputs, "
            "then removes duplicate site/output and chain data copies."
        ),
    )
    parser.add_argument("--dry-run", action="store_true", help="Validate paths and emit the execution plan without running the engine.")
    return parser.parse_args()


def package_root() -> Path:
    return Path(__file__).resolve().parents[1]


def engine_path() -> Path:
    return package_root() / "pv_ae" / "panel_day_engine.py"


def fixed_fault6_table_path() -> Path:
    return package_root() / "artifacts" / "fault6_fixed_result_table_v1.csv"


def fixed_fault6_preview_path() -> Path:
    return package_root() / "artifacts" / "fault6_label_and_algorithm_preview_v1.csv"


def baseline_manifest_path() -> Path:
    return package_root() / "artifacts" / "input_baseline_manifest_v1.json"


def baseline_core_digest_path() -> Path:
    return package_root() / "artifacts" / "panel_day_core_baseline_digest_v1.json"


def fault6_provenance_path() -> Path:
    return package_root() / "artifacts" / "fault6_fixed_result_provenance_v1.json"


def dependency_audit_json_path() -> Path:
    return package_root() / "artifacts" / "runtime_chain_dependency_audit_v1.json"


def dependency_audit_md_path() -> Path:
    return package_root() / "artifacts" / "runtime_chain_dependency_audit_v1.md"


def packaged_share_root() -> Path:
    return package_root() / "_share"


def packaged_pipeline_root() -> Path:
    return package_root() / "research" / "prognostics"


def packaged_script_path(name: str) -> Path:
    return packaged_pipeline_root() / name


def extract_date_from_name(path: Path) -> pd.Timestamp:
    match = DATE_RE.search(path.name)
    if not match:
        return pd.NaT
    return pd.to_datetime(match.group(1), errors="coerce").normalize()


def normalize_sites(raw_sites: str) -> list[str]:
    sites = [token.strip() for token in str(raw_sites).split(",") if token.strip()]
    if not sites:
        raise SystemExit("at least one site must be provided")
    return sites


def scan_site_files(data_root: Path, site: str, pattern: str) -> tuple[pd.Timestamp, pd.Timestamp, list[Path]]:
    raw_dir = data_root / site / "raw"
    if not raw_dir.exists():
        raise SystemExit(f"missing raw dir for site={site}: {raw_dir}")
    files = sorted(path for path in raw_dir.glob(pattern) if path.is_file() and path.suffix.lower() == ".csv")
    if not files:
        raise SystemExit(f"raw csv not found for site={site}: {raw_dir}")
    valid_dates = [value for value in (extract_date_from_name(path) for path in files) if pd.notna(value)]
    if not valid_dates:
        raise SystemExit(f"no YYYY-MM-DD found in filenames for site={site}: {raw_dir}")
    return min(valid_dates), max(valid_dates), files


def propose_windows(min_date: pd.Timestamp, max_date: pd.Timestamp, train_days: int) -> dict[str, str]:
    span_days = int((max_date - min_date).days)
    proposed = min(int(train_days) - 1, max(14, int(span_days * 0.30)))
    if proposed < 1:
        proposed = 1

    train_start = min_date
    train_end = min_date + pd.Timedelta(days=proposed)
    if train_end >= max_date:
        train_end = max_date - pd.Timedelta(days=1)
    if train_end < min_date:
        raise SystemExit("date span too short to propose train/eval windows")

    eval_start = train_end + pd.Timedelta(days=1)
    eval_end = max_date
    if eval_start > eval_end:
        raise SystemExit("date span too short to propose eval window")

    return {
        "train_start": str(train_start.date()),
        "train_end": str(train_end.date()),
        "eval_start": str(eval_start.date()),
        "eval_end": str(eval_end.date()),
        "input_date_min": str(min_date.date()),
        "input_date_max": str(max_date.date()),
    }


def site_manifest(files: list[Path]) -> dict[str, object]:
    date_tokens = [match.group(1) for path in files if (match := DATE_RE.search(path.name))]
    return {
        "file_count": int(len(files)),
        "total_bytes": int(sum(path.stat().st_size for path in files)),
        "first_filenames": [path.name for path in files[:5]],
        "last_filenames": [path.name for path in files[-5:]],
        "min_date": min(date_tokens) if date_tokens else "",
        "max_date": max(date_tokens) if date_tokens else "",
    }


def build_site_plan(args: argparse.Namespace, site: str) -> tuple[dict[str, object], list[str]]:
    data_root = args.data_root.expanduser().resolve()
    output_root = args.output_root.expanduser().resolve()
    site_output_dir = output_root / "sites" / site / "output"
    site_log_dir = output_root / "sites" / site / "log"
    site_output_dir.mkdir(parents=True, exist_ok=True)
    site_log_dir.mkdir(parents=True, exist_ok=True)

    min_date, max_date, files = scan_site_files(data_root, site, args.pattern)
    windows = propose_windows(min_date, max_date, args.train_days)
    cmd = [
        sys.executable,
        str(engine_path()),
        "--site",
        site,
        "--data-root",
        str(data_root),
        "--out-dir",
        str(site_output_dir),
        "--log-dir",
        str(site_log_dir),
        "--pattern",
        args.pattern,
        "--train-start",
        windows["train_start"],
        "--train-end",
        windows["train_end"],
        "--eval-start",
        windows["eval_start"],
        "--eval-end",
        windows["eval_end"],
        "--epochs",
        str(args.epochs),
        "--latent",
        str(args.latent),
        "--device",
        args.device,
    ]
    plan = {
        "site": site,
        "raw_dir": str(data_root / site / "raw"),
        "output_dir": str(site_output_dir),
        "log_dir": str(site_log_dir),
        "windows": windows,
        "file_manifest": site_manifest(files),
        "command": cmd,
    }
    return plan, cmd


def write_json(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def emit_progress(percent: int, message: str) -> None:
    safe_percent = max(0, min(100, int(percent)))
    print(f"[{safe_percent:03d}%] {message}", flush=True)


def load_baseline_manifest() -> dict[str, object]:
    path = baseline_manifest_path()
    if not path.exists():
        raise SystemExit(f"missing packaged baseline manifest: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def compare_to_baseline(site_plans: list[dict[str, object]]) -> dict[str, object]:
    baseline = load_baseline_manifest()
    comparison: dict[str, object] = {"all_sites_match": True, "sites": {}}
    baseline_sites = baseline.get("sites", {})
    for plan in site_plans:
        site = str(plan["site"])
        actual = plan["file_manifest"]
        expected = baseline_sites.get(site, {})
        site_match = True
        diffs: list[str] = []
        for key in ["file_count", "total_bytes", "min_date", "max_date"]:
            if actual.get(key) != expected.get(key):
                site_match = False
                diffs.append(f"{key}: expected={expected.get(key)} actual={actual.get(key)}")
        comparison["sites"][site] = {
            "match": site_match,
            "expected": expected,
            "actual": actual,
            "diffs": diffs,
        }
        if not site_match:
            comparison["all_sites_match"] = False
    comparison["note_ko"] = (
        "all_sites_match=1 이면 packaged fixed result table을 만든 baseline raw corpus와 현재 입력의 경량 fingerprint가 일치한다. "
        "일치하지 않으면 engine은 실행될 수 있어도 fixed result table exact replay 보장은 약해진다."
    )
    return comparison


def copy_fixed_results(output_root: Path) -> dict[str, str]:
    output_dir = output_root / "result"
    output_dir.mkdir(parents=True, exist_ok=True)
    fault6_dest = output_dir / "fault6_fixed_result_table_v1.csv"
    preview_dest = output_dir / "fault6_label_and_algorithm_preview_v1.csv"
    shutil.copy2(fixed_fault6_table_path(), fault6_dest)
    if fixed_fault6_preview_path().exists():
        preview_df = pd.read_csv(fixed_fault6_preview_path(), encoding="utf-8-sig", low_memory=False)
        to_user_preview_schema(preview_df).to_csv(preview_dest, index=False, encoding="utf-8-sig")
    return {
        "fault6_fixed_result_table_v1": str(fault6_dest),
        "fault6_label_and_algorithm_preview_v1": str(preview_dest),
    }


def copy_tree(source: Path, target: Path) -> None:
    if not source.exists():
        raise SystemExit(f"missing source tree: {source}")
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source, target, dirs_exist_ok=True)


def path_size_bytes(path: Path) -> int:
    if not path.exists():
        return 0
    if path.is_file() or path.is_symlink():
        return int(path.lstat().st_size)
    total = 0
    for child in path.rglob("*"):
        try:
            total += int(child.lstat().st_size)
        except OSError:
            continue
    return total


def remove_workspace_path(path: Path) -> dict[str, object]:
    exists_before = path.exists() or path.is_symlink()
    size_before = path_size_bytes(path)
    if exists_before:
        if path.is_dir() and not path.is_symlink():
            shutil.rmtree(path)
        else:
            path.unlink()
    return {
        "path": str(path),
        "exists_before": bool(exists_before),
        "size_bytes_before": int(size_before),
        "removed": bool(exists_before),
    }


def apply_workspace_retention(output_root: Path, retention: str) -> dict[str, object]:
    report: dict[str, object] = {
        "workspace_retention": retention,
        "status": "full_workspace_retained",
        "removed_paths": [],
        "kept_paths": [
            str(output_root / "result"),
            str(output_root / "shadow_compare_v1.json"),
            str(output_root / "run_metadata_v1.json"),
        ],
        "bytes_removed_estimate": 0,
        "note_ko": (
            "full 모드는 기존처럼 site output/workspace data를 모두 보존한다. "
            "result-only 모드는 재생성 가능한 대용량 중복 data copy만 제거하고 result 및 _share 산출물은 보존한다."
        ),
    }
    if retention == "full":
        return report

    removable_paths = [
        output_root / "sites",
        output_root / "live_chain_workspace" / "data",
        output_root / "raw_only_chain_workspace" / "data",
    ]
    removed = [remove_workspace_path(path) for path in removable_paths]
    report["removed_paths"] = removed
    report["bytes_removed_estimate"] = int(sum(int(item["size_bytes_before"]) for item in removed))
    report["status"] = "result_and_share_artifacts_retained"
    for share_path in [
        output_root / "live_chain_workspace" / "_share",
        output_root / "raw_only_chain_workspace" / "_share",
    ]:
        if share_path.exists():
            report["kept_paths"].append(str(share_path))
    return report


def copy_existing_site_outs(reuse_root: Path, output_root: Path, sites: list[str]) -> dict[str, str]:
    copied: dict[str, str] = {}
    for site in sites:
        source = reuse_root / site / "out"
        target = output_root / "sites" / site / "output"
        if not source.exists():
            raise SystemExit(f"missing precomputed out dir for site={site}: {source}")
        if target.exists():
            shutil.rmtree(target)
        copy_tree(source, target)
        copied[site] = str(target)
    return copied


def site_outs_available(root: Path, sites: list[str]) -> bool:
    for site in sites:
        if not (root / site / "out" / "panel_day_core.csv").exists():
            return False
    return True


def raw_latest_mtime(root: Path, site: str) -> float | None:
    raw_dir = root / site / "raw"
    if not raw_dir.exists():
        return None
    mtimes = [path.stat().st_mtime for path in raw_dir.glob("*.csv") if path.is_file()]
    return max(mtimes) if mtimes else None


def site_outs_freshness(root: Path, sites: list[str]) -> dict[str, object]:
    site_entries: dict[str, object] = {}
    all_fresh = True
    for site in sites:
        out_path = root / site / "out" / "panel_day_core.csv"
        raw_mtime = raw_latest_mtime(root, site)
        out_exists = out_path.exists()
        out_mtime = out_path.stat().st_mtime if out_exists else None
        fresh = bool(out_exists and raw_mtime is not None and out_mtime is not None and out_mtime >= raw_mtime)
        site_entries[site] = {
            "panel_day_core_exists": out_exists,
            "raw_latest_mtime": raw_mtime,
            "panel_day_core_mtime": out_mtime,
            "fresh_enough": fresh,
        }
        if not fresh:
            all_fresh = False
    return {"all_fresh": all_fresh, "sites": site_entries}


def resolve_reuse_existing_site_outs_root(
    data_root: Path,
    explicit_reuse_root: Path | None,
    prefer_existing_site_outs: str,
    sites: list[str],
) -> tuple[Path | None, str, dict[str, object]]:
    if explicit_reuse_root is not None:
        return explicit_reuse_root, "explicit", {"mode": "explicit", "sites": {}}

    if prefer_existing_site_outs == "off":
        return None, "disabled", {"mode": "disabled", "sites": {}}

    if site_outs_available(data_root, sites):
        freshness = site_outs_freshness(data_root, sites)
        if freshness["all_fresh"]:
            return data_root, "auto_fresh" if prefer_existing_site_outs == "auto" else "forced_fresh", freshness
        if prefer_existing_site_outs == "on":
            raise SystemExit(
                "prefer-existing-site-outs=on 이지만 data-root/<site>/out 가 raw보다 오래되었음"
            )
        return None, "auto_stale_out", freshness

    if prefer_existing_site_outs == "on":
        raise SystemExit(
            f"prefer-existing-site-outs=on 이지만 data-root 아래 precomputed out를 찾지 못함: {data_root}"
        )

    return None, "not_available", {"mode": "not_available", "sites": {}}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def truthy_mask(series: pd.Series) -> pd.Series:
    lowered = series.astype(str).str.strip().str.lower()
    return lowered.isin({"1", "true", "t", "yes"})


def ensure_columns(df: pd.DataFrame, required: list[str], name: str) -> None:
    missing = [column for column in required if column not in df.columns]
    if missing:
        raise SystemExit(f"{name} missing columns: {missing}")


def row_key(site: object, panel_id: object) -> tuple[str, str]:
    return normalize_text(site), normalize_text(panel_id)


def display_heuristic_name(raw_label: object) -> str:
    return shared_display_heuristic_name(raw_label)


def display_heuristic_note(raw_label: object) -> str:
    return shared_display_heuristic_note(raw_label)


def choose_display_precursor_date(
    event_type_ko: object,
    interpreted_onset_date: object,
    first_warning_date: object,
) -> str:
    if normalize_text(event_type_ko) != "전조형 고장":
        return ""
    onset_date = normalize_text(interpreted_onset_date)
    if onset_date:
        return onset_date
    return normalize_text(first_warning_date)


def choose_display_fault_date(
    fault_date: object,
    strict_trigger_date: object,
    first_final_fault_date: object,
) -> str:
    for candidate in [fault_date, strict_trigger_date, first_final_fault_date]:
        text = normalize_text(candidate)
        if text:
            return text
    return ""


def display_preview_precursor_date(value: object) -> str:
    text = normalize_text(value)
    return text if text else "전조없음"


def display_signal_grade(row: pd.Series) -> str:
    grade = normalize_text(row.get("운영해석등급_ko"))
    if not grade:
        grade = normalize_text(row.get("운영 판정"))
    if not grade:
        grade = normalize_text(row.get("현재상태"))
    if grade:
        if grade in {"고장 신호 포착", "고장 확정"}:
            return "확정"
        if grade == "강한 이상징후":
            return "고위험 관찰"
        if grade == "이상징후":
            return "관찰"
        return grade
    if normalize_text(row.get("패널고장여부_ko")) == "고장":
        return "확정"
    return ""


def display_existing_algorithm_source(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return "미검출"
    if text.lower() == "none":
        return "미검출"
    if text == "기존 알고리즘 미검출":
        return "미검출"
    return text


def as_int(value: object) -> int:
    try:
        parsed = int(float(value))
    except (TypeError, ValueError):
        return 0
    return parsed


def is_truthy_scalar(value: object) -> bool:
    text = normalize_text(value).lower()
    return text in {"1", "true", "t", "yes", "y"}


def event_summary_from_labels(event_type: object, terminal_pattern: object) -> str:
    event = normalize_text(event_type)
    terminal = normalize_text(terminal_pattern)
    mapping = {
        ("전조형 고장", "급격 종료"): "전조 후 급격 종료",
        ("전조형 고장", "진행성 악화"): "전조 후 진행 악화",
        ("급작 고장", "급작 발생"): "급작 발생",
    }
    return mapping.get((event, terminal), "")


def event_display_fields(record: pd.Series | dict[str, object]) -> dict[str, str]:
    existing_abrupt = normalize_text(record.get("급락 종결 관측"))
    existing_progressive = normalize_text(record.get("점진 저하 누적"))
    existing_summary = normalize_text(record.get("사건 종결 요약"))
    if existing_abrupt or existing_progressive or existing_summary:
        return {
            "급락 종결 관측": existing_abrupt or "없음",
            "점진 저하 누적": existing_progressive or "없음",
            "사건 종결 요약": existing_summary,
        }

    event_type = normalize_text(record.get("사건유형_ko")) or normalize_text(record.get("사건 해석"))
    terminal_pattern = normalize_text(record.get("최종고장양상_ko")) or normalize_text(
        record.get("최종고장양상")
    )
    precursor_date = display_preview_precursor_date(record.get("전조날짜"))
    grade = normalize_text(record.get("운영해석등급_ko")) or normalize_text(record.get("운영 판정"))
    if not grade and isinstance(record, pd.Series):
        grade = display_signal_grade(record)

    abrupt_observed = (
        terminal_pattern in {"급격 종료", "급작 발생"}
        or as_int(record.get("final_days")) > 0
        or is_truthy_scalar(record.get("대표final_fault"))
        or is_truthy_scalar(record.get("final_fault"))
    )
    progressive_observed = (
        terminal_pattern == "진행성 악화"
        or event_type == "전조형 고장"
        or "degradation" in normalize_text(record.get("anom_subtypes_csv")).lower()
        or "degradation" in normalize_text(record.get("대표anom_subtype")).lower()
        or as_int(record.get("ews_warning_days")) > 0
        or as_int(record.get("pre_alarm_days")) > 0
        or as_int(record.get("pre_ews_days")) > 0
        or as_int(record.get("prefault_cond_ae_days")) > 0
        or as_int(record.get("prefault_cond_dtw_days")) > 0
        or precursor_date != "전조없음"
    )

    summary = ""
    if grade == "확정" or normalize_text(record.get("패널고장여부_ko")) == "고장":
        summary = event_summary_from_labels(event_type, terminal_pattern)
        if not summary:
            if abrupt_observed and progressive_observed and precursor_date != "전조없음":
                summary = "전조 후 급격 종료"
            elif progressive_observed and precursor_date != "전조없음":
                summary = "전조 후 진행 악화"
            elif abrupt_observed:
                summary = "급작 발생" if precursor_date == "전조없음" else "급격 종료 관측"

    return {
        "급락 종결 관측": "있음" if abrupt_observed else "없음",
        "점진 저하 누적": "있음" if progressive_observed else "없음",
        "사건 종결 요약": summary,
    }


def has_precursor_signal(record: dict[str, object] | pd.Series) -> bool:
    if normalize_text(record.get("전조날짜")):
        return True
    for field in [
        "ews_warning_days",
        "pre_alarm_days",
        "pre_ews_days",
        "prefault_cond_ae_days",
        "prefault_cond_dtw_days",
        "prefault_cond_ews_days",
    ]:
        if as_int(record.get(field)) > 0:
            return True
    return False


def has_hard_fault_evidence(record: dict[str, object] | pd.Series) -> bool:
    return any(
        [
            as_int(record.get("final_days")) > 0,
            as_int(record.get("critical_days")) > 0,
            as_int(record.get("critical_confirmed_days")) > 0,
            is_truthy_scalar(record.get("final_fault")),
            is_truthy_scalar(record.get("critical_fault")),
            is_truthy_scalar(record.get("critical_confirmed")),
            is_truthy_scalar(record.get("대표final_fault")),
            is_truthy_scalar(record.get("대표critical_fault")),
            is_truthy_scalar(record.get("대표critical_confirmed")),
        ]
    )


def as_float(value: object) -> float | None:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(parsed):
        return None
    return parsed


def format_ratio(value: object, digits: int = 2) -> str:
    parsed = as_float(value)
    if parsed is None:
        return ""
    return f"{parsed:.{digits}f}"


def representative_signal_row(panel_core: pd.DataFrame) -> pd.Series:
    panel_df = panel_core.sort_values("date").copy()
    if panel_df.empty:
        return pd.Series(dtype=object)
    subtype_mask = panel_df.get("anom_subtype", pd.Series(dtype=object)).astype(str).str.contains(
        "degradation|fault_like|shadow_like|critical|confirmed_fault",
        case=False,
        na=False,
    )
    signal_mask = (
        truthy_mask(panel_df["final_fault"])
        | truthy_mask(panel_df["critical_fault"])
        | truthy_mask(panel_df["fault_like_day"])
        | truthy_mask(panel_df.get("event_A", pd.Series(False, index=panel_df.index)))
        | subtype_mask
    )
    focus_df = panel_df.loc[signal_mask].copy()
    if focus_df.empty:
        focus_df = panel_df.copy()
    if "mid_ratio" in focus_df.columns and focus_df["mid_ratio"].notna().any():
        return focus_df.sort_values(["mid_ratio", "date"], ascending=[True, True]).iloc[0]
    if "dtw_dist" in focus_df.columns and focus_df["dtw_dist"].notna().any():
        return focus_df.sort_values(["dtw_dist", "date"], ascending=[False, True]).iloc[0]
    return focus_df.iloc[0]


def signal_grade_explainer(evidence_row: dict[str, object]) -> str:
    text = normalize_text(evidence_row.get("운영해석등급_ko"))
    final_days = int(evidence_row.get("final_days", 0) or 0)
    critical_days = int(evidence_row.get("critical_days", 0) or 0)
    critical_confirmed_days = int(evidence_row.get("critical_confirmed_days", 0) or 0)
    ews_warning_days = int(evidence_row.get("ews_warning_days", 0) or 0)
    pre_alarm_days = int(evidence_row.get("pre_alarm_days", 0) or 0)
    pre_ews_days = int(evidence_row.get("pre_ews_days", 0) or 0)
    prefault_cond_ae_days = int(evidence_row.get("prefault_cond_ae_days", 0) or 0)
    prefault_cond_dtw_days = int(evidence_row.get("prefault_cond_dtw_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))
    if text == "확정":
        signal_labels: list[str] = []
        if final_days > 0:
            signal_labels.append("최종 고장 신호")
        if critical_confirmed_days > 0:
            signal_labels.append("강한 고장 신호 확정")
        elif critical_days > 0:
            signal_labels.append("강한 고장 신호")
        if "vdrop" in critical_sources:
            signal_labels.append("vdrop 전기 신호")
        signal_summary = " / ".join(signal_labels) if signal_labels else "확정 신호"
        return (
            f"다음 확정 신호가 관측돼 확정({final_days + critical_days + critical_confirmed_days}일): "
            f"{signal_summary}. 원인명은 후보 단계"
        )
    if text == "고위험 관찰":
        return (
            f"EWS({ews_warning_days}일)·pre_alarm({pre_alarm_days}일)·pre_ews({pre_ews_days}일)"
            f"와 AE/DTW 전조 조건(ae={prefault_cond_ae_days}, dtw={prefault_cond_dtw_days})이 누적돼 강한 이상징후로 분류"
        )
    if text == "관찰":
        return "약한 전조 신호만 보여 계속 관찰이 필요한 상태로 분류"
    if normalize_text(evidence_row.get("패널고장여부_ko")) == "고장":
        return "고정 결과표 기준 fault. 원인명은 후보 단계"
    return ""


def pattern_explainer(
    evidence_row: dict[str, object], *, soften_hard_language: bool = False
) -> str:
    mid_v_ratio = as_float(evidence_row.get("대표mid_v_ratio"))
    mid_i_ratio = as_float(evidence_row.get("대표mid_i_ratio"))
    mid_ratio = as_float(evidence_row.get("대표mid_ratio"))
    recon_error = as_float(evidence_row.get("대표recon_error"))
    dtw_dist = as_float(evidence_row.get("대표dtw_dist"))
    hs_score = as_float(evidence_row.get("대표hs_score"))
    critical_source = normalize_text(evidence_row.get("대표critical_source"))
    anom_subtype = normalize_text(evidence_row.get("대표anom_subtype"))
    final_flag = normalize_text(evidence_row.get("대표final_fault")) == "True"
    critical_flag = normalize_text(evidence_row.get("대표critical_fault")) == "True"
    event_flag = normalize_text(evidence_row.get("대표event_A")) == "True"

    reasons: list[str] = []
    if "vdrop" in critical_source:
        if soften_hard_language:
            reasons.append("상대 전압 이탈 징후가 반복 관측됨")
        else:
            reasons.append("전압강하형 전기 신호가 직접 관측됨")
    if mid_v_ratio is not None and mid_i_ratio is not None:
        if mid_v_ratio >= 0.9 and mid_i_ratio <= 0.4:
            if soften_hard_language:
                reasons.append(
                    f"전압 대비 전류 저하 징후가 나타남(mid_v={mid_v_ratio:.2f}, mid_i={mid_i_ratio:.2f})"
                )
            else:
                reasons.append(
                    f"전압은 비교적 유지되지만 전류가 크게 낮아짐(mid_v={mid_v_ratio:.2f}, mid_i={mid_i_ratio:.2f})"
                )
        elif mid_v_ratio <= 0.8 and mid_i_ratio <= 0.8:
            if soften_hard_language:
                reasons.append(
                    f"전압과 전류가 함께 낮아지는 징후가 이어짐(mid_v={mid_v_ratio:.2f}, mid_i={mid_i_ratio:.2f})"
                )
            else:
                reasons.append(
                    f"전압과 전류가 함께 낮아짐(mid_v={mid_v_ratio:.2f}, mid_i={mid_i_ratio:.2f})"
                )
        elif mid_i_ratio <= 0.4:
            if soften_hard_language:
                reasons.append(f"전류 저하 징후가 두드러짐(mid_i={mid_i_ratio:.2f})")
            else:
                reasons.append(f"전류가 크게 낮아진 패턴(mid_i={mid_i_ratio:.2f})")
    if mid_ratio is not None:
        if mid_ratio <= 0.1:
            reasons.append(f"중간 출력이 거의 0에 가까움(mid_ratio={mid_ratio:.2f})")
        elif mid_ratio <= 0.5:
            reasons.append(f"중간 출력이 뚜렷하게 낮아짐(mid_ratio={mid_ratio:.2f})")
    if final_flag:
        reasons.append("급락 종결 패턴이 직접 관측됨")
    elif critical_flag:
        reasons.append("critical fault 신호가 직접 나타남")
    elif event_flag:
        reasons.append("이상 이벤트(event_A)가 반복적으로 나타남")
    if "degradation" in anom_subtype:
        reasons.append("degradation subtype이 반복돼 점진적 저하 경향이 보임")
    if recon_error is not None and recon_error >= 0.05:
        reasons.append(f"정상 곡선 대비 복원 오차가 큼(recon={recon_error:.3f})")
    if dtw_dist is not None and dtw_dist >= 20:
        reasons.append(f"기준 곡선과 형태 차이가 큼(dtw={dtw_dist:.1f})")
    if hs_score is not None and hs_score >= 0.3:
        reasons.append(f"시계열 흔들림이 큼(hs={hs_score:.3f})")
    if not reasons:
        reasons.append("대표 관측일의 곡선/출력 변화가 정상 패턴과 다르게 나타남")
    return " / ".join(reasons[:3])


def to_user_preview_schema(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=USER_PREVIEW_OUTPUT_COLS)

    def pick_text(row: pd.Series, *columns: str) -> str:
        for column in columns:
            if column in row.index:
                text = normalize_text(row.get(column))
                if text:
                    return text
        return ""

    def pick_algorithm_source(row: pd.Series) -> str:
        source = pick_text(
            row,
            "기존 알고리즘 source",
            "커널로그 기존 알고리즘 판정",
            "커널로그 기존 알고리즘",
            "critical_source",
        )
        if not source:
            source = MAIL_BUCKET_ALGORITHM_MAP.get(
                (normalize_text(row.get("site")), normalize_text(row.get("panel_id"))),
                "",
            )
        return display_existing_algorithm_source(source)

    rows: list[dict[str, str]] = []
    for _, row in df.fillna("").iterrows():
        event_fields = event_display_fields(row)
        rows.append(
            {
                "site": normalize_text(row.get("site")),
                "panel_id": normalize_text(row.get("panel_id")),
                "전조날짜": display_preview_precursor_date(row.get("전조날짜")),
                "고장 기준일": pick_text(row, "고장 기준일", "고장날짜", "신호 기준일"),
                "운영 판정": display_signal_grade(row),
                **event_fields,
                "상위 해석 후보": pick_text(
                    row,
                    "상위 해석 후보",
                    "원인 추정",
                    "알고리즘 해석 원인",
                    "원인",
                    "1순위_의심원인_ko",
                ),
                "기존 알고리즘 source": pick_algorithm_source(row),
            }
        )
    return pd.DataFrame(rows).reindex(columns=USER_PREVIEW_OUTPUT_COLS)


def to_signal_preview_schema(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=SIGNAL_PREVIEW_OUTPUT_COLS)

    def pick_text(row: pd.Series, *columns: str) -> str:
        for column in columns:
            if column in row.index:
                text = normalize_text(row.get(column))
                if text:
                    return text
        return ""

    def pick_algorithm_source(row: pd.Series) -> str:
        source = pick_text(
            row,
            "기존 알고리즘 source",
            "커널로그 기존 알고리즘 판정",
            "커널로그 기존 알고리즘",
            "critical_source",
        )
        if not source:
            source = MAIL_BUCKET_ALGORITHM_MAP.get(
                (normalize_text(row.get("site")), normalize_text(row.get("panel_id"))),
                "",
            )
        return display_existing_algorithm_source(source)

    rows: list[dict[str, str]] = []
    for _, row in df.fillna("").iterrows():
        event_fields = event_display_fields(row)
        rows.append(
            {
                "site": normalize_text(row.get("site")),
                "panel_id": normalize_text(row.get("panel_id")),
                "전조날짜": display_preview_precursor_date(row.get("전조날짜")),
                "신호 기준일": pick_text(row, "신호 기준일", "고장날짜", "고장 기준일"),
                "운영 판정": display_signal_grade(row),
                **event_fields,
                "상위 해석 후보": pick_text(
                    row,
                    "상위 해석 후보",
                    "원인 추정",
                    "알고리즘 해석 원인",
                    "원인",
                    "1순위_의심원인_ko",
                ),
                "기존 알고리즘 source": pick_algorithm_source(row),
            }
        )
    return pd.DataFrame(rows).reindex(columns=SIGNAL_PREVIEW_OUTPUT_COLS)


def load_raw_only_common_module():
    package = package_root()
    if str(package) not in sys.path:
        sys.path.insert(0, str(package))
    from research.prognostics import runtime_rawonly_chain_common_v1 as raw_only_common_mod

    return raw_only_common_mod


def load_runtime_heuristic_module():
    package = package_root()
    if str(package) not in sys.path:
        sys.path.insert(0, str(package))
    from research.prognostics import (
        build_panel_day_engine_runtime_heuristic_v1 as runtime_heuristic_mod,
    )

    return runtime_heuristic_mod


def packaged_live_chain_support() -> dict[str, object]:
    required_scripts = [
        "build_panel_day_engine_bootstrap_verdict_v1.py",
        "build_panel_day_engine_fault_panel_event_audit_v1.py",
        "build_panel_day_engine_panel_multiaxis_verdict_v1.py",
        "build_panel_day_engine_gpvs_evidence_pack_v1.py",
        "build_panel_day_engine_cause_candidate_heuristics_v1.py",
    ]
    required_share_inputs = [
        "panel_day_engine_operator_workflow_default_v1.csv",
        "panel_day_engine_abrupt6_symptom_map_v1.csv",
        "panel_day_engine_kernellog_project_mapping_v1.csv",
        "panel_day_engine_gpv7_perf_summary_v1.csv",
        "panel_day_engine_project_final_decision_pack_v1.csv",
        "panel_day_engine_precursor_onset_truth_v1.csv",
        "panel_day_engine_non_precursor_performance_cases_v1.csv",
        "panel_day_engine_common_cause_descriptive_retrofit_cases_v1.csv",
        "panel_day_engine_gpvs_panel_attach_inventory_v1.csv",
        "panel_day_engine_gpvs_panel_attach_feasibility_v1.csv",
        "panel_day_engine_gpvs_panel_attach_candidates_v1.csv",
        "panel_day_engine_precursor_abrupt_consistency_cases_v1.csv",
        "panel_day_engine_precursor_abrupt_consistency_summary_v1.csv",
        "panel_day_engine_precursor_abrupt_consistency_recommendation_v1.csv",
        "panel_day_engine_c42997_1_1_forensic_summary_v1.csv",
        "panel_day_engine_fault_panel_event_audit_v1.csv",
        "panel_day_engine_detailed_fault_bridge_audit_v1.csv",
        "panel_day_engine_detailed_fault_bridge_summary_v1.csv",
        "panel_day_engine_gpvs_bytype_rebuild_summary_v1.csv",
        "panel_day_engine_gpvs_detailed_type_inference_audit_v1.csv",
        "panel_day_engine_gpvs_detailed_type_realpanel_sanity_v1.csv",
        "panel_day_engine_gpvs_mlpe_panel_agreement_v1.csv",
        "panel_day_engine_gpvs_canonical_dictionary_v1.csv",
        "panel_day_engine_gpvs_mlpe_fault_matching_table_v1.csv",
        "panel_day_engine_gpvs_mlpe_compatibility_summary_v1.csv",
        "panel_day_engine_gpvs_mlpe_fault_matching_summary_v1.csv",
        "panel_day_engine_gpvs_evidence_pack_v1.csv",
        "panel_day_engine_panel_multiaxis_verdict_v1.csv",
        "panel_date_reaudit_working.csv",
    ]
    missing_scripts = [name for name in required_scripts if not packaged_script_path(name).exists()]
    missing_share = [name for name in required_share_inputs if not (packaged_share_root() / name).exists()]
    supported = not missing_scripts and not missing_share
    return {
        "supported": supported,
        "required_scripts": required_scripts,
        "required_share_inputs": required_share_inputs,
        "missing_scripts": missing_scripts,
        "missing_share_inputs": missing_share,
        "note_ko": (
            "live chain은 package 내부에 복사된 bootstrap/audit/verdict/evidence/heuristic 스크립트와 "
            "frozen share 입력을 사용해 workspace-only로 재계산한다."
        ),
    }


def packaged_raw_only_chain_support() -> dict[str, object]:
    required_scripts = [
        "runtime_rawonly_chain_common_v1.py",
        "build_panel_day_engine_runtime_fault_event_audit_v1.py",
        "build_panel_day_engine_runtime_final_verdict_v1.py",
        "build_panel_day_engine_runtime_heuristic_v1.py",
    ]
    missing_scripts = [name for name in required_scripts if not packaged_script_path(name).exists()]
    return {
        "supported": not missing_scripts,
        "required_scripts": required_scripts,
        "missing_scripts": missing_scripts,
        "note_ko": (
            "raw-only chain은 package 내부에 복사된 runtime audit/verdict/heuristic 스크립트만 사용한다. "
            "frozen share truth/support asset은 참조하지 않는다."
        ),
    }


def normalize_core_digest_frame(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    ensure_columns(df, CORE_DIGEST_COLUMNS, source_name)
    digest_df = df.loc[:, CORE_DIGEST_COLUMNS].copy()
    digest_df["date"] = pd.to_datetime(digest_df["date"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
    for column in CORE_DIGEST_COLUMNS:
        if column == "date":
            continue
        digest_df[column] = digest_df[column].map(normalize_text)
    digest_df["panel_id"] = digest_df["panel_id"].astype(str)
    return digest_df.sort_values(["panel_id", "date"]).reset_index(drop=True)


def build_core_digest_payload(df: pd.DataFrame, source_name: str) -> dict[str, object]:
    digest_df = normalize_core_digest_frame(df, source_name)
    joined_rows = "\n".join(
        "|".join(normalize_text(value) for value in row)
        for row in digest_df.itertuples(index=False, name=None)
    )
    return {
        "columns": CORE_DIGEST_COLUMNS,
        "row_count": int(len(digest_df)),
        "digest_sha256": hashlib.sha256(joined_rows.encode("utf-8")).hexdigest(),
        "critical_source_counts": {
            key: int(value)
            for key, value in digest_df["critical_source"].value_counts(dropna=False).sort_index().items()
        },
        "anom_level_counts": {
            key: int(value)
            for key, value in digest_df["anom_level"].value_counts(dropna=False).sort_index().items()
        },
        "confirmed_fault_true_count": int(truthy_mask(digest_df["confirmed_fault"]).sum()),
        "critical_fault_true_count": int(truthy_mask(digest_df["critical_fault"]).sum()),
        "final_fault_true_count": int(truthy_mask(digest_df["final_fault"]).sum()),
    }


def load_core_baseline_digest() -> dict[str, object]:
    path = baseline_core_digest_path()
    if not path.exists():
        raise SystemExit(f"missing packaged core baseline digest: {path}")
    return json.loads(path.read_text(encoding="utf-8"))


def compare_single_site_digest(expected: dict[str, object], actual: dict[str, object]) -> list[str]:
    diffs: list[str] = []
    for key in [
        "row_count",
        "digest_sha256",
        "confirmed_fault_true_count",
        "critical_fault_true_count",
        "final_fault_true_count",
    ]:
        if expected.get(key) != actual.get(key):
            diffs.append(f"{key}: expected={expected.get(key)} actual={actual.get(key)}")
    if expected.get("columns") != actual.get("columns"):
        diffs.append("columns: expected reference columns differ from actual columns")
    if expected.get("critical_source_counts") != actual.get("critical_source_counts"):
        diffs.append("critical_source_counts: expected reference counts differ from actual counts")
    if expected.get("anom_level_counts") != actual.get("anom_level_counts"):
        diffs.append("anom_level_counts: expected reference counts differ from actual counts")
    return diffs


def load_panel_day_core_from_workspace(workspace_root: Path, site: str) -> pd.DataFrame:
    path = workspace_root / "data" / site / "out" / "panel_day_core.csv"
    if not path.exists():
        raise SystemExit(f"missing workspace panel_day_core: {path}")
    df = pd.read_csv(path, low_memory=False)
    ensure_columns(
        df,
        ["panel_id", "date", "final_fault", "critical_fault", "fault_like_day", "critical_source"],
        path.name,
    )
    df["panel_id"] = df["panel_id"].astype(str)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    return df


def representative_algorithm_fields(site: str, core_df: pd.DataFrame, panel_id: str) -> dict[str, str]:
    mapped = MAIL_BUCKET_ALGORITHM_MAP.get((normalize_text(site), normalize_text(panel_id)), "")
    if mapped:
        return {"커널로그 기존 알고리즘": mapped}
    panel_df = core_df.loc[core_df["panel_id"].eq(str(panel_id))].copy().sort_values("date")
    if panel_df.empty:
        return {"커널로그 기존 알고리즘": ""}

    final_days = panel_df.loc[truthy_mask(panel_df["final_fault"])]
    critical_days = panel_df.loc[truthy_mask(panel_df["critical_fault"])]
    fault_like_days = panel_df.loc[truthy_mask(panel_df["fault_like_day"])]

    if not final_days.empty:
        representative = final_days.iloc[0]
    elif not critical_days.empty:
        representative = critical_days.iloc[0]
    elif not fault_like_days.empty:
        representative = fault_like_days.iloc[0]
    else:
        representative = panel_df.iloc[-1]

    return {"커널로그 기존 알고리즘": normalize_text(representative.get("critical_source"))}


def build_live_fault_table(workspace_root: Path) -> pd.DataFrame:
    verdict_path = workspace_root / "_share" / "panel_day_engine_panel_multiaxis_verdict_v1.csv"
    heuristic_path = workspace_root / "_share" / "panel_day_engine_cause_candidate_heuristics_v1.csv"
    audit_path = workspace_root / "_share" / "panel_day_engine_fault_panel_event_audit_v1.csv"
    verdict_df = pd.read_csv(verdict_path, encoding="utf-8-sig", low_memory=False)
    heuristic_df = pd.read_csv(heuristic_path, encoding="utf-8-sig", low_memory=False)
    audit_df = pd.read_csv(audit_path, encoding="utf-8-sig", low_memory=False)
    ensure_columns(
        verdict_df,
        ["site", "panel_id", "패널고장여부_ko", "사건유형_ko", "최종고장양상_ko", "커널로그_원인군_ko"],
        verdict_path.name,
    )
    ensure_columns(
        heuristic_df,
        ["site", "panel_id", "원인후보_top1_ko", "원인후보_top2_ko", "원인후보_top3_ko"],
        heuristic_path.name,
    )
    ensure_columns(
        audit_df,
        [
            "site",
            "panel_id",
            "earliest_warning_date",
            "strict_trigger_date",
            "first_final_fault_date",
        ],
        audit_path.name,
    )

    heuristic_lookup = {
        row_key(row["site"], row["panel_id"]): row
        for row in heuristic_df.to_dict(orient="records")
    }
    audit_lookup = {
        row_key(row["site"], row["panel_id"]): row
        for row in audit_df.to_dict(orient="records")
    }
    rows: list[dict[str, str]] = []
    for row in verdict_df.loc[verdict_df["패널고장여부_ko"].map(normalize_text).eq("고장")].to_dict(orient="records"):
        key = row_key(row["site"], row["panel_id"])
        heuristic_row = heuristic_lookup.get(key)
        if heuristic_row is None:
            raise SystemExit(f"missing heuristic row for fault panel: {key}")
        audit_row = audit_lookup.get(key, {})
        rows.append(
            {
                "site": normalize_text(row["site"]),
                "panel_id": normalize_text(row["panel_id"]),
                "패널고장여부_ko": normalize_text(row["패널고장여부_ko"]),
                "사건유형_ko": normalize_text(row["사건유형_ko"]),
                "최종고장양상_ko": normalize_text(row["최종고장양상_ko"]),
                "커널로그_원인군_ko": normalize_text(row["커널로그_원인군_ko"]),
                "1순위_의심원인_ko": display_heuristic_name(heuristic_row["원인후보_top1_ko"]),
                "2순위_의심원인_ko": display_heuristic_name(heuristic_row["원인후보_top2_ko"]),
                "3순위_의심원인_ko": display_heuristic_name(heuristic_row["원인후보_top3_ko"]),
                "전조날짜": choose_display_precursor_date(
                    event_type_ko=row.get("사건유형_ko"),
                    interpreted_onset_date=row.get("사건해석상전조시작일"),
                    first_warning_date=audit_row.get("earliest_warning_date"),
                ),
                "고장날짜": choose_display_fault_date(
                    fault_date=row.get("세부fault_기준일"),
                    strict_trigger_date=audit_row.get("strict_trigger_date"),
                    first_final_fault_date=audit_row.get("first_final_fault_date"),
                ),
            }
        )
    return (
        pd.DataFrame(rows)
        .reindex(columns=LIVE_FAULT_OUTPUT_COLS)
        .sort_values(["site", "panel_id"], ascending=[True, True])
        .reset_index(drop=True)
    )


def build_live_fault_preview(workspace_root: Path, fault_df: pd.DataFrame) -> pd.DataFrame:
    per_site_core = {
        site: load_panel_day_core_from_workspace(workspace_root, site)
        for site in sorted(fault_df["site"].astype(str).unique())
    }
    rows: list[dict[str, str]] = []
    for _, row in fault_df.iterrows():
        site = normalize_text(row["site"])
        panel_id = normalize_text(row["panel_id"])
        rows.append(
            {
                "site": site,
                "panel_id": panel_id,
                "패널고장여부_ko": normalize_text(row["패널고장여부_ko"]),
                "사건유형_ko": normalize_text(row["사건유형_ko"]),
                "최종고장양상_ko": normalize_text(row["최종고장양상_ko"]),
                "전조날짜": normalize_text(row.get("전조날짜")),
                "고장날짜": normalize_text(row.get("고장날짜")),
                "라벨된 fault": normalize_text(row["커널로그_원인군_ko"]),
                "1순위_의심원인_ko": normalize_text(row["1순위_의심원인_ko"]),
                "2순위_의심원인_ko": normalize_text(row["2순위_의심원인_ko"]),
                "3순위_의심원인_ko": normalize_text(row["3순위_의심원인_ko"]),
                **representative_algorithm_fields(site, per_site_core[site], panel_id),
            }
        )
    return pd.DataFrame(rows).reindex(columns=LIVE_PREVIEW_OUTPUT_COLS)


def compare_live_fault_to_fixed(live_fault_df: pd.DataFrame) -> dict[str, object]:
    fixed_path = fixed_fault6_table_path()
    if not fixed_path.exists():
        return {
            "fixed_reference_available": False,
            "exact_match": False,
            "diff_columns": [],
        }
    fixed_df = pd.read_csv(fixed_path, encoding="utf-8-sig", low_memory=False).sort_values(["site", "panel_id"]).reset_index(drop=True)
    live_df = live_fault_df.sort_values(["site", "panel_id"]).reset_index(drop=True)
    diff_columns: list[str] = []
    if len(fixed_df) != len(live_df):
        diff_columns.append("__row_count__")
    else:
        for column in LIVE_FAULT_OUTPUT_COLS:
            if column not in LIVE_FAULT_COMPARE_COLS:
                continue
            left = fixed_df[column].fillna("").astype(str)
            right = live_df[column].fillna("").astype(str)
            if not left.equals(right):
                diff_columns.append(column)
    return {
        "fixed_reference_available": True,
        "exact_match": not diff_columns,
        "diff_columns": diff_columns,
        "fixed_row_count": int(len(fixed_df)),
        "live_row_count": int(len(live_df)),
    }


def publish_live_chain_outputs(output_root: Path, result_dir: Path, summary_path: Path) -> dict[str, str]:
    root_result_dir = output_root / "result"
    root_result_dir.mkdir(parents=True, exist_ok=True)

    mapping = {
        result_dir / "fault_panel_result_live_v1.csv": root_result_dir / ROOT_LIVE_FAULT_NAME,
        result_dir / "fault_panel_result_live_preview_v1.csv": root_result_dir / ROOT_LIVE_PREVIEW_NAME,
    }
    published: dict[str, str] = {}
    for source, target in mapping.items():
        if not source.exists():
            raise SystemExit(f"missing live chain output for publish step: {source}")
        shutil.copy2(source, target)
        published[target.name] = str(target)
    return published


def publish_raw_only_chain_outputs(output_root: Path, result_dir: Path) -> dict[str, str]:
    root_result_dir = output_root / "result"
    root_result_dir.mkdir(parents=True, exist_ok=True)
    mapping = {
        result_dir / "fault_panel_result_raw_only_v1.csv": root_result_dir / ROOT_RAWONLY_FAULT_NAME,
        result_dir / "fault_panel_result_raw_only_preview_v1.csv": root_result_dir / ROOT_RAWONLY_PREVIEW_NAME,
    }
    published: dict[str, str] = {}
    for source, target in mapping.items():
        if not source.exists():
            raise SystemExit(f"missing raw-only chain output for publish step: {source}")
        shutil.copy2(source, target)
        published[target.name] = str(target)
    return published


def build_strict_raw_only_current_outputs(
    raw_only_chain_result: dict[str, object],
    evidence_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, object]]:
    candidate_fault_path = Path(
        str(raw_only_chain_result.get("generated_outputs", {}).get("fault_panel_result_raw_only_v1", ""))
    )
    candidate_preview_path = Path(
        str(raw_only_chain_result.get("generated_outputs", {}).get("fault_panel_result_raw_only_preview_v1", ""))
    )
    if not candidate_fault_path.exists() or not candidate_preview_path.exists():
        raise SystemExit("missing candidate raw-only outputs for strict current publish")

    candidate_fault_df = pd.read_csv(candidate_fault_path, encoding="utf-8-sig", low_memory=False)
    candidate_preview_df = pd.read_csv(candidate_preview_path, encoding="utf-8-sig", low_memory=False)
    strict_keys = {
        row_key(row["site"], row["panel_id"])
        for row in evidence_df.to_dict(orient="records")
        if normalize_text(row.get("운영해석등급_ko")) in RAW_ONLY_STRICT_CURRENT_GRADES
    }
    if strict_keys:
        strict_fault_df = candidate_fault_df.loc[
            candidate_fault_df.apply(lambda row: row_key(row["site"], row["panel_id"]) in strict_keys, axis=1)
        ].copy()
        strict_preview_df = candidate_preview_df.loc[
            candidate_preview_df.apply(lambda row: row_key(row["site"], row["panel_id"]) in strict_keys, axis=1)
        ].copy()
    else:
        strict_fault_df = candidate_fault_df.iloc[0:0].copy()
        strict_preview_df = candidate_preview_df.iloc[0:0].copy()

    strict_fault_df = strict_fault_df.sort_values(["site", "panel_id"]).reset_index(drop=True)
    strict_preview_df = strict_preview_df.sort_values(["site", "panel_id"]).reset_index(drop=True)
    date_lookup = {
        row_key(row["site"], row["panel_id"]): {
            "전조날짜": normalize_text(row.get("전조날짜")),
            "고장날짜": normalize_text(row.get("고장날짜")),
        }
        for row in evidence_df.to_dict(orient="records")
    }
    for df in [strict_fault_df, strict_preview_df]:
        if df.empty:
            continue
        df["전조날짜"] = df.apply(
            lambda row: date_lookup.get(row_key(row["site"], row["panel_id"]), {}).get("전조날짜", ""),
            axis=1,
        )
        df["고장날짜"] = df.apply(
            lambda row: date_lookup.get(row_key(row["site"], row["panel_id"]), {}).get("고장날짜", ""),
            axis=1,
        )
        ordered_cols = [column for column in df.columns if column not in {"전조날짜", "고장날짜"}]
        insert_at = ordered_cols.index("최종고장양상_ko") + 1 if "최종고장양상_ko" in ordered_cols else len(ordered_cols)
        ordered_cols[insert_at:insert_at] = ["전조날짜", "고장날짜"]
        df = df.reindex(columns=ordered_cols)
        if df is strict_fault_df:
            strict_fault_df = df
        else:
            strict_preview_df = df
    meta = {
        "publish_policy_ko": "raw_only current는 운영해석등급_ko=확정 strict subset만 노출",
        "strict_grade_csv": ",".join(sorted(RAW_ONLY_STRICT_CURRENT_GRADES)),
        "candidate_row_count": int(len(candidate_fault_df)),
        "published_current_row_count": int(len(strict_fault_df)),
        "dropped_candidate_row_count": int(len(candidate_fault_df) - len(strict_fault_df)),
    }
    return strict_fault_df, strict_preview_df, meta


def publish_raw_only_current_outputs(
    output_root: Path,
    strict_fault_df: pd.DataFrame,
    strict_preview_df: pd.DataFrame,
) -> dict[str, str]:
    root_result_dir = output_root / "result"
    root_result_dir.mkdir(parents=True, exist_ok=True)
    fault_path = root_result_dir / ROOT_RAWONLY_FAULT_NAME
    preview_path = root_result_dir / ROOT_RAWONLY_PREVIEW_NAME
    strict_fault_df.to_csv(fault_path, index=False, encoding="utf-8-sig")
    strict_preview_df.to_csv(preview_path, index=False, encoding="utf-8-sig")
    return {
        ROOT_RAWONLY_FAULT_NAME: str(fault_path),
        ROOT_RAWONLY_PREVIEW_NAME: str(preview_path),
    }


def markdown_table_from_df(df: pd.DataFrame) -> str:
    if df.empty:
        return "_empty_"
    safe_df = df.fillna("").astype(str)
    headers = safe_df.columns.tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in safe_df.itertuples(index=False, name=None):
        lines.append("| " + " | ".join(str(value) for value in row) + " |")
    return "\n".join(lines)


def truncate_report_df(df: pd.DataFrame, limit: int = 20) -> pd.DataFrame:
    if df.empty or len(df) <= limit:
        return df
    return df.head(limit).copy()


def build_live_report_markdown(
    sites: list[str],
    baseline_comparison: dict[str, object],
    compare: dict[str, object],
    published_outputs: dict[str, str],
    live_preview_df: pd.DataFrame,
) -> str:
    site_lines = "\n".join(f"- `{site}`" for site in sites)
    baseline_site_lines = []
    for site in sites:
        site_entry = baseline_comparison.get("sites", {}).get(site, {})
        baseline_site_lines.append(
            f"- `{site}`: `match={site_entry.get('match')}`"
        )
    baseline_block = "\n".join(baseline_site_lines)
    output_lines = "\n".join(
        f"- `{name}`: `{path}`" for name, path in sorted(published_outputs.items())
    )
    return (
        "# fault_panel_result_current_report_v1\n\n"
        "## 목적\n"
        "현재 runtime 실행에서 운영자가 바로 확인할 `운영 공식 current` 결과를 한 곳에 모아 보여준다.\n\n"
        "## 실행 대상 site\n"
        f"{site_lines}\n\n"
        "## baseline 입력 비교\n"
        f"- `all_sites_match`: `{baseline_comparison.get('all_sites_match')}`\n"
        f"{baseline_block}\n\n"
        "## live chain 상태\n"
        f"- `fixed_fault_reference_exact_match`: `{compare.get('exact_match')}`\n"
        f"- `baseline_input_all_sites_match`: `{compare.get('baseline_input_all_sites_match')}`\n"
        f"- `diff_columns`: `{compare.get('diff_columns', [])}`\n\n"
        "## 읽는 법\n"
        "- 이 report는 `official current` 설명용 문서다.\n"
        "- `fault_panel_result_current_preview_v1.csv`와 함께 현재 운영 공식 결과를 먼저 읽는 기본 문서다.\n"
        "- `raw-only` 보조표나 analyst artifact를 대신하지 않는다.\n"
        "- `fault_panel_result_master_report_v1.md`는 artifact 안내와 fallback 설명용 문서이며, 이 report를 대체하지 않는다.\n\n"
        "## 주요 산출물\n"
        f"{output_lines}\n\n"
        "## 현재 preview 표\n"
        f"{markdown_table_from_df(live_preview_df)}\n"
    )


def build_raw_only_report_markdown(
    sites: list[str],
    compare: dict[str, object],
    published_outputs: dict[str, str],
    live_preview_df: pd.DataFrame,
    publish_meta: dict[str, object] | None = None,
) -> str:
    site_lines = "\n".join(f"- `{site}`" for site in sites)
    output_lines = "\n".join(
        f"- `{name}`: `{path}`" for name, path in sorted(published_outputs.items())
    )
    publish_meta = publish_meta or {}
    return (
        "# fault_panel_result_raw_only_current_report_v1\n\n"
        "## 목적\n"
        "raw-only algorithm candidate chain 중 운영 strict current로 승격된 현재 결과를 `분석용/운영 보조표`로 확인한다.\n\n"
        "## 실행 대상 site\n"
        f"{site_lines}\n\n"
        "## raw-only vs fixed reference 비교\n"
        f"- `status_ko`: `{compare.get('status_ko')}`\n"
        f"- `reference_available`: `{compare.get('reference_available')}`\n"
        f"- `row_key_match`: `{compare.get('row_key_match')}`\n"
        f"- `decision_columns_match`: `{compare.get('decision_columns_match')}`\n"
        f"- `overlap_decision_columns_match`: `{compare.get('overlap_decision_columns_match')}`\n"
        f"- `exact_match`: `{compare.get('exact_match')}`\n"
        f"- `reference_row_count`: `{compare.get('reference_row_count')}`\n"
        f"- `candidate_row_count`: `{compare.get('candidate_row_count')}`\n"
        f"- `matched_row_key_count`: `{compare.get('matched_row_key_count')}`\n"
        f"- `diff_columns`: `{compare.get('diff_columns', [])}`\n\n"
        f"- `overlap_diff_columns`: `{compare.get('overlap_diff_columns', [])}`\n\n"
        "## current 출력 정책\n"
        f"- `publish_policy_ko`: `{publish_meta.get('publish_policy_ko', '')}`\n"
        f"- `strict_grade_csv`: `{publish_meta.get('strict_grade_csv', '')}`\n"
        f"- `published_current_row_count`: `{publish_meta.get('published_current_row_count', '')}`\n"
        f"- `candidate_row_count`: `{publish_meta.get('candidate_row_count', '')}`\n"
        f"- `dropped_candidate_row_count`: `{publish_meta.get('dropped_candidate_row_count', '')}`\n\n"
        "## 주의\n"
        "- `커널로그_원인군_ko` 컬럼명은 유지하지만, 이 report에서는 raw-only algorithm-derived family 의미다.\n"
        "- 이 chain은 frozen truth/support asset을 참조하지 않는다.\n\n"
        "- 이 report는 `official current report`가 아니며, 운영 공식 결과를 대체하지 않는다.\n"
        "- 운영자 기본 진입점은 `fault_panel_result_current_*` 계열이고, 이 report는 analyst/support 확인용이다.\n\n"
        "- preview 표의 `사건 종결 요약`은 관측 플래그를 먼저 본 뒤, 확정 row에서만 채워지는 요약이다.\n\n"
        "- `result/raw_only_chain/*`에는 전체 candidate가 남고, `result/fault_panel_result_raw_only_current_*`는 strict current subset만 노출한다.\n\n"
        "## 주요 산출물\n"
        f"{output_lines}\n\n"
        "## 현재 preview 표\n"
        f"{markdown_table_from_df(truncate_report_df(live_preview_df))}\n"
    )


def build_master_report_markdown(
    sites: list[str],
    baseline_comparison: dict[str, object],
    live_chain_result: dict[str, object],
    raw_only_chain_result: dict[str, object],
    live_preview_df: pd.DataFrame,
    raw_only_preview_df: pd.DataFrame,
    precursor_report_df: pd.DataFrame | None = None,
    fault_signal_report_df: pd.DataFrame | None = None,
    detailed_report_path: Path | None = None,
    precursor_report_path: Path | None = None,
    fault_signal_report_path: Path | None = None,
) -> str:
    site_lines = "\n".join(f"- `{site}`" for site in sites)
    baseline_site_lines = []
    for site in sites:
        site_entry = baseline_comparison.get("sites", {}).get(site, {})
        baseline_site_lines.append(f"- `{site}`: `match={site_entry.get('match')}`")
    baseline_block = "\n".join(baseline_site_lines)
    live_compare = live_chain_result.get("fixed_fault_reference_compare", {})
    raw_only_compare = raw_only_chain_result.get("fixed_fault_reference_compare", {})
    primary_output_lines = []
    analyst_output_lines = []
    for name, path in sorted(live_chain_result.get("published_outputs", {}).items()):
        primary_output_lines.append(f"- `live::{name}`: `{path}`")
    for name, path in sorted(raw_only_chain_result.get("published_outputs", {}).items()):
        analyst_output_lines.append(f"- `raw_only::{name}`: `{path}`")
    primary_output_block = "\n".join(primary_output_lines) if primary_output_lines else "_none_"
    analyst_output_block = "\n".join(analyst_output_lines) if analyst_output_lines else "_none_"
    precursor_report_df = precursor_report_df if precursor_report_df is not None else pd.DataFrame()
    fault_signal_report_df = fault_signal_report_df if fault_signal_report_df is not None else pd.DataFrame()
    precursor_keys = set(
        zip(
            precursor_report_df.get("site", pd.Series(dtype=object)).astype(str),
            precursor_report_df.get("panel_id", pd.Series(dtype=object)).astype(str),
        )
    )
    fault_signal_keys = set(
        zip(
            fault_signal_report_df.get("site", pd.Series(dtype=object)).astype(str),
            fault_signal_report_df.get("panel_id", pd.Series(dtype=object)).astype(str),
        )
    )
    overlap_row_count = len(precursor_keys & fault_signal_keys)
    fault_signal_subgroup_summary = pd.DataFrame(
        columns=["site", "group root", "subgroup base", "row_count"]
    )
    fault_signal_cluster_summary = pd.DataFrame(
        columns=[
            "site",
            "group root",
            "subgroup base",
            "subgroup cluster",
            "row_count",
            "min_signal_date",
            "max_signal_date",
        ]
    )
    if fault_signal_report_df is not None and not fault_signal_report_df.empty:
        working = fault_signal_report_df.copy()
        working["group root"] = working["group root"].map(normalize_text)
        working["subgroup base"] = working["subgroup base"].map(normalize_text)
        working["subgroup cluster"] = working["subgroup cluster"].map(normalize_text)
        working["신호 기준일_dt"] = pd.to_datetime(working["신호 기준일"], errors="coerce")
        working = working.loc[working["subgroup base"].ne("")].copy()
        if not working.empty:
            fault_signal_subgroup_summary = (
                working.groupby(["site", "group root", "subgroup base"], dropna=False)
                .size()
                .rename("row_count")
                .reset_index()
                .sort_values(
                    ["row_count", "site", "group root", "subgroup base"],
                    ascending=[False, True, True, True],
                )
                .reset_index(drop=True)
            )
            fault_signal_cluster_summary = (
                working.groupby(
                    ["site", "group root", "subgroup base", "subgroup cluster"], dropna=False
                )
                .agg(
                    row_count=("panel_id", "size"),
                    min_signal_date=("신호 기준일_dt", "min"),
                    max_signal_date=("신호 기준일_dt", "max"),
                )
                .reset_index()
                .sort_values(
                    ["row_count", "site", "group root", "subgroup base", "subgroup cluster"],
                    ascending=[False, True, True, True, True],
                )
                .reset_index(drop=True)
            )
            for column in ["min_signal_date", "max_signal_date"]:
                fault_signal_cluster_summary[column] = pd.to_datetime(
                    fault_signal_cluster_summary[column], errors="coerce"
                ).dt.strftime("%Y-%m-%d")
    fault_signal_unique_group_root_count = (
        int(len(fault_signal_subgroup_summary[["site", "group root"]].drop_duplicates()))
        if not fault_signal_subgroup_summary.empty
        else 0
    )
    fault_signal_unique_subgroup_base_count = (
        int(len(fault_signal_subgroup_summary[["site", "subgroup base"]].drop_duplicates()))
        if not fault_signal_subgroup_summary.empty
        else 0
    )
    fault_signal_unique_subgroup_cluster_count = (
        int(len(fault_signal_cluster_summary[["site", "subgroup cluster"]].drop_duplicates()))
        if not fault_signal_cluster_summary.empty
        else 0
    )
    fault_signal_top_subgroup_block = (
        markdown_table_from_df(fault_signal_subgroup_summary.head(10))
        if not fault_signal_subgroup_summary.empty
        else "_none_"
    )
    fault_signal_top_cluster_block = (
        markdown_table_from_df(fault_signal_cluster_summary.head(10))
        if not fault_signal_cluster_summary.empty
        else "_none_"
    )
    detailed_report_block = (
        f"- `fault_panel_result_detailed_report_v1.xlsx`: `{detailed_report_path}`\n\n"
        if detailed_report_path is not None
        else ""
    )
    precursor_report_block = (
        f"- `fault_panel_result_precursor_report_v1.csv`: `{precursor_report_path}`\n\n"
        if precursor_report_path is not None
        else ""
    )
    fault_signal_report_block = (
        f"- `{ROOT_FAULT_SIGNAL_REPORT_NAME}`: `{fault_signal_report_path}`\n\n"
        if fault_signal_report_path is not None
        else ""
    )
    return (
        "# fault_panel_result_master_report_v1\n\n"
        "## 목적\n"
        "frozen-support live chain과 raw-only algorithm candidate chain을 비교하고, 어떤 artifact를 어떤 순서로 읽을지 안내한다.\n\n"
        "## 실행 대상 site\n"
        f"{site_lines}\n\n"
        "## baseline 입력 비교\n"
        f"- `all_sites_match`: `{baseline_comparison.get('all_sites_match')}`\n"
        f"{baseline_block}\n\n"
        "## frozen-support live chain 요약\n"
        f"- `status_ko`: `{live_chain_result.get('status_ko')}`\n"
        f"- `fixed_fault_reference_exact_match`: `{live_compare.get('exact_match')}`\n"
        f"- `baseline_input_all_sites_match`: `{live_compare.get('baseline_input_all_sites_match')}`\n"
        f"- `diff_columns`: `{live_compare.get('diff_columns', [])}`\n\n"
        "## raw-only algorithm candidate chain 요약\n"
        f"- `status_ko`: `{raw_only_compare.get('status_ko')}`\n"
        f"- `reference_available`: `{raw_only_compare.get('reference_available')}`\n"
        f"- `overlap_decision_columns_match`: `{raw_only_compare.get('overlap_decision_columns_match')}`\n"
        f"- `reference_row_count`: `{raw_only_compare.get('reference_row_count')}`\n"
        f"- `candidate_row_count`: `{raw_only_compare.get('candidate_row_count')}`\n"
        f"- `published_current_row_count`: `{raw_only_chain_result.get('publish_meta', {}).get('published_current_row_count', '')}`\n"
        f"- `matched_row_key_count`: `{raw_only_compare.get('matched_row_key_count')}`\n"
        f"- `overlap_diff_columns`: `{raw_only_compare.get('overlap_diff_columns', [])}`\n\n"
        "## report split 요약\n"
        f"- `precursor_candidate_row_count`: `{len(precursor_report_df)}`\n"
        f"- `raw_only_fault_signal_row_count`: `{len(fault_signal_report_df)}`\n"
        f"- `raw_only_fault_signal_unique_group_root_count`: `{fault_signal_unique_group_root_count}`\n"
        f"- `raw_only_fault_signal_unique_subgroup_base_count`: `{fault_signal_unique_subgroup_base_count}`\n"
        f"- `raw_only_fault_signal_unique_subgroup_cluster_count`: `{fault_signal_unique_subgroup_cluster_count}`\n"
        f"- `report_row_overlap_count`: `{overlap_row_count}`\n\n"
        "## 먼저 보는 법\n"
        "- `fault_panel_result_current_*`: frozen-support live chain 기준의 공식 current 결과를 먼저 확인한다. current preview/current report가 있으면 그쪽이 공식 current 설명의 주 문서다.\n"
        "- `fault_panel_result_precursor_report_v1.csv`: 아직 고장 신호는 없지만 추적 가치가 있는 precursor candidate를 본다.\n"
        "- raw-only artifact는 operator 기본 읽기 순서가 아니라 아래 `analyst/support 추가 자료` 섹션에서 필요 시 확인한다.\n\n"
        "## 해석 가이드\n"
        "- 이 문서는 공식 current 설명 문서를 대체하지 않는 안내/fallback 문서다. current preview/current report가 있으면 그쪽을 먼저 읽는다.\n"
        "- `fault_panel_result_current_*`는 frozen-support live chain 기준 결과다.\n"
        "- `fault_panel_result_raw_only_current_*`는 raw-only candidate 중 strict current subset만 보여준다.\n"
        "- `fault_panel_result_precursor_report_v1.csv`는 고장 신호가 아직 없는 precursor candidate만 보여준다.\n"
        f"- `{ROOT_FAULT_SIGNAL_REPORT_NAME}`는 raw-only candidate 우주에서 고장 신호가 이미 관측된 panel만 모은 analyst/support 보조표다.\n"
        "- raw-only chain의 `커널로그_원인군_ko`는 기존 라벨 family가 아니라 algorithm-derived family 의미다.\n"
        "- preview 표의 `운영 판정`은 현재 신호 단계, `상위 해석 후보`는 가장 가까운 원인 후보를 뜻한다.\n"
        "- `급락 종결 관측`과 `점진 저하 누적`은 관측 축이고, `사건 종결 요약`은 확정 row에서만 채워지는 사건 요약이다.\n"
        "- `고장 기준일`은 확정 고장일만 뜻하는 칼럼이 아니라 판단 기준으로 삼은 날짜다.\n"
        "- `기존 알고리즘 source`의 `미검출`은 legacy source 태그가 없다는 뜻이다.\n"
        "- precursor report와 raw-only fault signal report는 row가 중복되지 않게 분리해 읽어야 한다.\n"
        "- raw-only fault signal report의 row 수는 `panel_id` 기준 count이고, 같은 `subgroup base` 아래 여러 panel이 함께 잡히면 여러 row로 보일 수 있다.\n"
        f"- `subgroup cluster`는 같은 subgroup base 안에서 `신호 기준일` 간격이 `{FAULT_SIGNAL_CLUSTER_GAP_DAYS}`일 이하인 row를 하나의 보조 cluster로 묶은 analyst/support 휴리스틱이다.\n"
        "- 운영자는 기본적으로 current -> precursor 순서로 읽고, raw-only artifact는 analyst/support 추가 자료가 필요할 때만 연다.\n"
        "- 전체 candidate universe는 `result/raw_only_chain/*`와 detailed report 안에 그대로 남는다.\n\n"
        "## 컬럼 읽는 법\n"
        "- precursor report의 `전조 축`은 EWS/AE/DTW/규칙징후 중 어떤 축이 전조로 묶였는지 보여준다.\n"
        "- precursor report의 `대표 전조 신호`는 전조 후보를 만든 누적 신호를 짧게 풀어쓴 요약이다.\n"
        "- precursor report의 `모니터링 권고`는 다음 수집 주기에 무엇을 먼저 확인할지 알려주는 운영 메모다.\n"
        "- precursor report의 `공통원인 위험`과 `권고 검토 레인`은 panel-local precursor로 읽기 전에 공통 외란 가능성을 얼마나 먼저 볼지 정리한 보조 값이다.\n"
        "- raw-only fault signal report의 `group root`는 넓은 family root, `subgroup base`는 common-cause 검토에 더 가까운 하위 묶음이다.\n"
        "- raw-only fault signal report의 `동일 subgroup row 수`는 같은 subgroup base 아래 함께 잡힌 panel row 수다.\n"
        "- raw-only fault signal report의 `subgroup cluster`와 `동일 cluster row 수`는 `사건 수`를 직접 뜻하지 않고, 같은 subgroup base 안에서 가까운 날짜 row를 묶어 읽기 쉽게 만든 보조 값이다.\n"
        "- raw-only fault signal report의 `확정 경로`는 주 경로 하나만 보여주고, `고장 신호 요약`은 일수와 보조 근거를 덧붙인다.\n"
        "- raw-only fault signal report의 `근접 공통원인`은 strict_trigger 기준 ±3일 안의 common-cause만 적고, warning-anchor 기준 common-cause는 audit 전용으로 남긴다.\n"
        "- raw-only fault signal report의 `현장 점검 권고`는 첫 현장 액션의 우선순위를 짧게 적은 값이다.\n\n"
        "## 주요 산출물\n"
        f"{primary_output_block}\n\n"
        "## analyst/support 추가 자료\n"
        f"{analyst_output_block}\n\n"
        "## 상세 리포트\n"
        f"{detailed_report_block}"
        "## 전조 리포트\n"
        f"{precursor_report_block}"
        "## raw-only 고장 신호 리포트\n"
        f"{fault_signal_report_block}"
        "## raw-only 고장 신호 subgroup base 요약 (앞 10행)\n"
        f"{fault_signal_top_subgroup_block}\n\n"
        "## raw-only 고장 신호 subgroup cluster 요약 (앞 10행)\n"
        f"{fault_signal_top_cluster_block}\n\n"
        "## current preview 표\n"
        f"{markdown_table_from_df(truncate_report_df(live_preview_df))}\n\n"
        "## precursor 후보 표 (앞 20행)\n"
        f"{markdown_table_from_df(truncate_report_df(precursor_report_df, limit=20))}\n\n"
        "## analyst/support 참고 메모\n"
        f"- `{ROOT_FAULT_SIGNAL_REPORT_NAME}`와 `fault_panel_result_raw_only_current_*`는 master report에서 경로만 안내하는 보조 artifact다.\n"
        "- raw-only preview/fault signal row는 operator 기본 읽기 흐름에 직접 전개하지 않는다.\n"
        f"- raw-only strict current preview row count: `{len(raw_only_preview_df)}`\n"
        f"- raw-only fault signal row count: `{len(fault_signal_report_df)}`\n"
    )


def panel_group_root(panel_id: object) -> str:
    text = normalize_text(panel_id)
    tokens = text.split(".")
    if len(tokens) >= 3:
        return ".".join(tokens[:-2])
    return text


def panel_subgroup_base(panel_id: object) -> str:
    text = normalize_text(panel_id)
    tokens = text.split(".")
    if len(tokens) >= 2:
        return ".".join(tokens[:-1])
    return text


def attach_fault_signal_cluster_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        working = df.copy() if df is not None else pd.DataFrame()
        if "subgroup cluster" not in working.columns:
            working["subgroup cluster"] = pd.Series(dtype=object)
        if "동일 cluster row 수" not in working.columns:
            working["동일 cluster row 수"] = pd.Series(dtype=int)
        return working

    working = df.copy()
    working["신호 기준일_dt"] = pd.to_datetime(working["신호 기준일"], errors="coerce")
    cluster_key_by_index: dict[int, str] = {}
    cluster_size_by_index: dict[int, int] = {}

    for (_, subgroup_base), subgroup_rows in working.groupby(
        ["site", "subgroup base"], sort=False, dropna=False
    ):
        subgroup_rows = subgroup_rows.sort_values(
            ["신호 기준일_dt", "panel_id"], ascending=[True, True]
        ).copy()
        cluster_ids: list[int] = []
        cluster_id = 0
        prev_date = None
        for _, subgroup_row in subgroup_rows.iterrows():
            current_date = subgroup_row.get("신호 기준일_dt")
            if pd.isna(current_date):
                cluster_id += 1
            elif prev_date is None or (current_date - prev_date).days > FAULT_SIGNAL_CLUSTER_GAP_DAYS:
                cluster_id += 1
                prev_date = current_date
            else:
                prev_date = current_date
            cluster_ids.append(cluster_id)
        subgroup_rows["cluster_id"] = cluster_ids

        cluster_meta = (
            subgroup_rows.groupby("cluster_id", dropna=False)
            .agg(
                cluster_rows=("panel_id", "size"),
                start_date=("신호 기준일_dt", "min"),
                end_date=("신호 기준일_dt", "max"),
            )
            .reset_index()
        )
        label_map: dict[int, str] = {}
        size_map: dict[int, int] = {}
        for cluster_row in cluster_meta.to_dict(orient="records"):
            cid = int(cluster_row["cluster_id"])
            start_date = cluster_row.get("start_date")
            end_date = cluster_row.get("end_date")
            if pd.notna(start_date) and pd.notna(end_date):
                start_text = pd.Timestamp(start_date).strftime("%Y-%m-%d")
                end_text = pd.Timestamp(end_date).strftime("%Y-%m-%d")
                if start_text == end_text:
                    label = f"{normalize_text(subgroup_base)} @ {start_text}"
                else:
                    label = f"{normalize_text(subgroup_base)} @ {start_text}~{end_text}"
            else:
                label = f"{normalize_text(subgroup_base)} @ undated#{cid}"
            label_map[cid] = label
            size_map[cid] = int(cluster_row.get("cluster_rows", 0) or 0)

        subgroup_rows["subgroup cluster"] = subgroup_rows["cluster_id"].map(label_map)
        subgroup_rows["동일 cluster row 수"] = subgroup_rows["cluster_id"].map(size_map)
        cluster_key_by_index.update(subgroup_rows["subgroup cluster"].to_dict())
        cluster_size_by_index.update(subgroup_rows["동일 cluster row 수"].to_dict())

    working["subgroup cluster"] = working.index.map(cluster_key_by_index.get)
    working["동일 cluster row 수"] = working.index.map(cluster_size_by_index.get)
    return working.drop(columns=["신호 기준일_dt"], errors="ignore")


def bool_count(df: pd.DataFrame, column: str) -> int:
    if column not in df.columns or df.empty:
        return 0
    return int(truthy_mask(df[column]).sum())


def unique_csv(series: pd.Series) -> str:
    values = sorted({normalize_text(value) for value in series if normalize_text(value)})
    return ",".join(values)


def load_gate_from_workspace(workspace_root: Path, site: str) -> pd.DataFrame:
    path = workspace_root / "data" / site / "out" / "ae_simple_local_precursor_gate_daily.csv"
    if not path.exists():
        raise SystemExit(f"missing workspace precursor gate output: {path}")
    df = pd.read_csv(path, low_memory=False)
    ensure_columns(df, ["panel_id", "date"], path.name)
    df["panel_id"] = df["panel_id"].astype(str)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    return df


def report_attention_grade(evidence_row: dict[str, object]) -> str:
    final_days = int(evidence_row.get("final_days", 0) or 0)
    critical_days = int(evidence_row.get("critical_days", 0) or 0)
    critical_confirmed_days = int(evidence_row.get("critical_confirmed_days", 0) or 0)
    fault_like_days = int(evidence_row.get("fault_like_days", 0) or 0)
    ews_warning_days = int(evidence_row.get("ews_warning_days", 0) or 0)
    pre_alarm_days = int(evidence_row.get("pre_alarm_days", 0) or 0)
    pre_ews_days = int(evidence_row.get("pre_ews_days", 0) or 0)
    prefault_cond_ae_days = int(evidence_row.get("prefault_cond_ae_days", 0) or 0)
    prefault_cond_dtw_days = int(evidence_row.get("prefault_cond_dtw_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))

    if final_days > 0 or critical_days > 0 or critical_confirmed_days > 0:
        return "확정"
    if (
        "vdrop" in critical_sources
        or ews_warning_days >= 15
        or pre_alarm_days >= 10
        or pre_ews_days >= 50
        or prefault_cond_ae_days >= 120
        or prefault_cond_dtw_days >= 120
        or fault_like_days >= 2
    ):
        return "고위험 관찰"
    return "관찰"


def report_reason_text(evidence_row: dict[str, object]) -> str:
    grade = normalize_text(evidence_row.get("운영해석등급_ko"))
    top1 = normalize_text(evidence_row.get("1순위_의심원인_ko"))
    final_days = int(evidence_row.get("final_days", 0) or 0)
    critical_days = int(evidence_row.get("critical_days", 0) or 0)
    critical_confirmed_days = int(evidence_row.get("critical_confirmed_days", 0) or 0)
    ews_warning_days = int(evidence_row.get("ews_warning_days", 0) or 0)
    pre_ews_days = int(evidence_row.get("pre_ews_days", 0) or 0)
    prefault_B_effective_days = int(evidence_row.get("prefault_B_effective_days", 0) or 0)
    prefault_B_common_cause_overlap_days = int(evidence_row.get("prefault_B_common_cause_overlap_days", 0) or 0)
    prefault_cond_ae_days = int(evidence_row.get("prefault_cond_ae_days", 0) or 0)
    prefault_cond_dtw_days = int(evidence_row.get("prefault_cond_dtw_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))
    subtypes = normalize_text(evidence_row.get("anom_subtypes_csv"))
    subgroup_candidate_count = int(evidence_row.get("subgroup_candidate_panel_count", 0) or 0)

    if grade == "확정":
        signal_labels: list[str] = []
        if final_days > 0:
            signal_labels.append("최종 고장 신호")
        if critical_confirmed_days > 0:
            signal_labels.append("강한 고장 신호 확정")
        elif critical_days > 0:
            signal_labels.append("강한 고장 신호")
        if "vdrop" in critical_sources:
            signal_labels.append("vdrop 전기 신호")
        signal_summary = " / ".join(signal_labels) if signal_labels else "확정 신호"
        return f"{signal_summary}가 나타나 고장 신호가 뚜렷하게 포착됨"

    reasons: list[str] = []
    if "degradation" in subtypes:
        reasons.append("degradation subtype 반복")
    if ews_warning_days > 0 or pre_ews_days > 0:
        reasons.append(f"EWS 전조 누적(ews={ews_warning_days}, pre_ews={pre_ews_days})")
    if prefault_B_effective_days > 0:
        reasons.append(f"Option B 유효 전조 누적({prefault_B_effective_days}일)")
    if prefault_cond_ae_days > 0 or prefault_cond_dtw_days > 0:
        reasons.append(
            f"AE/DTW 전조 조건 누적(ae={prefault_cond_ae_days}, dtw={prefault_cond_dtw_days})"
        )
    if prefault_B_common_cause_overlap_days > 0:
        reasons.append(f"공통원인 겹침 option B({prefault_B_common_cause_overlap_days}일)는 별도 분리")
    if subgroup_candidate_count >= 3:
        reasons.append(f"동일 subgroup 동시 흔들림({subgroup_candidate_count} panels)")
    if top1:
        reasons.append(f"가장 가까운 후보는 {top1}")
    if not reasons:
        reasons.append("약한 이상 신호만 있어 관찰 대상으로 해석")
    return " / ".join(reasons)


def report_precursor_axes_text(evidence_row: dict[str, object]) -> str:
    axes: list[str] = []
    ews_warning_days = int(evidence_row.get("ews_warning_days", 0) or 0)
    pre_alarm_days = int(evidence_row.get("pre_alarm_days", 0) or 0)
    pre_ews_days = int(evidence_row.get("pre_ews_days", 0) or 0)
    prefault_cond_ae_days = int(evidence_row.get("prefault_cond_ae_days", 0) or 0)
    prefault_cond_dtw_days = int(evidence_row.get("prefault_cond_dtw_days", 0) or 0)
    event_A_days = int(evidence_row.get("event_A_days", 0) or 0)
    fault_like_days = int(evidence_row.get("fault_like_days", 0) or 0)
    final_days = int(evidence_row.get("final_days", 0) or 0)
    critical_days = int(evidence_row.get("critical_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))
    subtypes = normalize_text(evidence_row.get("anom_subtypes_csv"))

    if ews_warning_days > 0 or pre_ews_days > 0:
        axes.append("EWS")
    if prefault_cond_ae_days > 0 or event_A_days > 0 or "degradation" in subtypes:
        axes.append("AE")
    if prefault_cond_dtw_days > 0:
        axes.append("DTW")
    if (
        pre_alarm_days > 0
        or fault_like_days > 0
        or final_days > 0
        or critical_days > 0
        or "vdrop" in critical_sources
    ):
        axes.append("규칙징후")
    return "+".join(axes)


def report_precursor_signal_text(evidence_row: dict[str, object]) -> str:
    signals: list[str] = []
    ews_warning_days = int(evidence_row.get("ews_warning_days", 0) or 0)
    pre_alarm_days = int(evidence_row.get("pre_alarm_days", 0) or 0)
    pre_ews_days = int(evidence_row.get("pre_ews_days", 0) or 0)
    prefault_B_effective_days = int(evidence_row.get("prefault_B_effective_days", 0) or 0)
    prefault_B_common_cause_overlap_days = int(evidence_row.get("prefault_B_common_cause_overlap_days", 0) or 0)
    prefault_cond_ae_days = int(evidence_row.get("prefault_cond_ae_days", 0) or 0)
    prefault_cond_dtw_days = int(evidence_row.get("prefault_cond_dtw_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))
    subtypes = normalize_text(evidence_row.get("anom_subtypes_csv"))

    if ews_warning_days > 0 or pre_ews_days > 0:
        signals.append(f"EWS 전조 누적(ews={ews_warning_days}, pre_ews={pre_ews_days})")
    if prefault_B_effective_days > 0:
        signals.append(f"Option B 유효 누적({prefault_B_effective_days}일)")
    if prefault_cond_ae_days > 0:
        signals.append(f"AE 전조 조건 누적({prefault_cond_ae_days}일)")
    if prefault_cond_dtw_days > 0:
        signals.append(f"DTW 전조 조건 누적({prefault_cond_dtw_days}일)")
    if pre_alarm_days > 0:
        signals.append(f"pre_alarm 누적({pre_alarm_days}일)")
    if "vdrop" in critical_sources:
        signals.append("상대 전압 이탈 징후")
    if "degradation" in subtypes:
        signals.append("degradation subtype 반복")
    if prefault_B_common_cause_overlap_days > 0:
        signals.append(f"공통원인 겹침 option B({prefault_B_common_cause_overlap_days}일)")
    return " / ".join(signals)


def precursor_common_cause_risk_text(evidence_row: dict[str, object]) -> str:
    prefault_B_common_cause_overlap_days = int(
        evidence_row.get("prefault_B_common_cause_overlap_days", 0) or 0
    )
    subgroup_candidate_count = int(evidence_row.get("subgroup_candidate_panel_count", 0) or 0)
    if prefault_B_common_cause_overlap_days > 0 and subgroup_candidate_count >= 3:
        return "높음"
    if prefault_B_common_cause_overlap_days > 0 or subgroup_candidate_count >= 3:
        return "중간"
    return "낮음"


def precursor_review_lane_text(evidence_row: dict[str, object]) -> str:
    risk = precursor_common_cause_risk_text(evidence_row)
    grade = normalize_text(evidence_row.get("운영해석등급_ko"))
    if risk == "높음":
        return "공통원인 검토"
    if risk == "중간":
        return "공통원인 우선 확인"
    if grade == "고위험 관찰":
        return "단일 패널 우선 추적"
    return "일반 모니터링"


def precursor_monitoring_action_text(evidence_row: dict[str, object]) -> str:
    grade = normalize_text(evidence_row.get("운영해석등급_ko"))
    top1 = normalize_text(evidence_row.get("1순위_의심원인_ko"))
    axes = report_precursor_axes_text(evidence_row)
    prefault_B_effective_days = int(evidence_row.get("prefault_B_effective_days", 0) or 0)
    prefault_B_common_cause_overlap_days = int(evidence_row.get("prefault_B_common_cause_overlap_days", 0) or 0)
    subgroup_candidate_count = int(evidence_row.get("subgroup_candidate_panel_count", 0) or 0)
    if prefault_B_common_cause_overlap_days > 0 and subgroup_candidate_count >= 3:
        return "site_event/group_off 및 동일 subgroup 동시 흔들림을 먼저 재확인"
    if prefault_B_common_cause_overlap_days > 0 and prefault_B_effective_days == 0:
        return "site_event/group_off 공통원인 여부를 먼저 재확인"
    if subgroup_candidate_count >= 3 and prefault_B_effective_days == 0:
        return "동일 subgroup 동시 흔들림과 공통 외란 여부를 먼저 재확인"
    if "오염" in top1:
        return "세척 전후 추세 비교와 추가 관찰 권고"
    if "음영" in top1:
        return "인접 음영 구조와 시간대별 반복 여부 재확인 권고"
    if "접촉" in top1 or "끊김" in top1:
        return "다음 수집 주기 재확인 후 접속부 점검 여부 판단"
    if grade == "고위험 관찰":
        return "가까운 주기 재확인과 현장 비교 점검 권고"
    if axes:
        return f"{axes} 축 모니터링 유지"
    return "지속 모니터링 유지"


def strict_trigger_common_cause_text(evidence_row: dict[str, object]) -> str:
    if bool(evidence_row.get("strict_trigger_proximal_common_cause_flag")):
        return "strict_trigger 근처 공통원인 흔들림 동반"
    return ""


def fault_signal_path_text(evidence_row: dict[str, object]) -> str:
    final_days = int(evidence_row.get("final_days", 0) or 0)
    critical_days = int(evidence_row.get("critical_days", 0) or 0)
    critical_confirmed_days = int(evidence_row.get("critical_confirmed_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))
    if final_days > 0:
        return "최종 고장 신호 경로"
    if critical_confirmed_days > 0:
        return "강한 고장 신호 확정 경로"
    if critical_days > 0:
        return "vdrop 강신호 경로" if "vdrop" in critical_sources else "강한 고장 신호 경로"
    return "고장 신호 관측"


def fault_signal_summary_text(evidence_row: dict[str, object]) -> str:
    final_days = int(evidence_row.get("final_days", 0) or 0)
    critical_days = int(evidence_row.get("critical_days", 0) or 0)
    critical_confirmed_days = int(evidence_row.get("critical_confirmed_days", 0) or 0)
    critical_sources = normalize_text(evidence_row.get("critical_sources_csv"))
    parts: list[str] = []
    if final_days > 0:
        parts.append(f"최종 고장 신호 {final_days}일")
        if critical_confirmed_days > 0:
            parts.append("강한 고장 신호 확정이 함께 관측됨")
        elif critical_days > 0:
            parts.append("강한 고장 신호가 함께 관측됨")
    elif critical_confirmed_days > 0:
        parts.append(f"강한 고장 신호 확정 {critical_confirmed_days}일")
    elif critical_days > 0:
        parts.append(f"강한 고장 신호 {critical_days}일")
    if "vdrop" in critical_sources:
        parts.append("vdrop 전기 신호 동반")
    if bool(evidence_row.get("strict_trigger_proximal_common_cause_flag")):
        parts.append("strict_trigger 근처 공통원인 흔들림 동반")
    return " / ".join(parts) if parts else "고장 신호 관측"


def fault_signal_action_text(evidence_row: dict[str, object]) -> str:
    top1 = normalize_text(evidence_row.get("1순위_의심원인_ko"))
    if bool(evidence_row.get("strict_trigger_proximal_common_cause_flag")):
        return "패널 국소 고장 신호와 함께 strict_trigger 근처 공통원인 여부도 동시 확인"
    if "다이오드" in top1 or "국소 회로" in top1:
        return "현장 점검 후 다이오드·국소 회로 이상 여부 우선 확인"
    if "접촉" in top1 or "끊김" in top1 or "개방" in top1:
        return "배선·접속부 우선 점검"
    if "측정" in top1 or "응답" in top1:
        return "MLPE/계측값과 접속 상태 동시 점검"
    if "외부 전원" in top1:
        return "패널 국소 이상보다 외부 전원/공통 원인 먼저 확인"
    return "현장 점검과 최근 작업 이력 확인 권고"


def build_precursor_report_df(evidence_df: pd.DataFrame) -> pd.DataFrame:
    if evidence_df is None or evidence_df.empty:
        return pd.DataFrame(columns=PRECURSOR_REPORT_OUTPUT_COLS)

    rows: list[dict[str, object]] = []
    for _, row in evidence_df.fillna("").iterrows():
        precursor_date = normalize_text(row.get("전조날짜"))
        evidence_row = row.to_dict()
        if not precursor_date or not has_precursor_signal(evidence_row) or has_hard_fault_evidence(evidence_row):
            continue
        rows.append(
            {
                "site": normalize_text(row.get("site")),
                "panel_id": normalize_text(row.get("panel_id")),
                "운영 판정": normalize_text(row.get("운영해석등급_ko")) or "전조 후보",
                "판정 근거": report_reason_text(evidence_row),
                "전조날짜": precursor_date,
                "전조 축": report_precursor_axes_text(evidence_row),
                "대표 전조 신호": report_precursor_signal_text(evidence_row),
                "전조 요약": normalize_text(row.get("근거요약_ko")),
                "상위 해석 후보": normalize_text(row.get("1순위_의심원인_ko")),
                "기존 알고리즘 source": display_existing_algorithm_source(
                    row.get("커널로그 기존 알고리즘")
                ),
                "패턴 설명": pattern_explainer(evidence_row, soften_hard_language=True),
                "모니터링 권고": precursor_monitoring_action_text(evidence_row),
                "공통원인 위험": precursor_common_cause_risk_text(evidence_row),
                "권고 검토 레인": precursor_review_lane_text(evidence_row),
                "EWS 전조 일수": int(row.get("ews_warning_days", 0) or 0),
                "pre_alarm 일수": int(row.get("pre_alarm_days", 0) or 0),
                "pre_ews 일수": int(row.get("pre_ews_days", 0) or 0),
                "Option B 유효 일수": int(row.get("prefault_B_effective_days", 0) or 0),
                "공통원인 겹침 일수": int(row.get("prefault_B_common_cause_overlap_days", 0) or 0),
                "AE 전조 조건 일수": int(row.get("prefault_cond_ae_days", 0) or 0),
                "DTW 전조 조건 일수": int(row.get("prefault_cond_dtw_days", 0) or 0),
            }
        )
    return (
        pd.DataFrame(rows)
        .reindex(columns=PRECURSOR_REPORT_OUTPUT_COLS)
        .sort_values(["site", "panel_id"], ascending=[True, True])
        .reset_index(drop=True)
    )


def build_fault_signal_report_df(evidence_df: pd.DataFrame) -> pd.DataFrame:
    if evidence_df is None or evidence_df.empty:
        return pd.DataFrame(columns=FAULT_SIGNAL_REPORT_OUTPUT_COLS)

    rows: list[dict[str, object]] = []
    for _, row in evidence_df.fillna("").iterrows():
        evidence_row = row.to_dict()
        if not has_hard_fault_evidence(evidence_row):
            continue
        event_fields = event_display_fields(evidence_row)
        rows.append(
            {
                "site": normalize_text(row.get("site")),
                "group root": normalize_text(row.get("group_root")),
                "subgroup base": normalize_text(row.get("subgroup_base")),
                "panel_id": normalize_text(row.get("panel_id")),
                "동일 subgroup row 수": int(row.get("subgroup_candidate_panel_count", 0) or 0),
                "운영 판정": normalize_text(row.get("운영해석등급_ko")) or display_signal_grade(row),
                "확정 경로": fault_signal_path_text(evidence_row),
                "고장 신호 요약": fault_signal_summary_text(evidence_row),
                "전조 시작일": normalize_text(row.get("전조날짜")),
                "신호 기준일": normalize_text(row.get("고장날짜")),
                "사건유형": normalize_text(row.get("사건유형_ko")),
                "사건 종결 요약": event_fields.get("사건 종결 요약", ""),
                "근접 공통원인": strict_trigger_common_cause_text(evidence_row),
                "상위 해석 후보": normalize_text(row.get("1순위_의심원인_ko")),
                "기존 알고리즘 source": display_existing_algorithm_source(
                    row.get("커널로그 기존 알고리즘")
                ),
                "패턴 설명": pattern_explainer(evidence_row),
                "현장 점검 권고": fault_signal_action_text(evidence_row),
            }
        )
    working = pd.DataFrame(rows)
    working = attach_fault_signal_cluster_columns(working)
    working = (
        working.reindex(columns=FAULT_SIGNAL_REPORT_OUTPUT_COLS)
        .sort_values(["site", "subgroup base", "신호 기준일", "panel_id"], ascending=[True, True, True, True])
        .reset_index(drop=True)
    )
    if "동일 cluster row 수" in working.columns:
        working["동일 cluster row 수"] = working["동일 cluster row 수"].fillna(0).astype(int)
    return working


def nonempty_sheet_df(df: pd.DataFrame, note: str) -> pd.DataFrame:
    if not df.empty:
        return df
    return pd.DataFrame([{"note": note}])


def signal_label_text(record: dict[str, object]) -> str:
    labels: list[str] = []
    if bool(record.get("event_A")):
        labels.append("event_A")
    if bool(record.get("v_drop")):
        labels.append("v_drop")
    if bool(record.get("critical_fault")):
        labels.append("critical_fault")
    if bool(record.get("critical_suspect")):
        labels.append("critical_suspect")
    if bool(record.get("critical_confirmed")):
        labels.append("critical_confirmed")
    if bool(record.get("fault_like_day")):
        labels.append("fault_like")
    if bool(record.get("final_fault")):
        labels.append("final_fault")
    if bool(record.get("ews_warning")):
        labels.append("ews_warning")
    if bool(record.get("pre_alarm")):
        labels.append("pre_alarm")
    if bool(record.get("pre_ews")):
        labels.append("pre_ews")
    if bool(record.get("site_event_soft")):
        labels.append("site_event_soft")
    if bool(record.get("site_event_hard")):
        labels.append("site_event_hard")
    if bool(record.get("group_off_date")):
        labels.append("group_off")
    if bool(record.get("prefault_B")):
        labels.append("prefault_B")
    if bool(record.get("prefault_B_effective")):
        labels.append("prefault_B_effective")
    if bool(record.get("prefault_B_common_cause_overlap")):
        labels.append("prefault_B_common_cause_overlap")
    if bool(record.get("prefault_cond_mid")):
        labels.append("prefault_mid")
    if bool(record.get("prefault_cond_ae")):
        labels.append("prefault_ae")
    if bool(record.get("prefault_cond_dtw")):
        labels.append("prefault_dtw")
    if bool(record.get("prefault_cond_ews")):
        labels.append("prefault_ews")
    subtype = normalize_text(record.get("anom_subtype"))
    if subtype:
        labels.append(f"subtype:{subtype}")
    return ",".join(labels)


def auto_fit_workbook_columns(path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise SystemExit(
            "openpyxl is required to generate fault_panel_result_detailed_report_v1.xlsx"
        ) from exc

    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        if worksheet.max_row >= 2:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 60)
    workbook.save(path)


def build_detailed_report_frames(
    output_root: Path,
    sites: list[str],
    baseline_comparison: dict[str, object],
    live_chain_result: dict[str, object],
    raw_only_chain_result: dict[str, object],
    live_preview_df: pd.DataFrame,
    raw_only_preview_df: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    overview_df = pd.DataFrame(
        [
            {"section": "sites", "key": "sites_csv", "value": ",".join(sites)},
            {
                "section": "baseline",
                "key": "all_sites_match",
                "value": str(baseline_comparison.get("all_sites_match")),
            },
            {
                "section": "live_chain",
                "key": "status_ko",
                "value": normalize_text(live_chain_result.get("status_ko")),
            },
            {
                "section": "live_chain",
                "key": "fixed_fault_reference_exact_match",
                "value": str(
                    live_chain_result.get("fixed_fault_reference_compare", {}).get("exact_match")
                ),
            },
            {
                "section": "raw_only_chain",
                "key": "status_ko",
                "value": normalize_text(raw_only_chain_result.get("status_ko")),
            },
            {
                "section": "raw_only_chain",
                "key": "compare_status_ko",
                "value": normalize_text(
                    raw_only_chain_result.get("fixed_fault_reference_compare", {}).get("status_ko")
                ),
            },
            {
                "section": "raw_only_chain",
                "key": "candidate_row_count",
                "value": str(
                    raw_only_chain_result.get("fixed_fault_reference_compare", {}).get(
                        "candidate_row_count"
                    )
                ),
            },
            {
                "section": "raw_only_chain",
                "key": "reference_row_count",
                "value": str(
                    raw_only_chain_result.get("fixed_fault_reference_compare", {}).get(
                        "reference_row_count"
                    )
                ),
            },
            {
                "section": "notes",
                "key": "attention_grade_note_ko",
                "value": (
                    "운영해석등급_ko는 상세 리포트용 보조 등급이다. core verdict를 바꾸지 않고 "
                    "확정/고위험 관찰/관찰을 사람이 읽기 쉽게 정리한다."
                ),
            },
        ]
    )

    frames: dict[str, pd.DataFrame] = {
        "overview": overview_df,
        "current_preview": nonempty_sheet_df(
            live_preview_df.copy(),
            "live current preview not available",
        ),
        "raw_only_preview": nonempty_sheet_df(
            raw_only_preview_df.copy(),
            "raw-only preview not available",
        ),
    }

    if not raw_only_chain_result.get("requested") or normalize_text(raw_only_chain_result.get("status_ko")) != "completed":
        frames["raw_only_evidence"] = pd.DataFrame(
            [{"note": "raw-only chain not completed; detailed evidence unavailable"}]
        )
        frames["raw_only_candidate_scores"] = pd.DataFrame(
            [{"note": "raw-only chain not completed; candidate score matrix unavailable"}]
        )
        frames["raw_only_timeline"] = pd.DataFrame(
            [{"note": "raw-only chain not completed; timeline unavailable"}]
        )
        frames["raw_only_daily_log"] = pd.DataFrame(
            [{"note": "raw-only chain not completed; all-date log unavailable"}]
        )
        frames["raw_only_cluster"] = pd.DataFrame(
            [{"note": "raw-only chain not completed; cluster summary unavailable"}]
        )
        frames["precursor_report"] = pd.DataFrame(columns=PRECURSOR_REPORT_OUTPUT_COLS)
        frames["fault_signal_report"] = pd.DataFrame(columns=FAULT_SIGNAL_REPORT_OUTPUT_COLS)
        frames["definitions"] = pd.DataFrame(
            [
                {
                    "항목": "확정",
                    "설명": "최종 고장 신호 또는 강한 고장 신호가 관측된 상태",
                },
                {
                    "항목": "고위험 관찰",
                    "설명": "즉시 확정에 쓰는 신호는 없지만 EWS/AE/DTW 전조가 강하게 누적",
                },
                {"항목": "관찰", "설명": "약한 이상 또는 간헐 이상으로 계속 관찰 필요"},
            ]
        )
        return frames

    workspace_root = Path(str(raw_only_chain_result["workspace_root"]))
    raw_only_common = load_raw_only_common_module()
    runtime_heuristic = load_runtime_heuristic_module()
    audit_path = workspace_root / "_share" / raw_only_common.RUNTIME_AUDIT_OUTPUT_NAME
    heuristic_path = workspace_root / "_share" / raw_only_common.RUNTIME_HEURISTIC_OUTPUT_NAME
    verdict_path = workspace_root / "_share" / raw_only_common.RUNTIME_VERDICT_OUTPUT_NAME
    audit_df = pd.read_csv(audit_path, encoding="utf-8-sig", low_memory=False)
    heuristic_df = pd.read_csv(heuristic_path, encoding="utf-8-sig", low_memory=False)
    verdict_df = pd.read_csv(verdict_path, encoding="utf-8-sig", low_memory=False)
    audit_df["site"] = audit_df["site"].astype(str)
    audit_df["panel_id"] = audit_df["panel_id"].astype(str)
    heuristic_df["site"] = heuristic_df["site"].astype(str)
    heuristic_df["panel_id"] = heuristic_df["panel_id"].astype(str)
    verdict_df["site"] = verdict_df["site"].astype(str)
    verdict_df["panel_id"] = verdict_df["panel_id"].astype(str)

    audit_lookup = {
        row_key(row["site"], row["panel_id"]): row for row in audit_df.to_dict(orient="records")
    }
    heuristic_lookup = {
        row_key(row["site"], row["panel_id"]): row for row in heuristic_df.to_dict(orient="records")
    }
    verdict_lookup = {
        row_key(row["site"], row["panel_id"]): row for row in verdict_df.to_dict(orient="records")
    }

    per_site_core = {site: load_panel_day_core_from_workspace(workspace_root, site) for site in sites}
    per_site_gate = {site: load_gate_from_workspace(workspace_root, site) for site in sites}
    preview_with_group_keys = raw_only_preview_df.copy()
    if not preview_with_group_keys.empty:
        preview_with_group_keys["group_root"] = preview_with_group_keys["panel_id"].map(
            panel_group_root
        )
        preview_with_group_keys["subgroup_base"] = preview_with_group_keys["panel_id"].map(
            panel_subgroup_base
        )
    group_root_counts = (
        preview_with_group_keys.groupby(["site", "group_root"]).size().to_dict()
        if not preview_with_group_keys.empty
        else {}
    )
    subgroup_counts = (
        preview_with_group_keys.groupby(["site", "subgroup_base"]).size().to_dict()
        if not preview_with_group_keys.empty
        else {}
    )

    evidence_rows: list[dict[str, object]] = []
    candidate_score_rows: list[dict[str, object]] = []
    timeline_rows: list[dict[str, object]] = []
    all_date_rows: list[dict[str, object]] = []
    for _, preview_row in raw_only_preview_df.iterrows():
        site = normalize_text(preview_row.get("site"))
        panel_id = normalize_text(preview_row.get("panel_id"))
        group_root = panel_group_root(panel_id)
        subgroup_base = panel_subgroup_base(panel_id)
        base = group_root
        key = row_key(site, panel_id)
        audit_row = audit_lookup.get(key, {})
        heuristic_row = heuristic_lookup.get(key, {})
        verdict_row = verdict_lookup.get(key, {})
        merged_for_scores = dict(verdict_row)
        merged_for_scores.update(audit_row)
        score_map, score_notes = runtime_heuristic.score_row(merged_for_scores)
        ranked_candidates = runtime_heuristic.choose_ranked_candidates(score_map)
        top_score = ranked_candidates[0][1] if ranked_candidates else 0
        panel_core = per_site_core[site].loc[per_site_core[site]["panel_id"].eq(panel_id)].copy()
        panel_gate = per_site_gate[site].loc[per_site_gate[site]["panel_id"].eq(panel_id)].copy()
        representative = representative_signal_row(panel_core)

        evidence_row: dict[str, object] = {
            "site": site,
            "panel_id": panel_id,
            "base": group_root,
            "group_root": group_root,
            "subgroup_base": subgroup_base,
            "base_candidate_panel_count": int(group_root_counts.get((site, group_root), 0)),
            "subgroup_candidate_panel_count": int(subgroup_counts.get((site, subgroup_base), 0)),
            "패널고장여부_ko": normalize_text(preview_row.get("패널고장여부_ko")),
            "사건유형_ko": normalize_text(preview_row.get("사건유형_ko")),
            "최종고장양상_ko": normalize_text(preview_row.get("최종고장양상_ko")),
            "라벨된 fault": normalize_text(preview_row.get("라벨된 fault")),
            "1순위_의심원인_ko": normalize_text(preview_row.get("1순위_의심원인_ko")),
            "2순위_의심원인_ko": normalize_text(preview_row.get("2순위_의심원인_ko")),
            "3순위_의심원인_ko": normalize_text(preview_row.get("3순위_의심원인_ko")),
            "커널로그 기존 알고리즘": normalize_text(preview_row.get("커널로그 기존 알고리즘")),
            "final_days": bool_count(panel_core, "final_fault"),
            "critical_days": bool_count(panel_core, "critical_fault"),
            "fault_like_days": bool_count(panel_core, "fault_like_day"),
            "event_A_days": bool_count(panel_core, "event_A"),
            "ews_warning_days": bool_count(panel_gate, "ews_warning"),
            "pre_alarm_days": bool_count(panel_gate, "pre_alarm"),
            "pre_ews_days": bool_count(panel_gate, "pre_ews"),
            "critical_confirmed_days": bool_count(panel_core, "critical_confirmed"),
            "prefault_B_days": bool_count(panel_gate, "prefault_B"),
            "prefault_B_effective_days": bool_count(panel_gate, "prefault_B_effective"),
            "prefault_B_common_cause_overlap_days": bool_count(panel_gate, "prefault_B_common_cause_overlap"),
            "prefault_cond_mid_days": bool_count(panel_gate, "prefault_cond_mid"),
            "prefault_cond_ae_days": bool_count(panel_gate, "prefault_cond_ae"),
            "prefault_cond_dtw_days": bool_count(panel_gate, "prefault_cond_dtw"),
            "prefault_cond_ews_days": bool_count(panel_gate, "prefault_cond_ews"),
            "critical_sources_csv": unique_csv(panel_core.get("critical_source", pd.Series(dtype=object))),
            "anom_subtypes_csv": unique_csv(panel_core.get("anom_subtype", pd.Series(dtype=object))),
            "원인후보_top1_score": heuristic_row.get("원인후보_top1_score", ""),
            "원인후보_top2_score": heuristic_row.get("원인후보_top2_score", ""),
            "원인후보_top3_score": heuristic_row.get("원인후보_top3_score", ""),
            "원인후보_경합상태_ko": normalize_text(heuristic_row.get("원인후보_경합상태_ko")),
            "원인후보_공동상위후보_csv": normalize_text(heuristic_row.get("원인후보_공동상위후보_csv")),
            "원인후보_실증우선확인_ko": normalize_text(heuristic_row.get("원인후보_실증우선확인_ko")),
            "원인후보_신뢰도_ko": normalize_text(heuristic_row.get("원인후보_신뢰도_ko")),
            "원인후보_해석메모_ko": normalize_text(heuristic_row.get("원인후보_해석메모_ko")),
            "사건이력_ko": normalize_text(verdict_row.get("사건이력_ko")),
            "대표판정_ko": normalize_text(verdict_row.get("대표판정_ko")),
            "운영최초전조발견일": normalize_text(verdict_row.get("운영최초전조발견일")),
            "사건해석상전조시작일": normalize_text(verdict_row.get("사건해석상전조시작일")),
            "세부fault_기준일": normalize_text(verdict_row.get("세부fault_기준일")),
            "판정주의_ko": normalize_text(verdict_row.get("판정주의_ko")),
            "strict_trigger_proximal_common_cause_flag": int(
                audit_row.get("strict_trigger_proximal_common_cause_flag", 0) or 0
            ),
            "warning_proximal_common_cause_flag": int(
                audit_row.get("warning_proximal_common_cause_flag", 0) or 0
            ),
            "대표critical_source": normalize_text(representative.get("critical_source")),
            "대표anom_subtype": normalize_text(representative.get("anom_subtype")),
            "대표mid_ratio": representative.get("mid_ratio", ""),
            "대표mid_v_ratio": representative.get("mid_v_ratio", ""),
            "대표mid_i_ratio": representative.get("mid_i_ratio", ""),
            "대표recon_error": representative.get("recon_error", ""),
            "대표dtw_dist": representative.get("dtw_dist", ""),
            "대표hs_score": representative.get("hs_score", ""),
            "대표event_A": normalize_text(representative.get("event_A")),
            "대표critical_fault": normalize_text(representative.get("critical_fault")),
            "대표critical_confirmed": normalize_text(representative.get("critical_confirmed")),
            "대표final_fault": normalize_text(representative.get("final_fault")),
        }
        evidence_row["전조날짜"] = choose_display_precursor_date(
            event_type_ko=preview_row.get("사건유형_ko"),
            interpreted_onset_date=verdict_row.get("사건해석상전조시작일"),
            first_warning_date=audit_row.get("earliest_warning_date"),
        )
        evidence_row["고장날짜"] = choose_display_fault_date(
            fault_date=verdict_row.get("세부fault_기준일"),
            strict_trigger_date=audit_row.get("strict_trigger_date"),
            first_final_fault_date=audit_row.get("first_final_fault_date"),
        )
        evidence_row["운영해석등급_ko"] = report_attention_grade(evidence_row)
        evidence_row["근거요약_ko"] = report_reason_text(evidence_row)
        evidence_rows.append(evidence_row)
        for rank_idx, (candidate, score) in enumerate(ranked_candidates, start=1):
            candidate_score_rows.append(
                {
                    "site": site,
                    "panel_id": panel_id,
                    "base": base,
                    "운영해석등급_ko": evidence_row["운영해석등급_ko"],
                    "패널고장여부_ko": evidence_row["패널고장여부_ko"],
                    "사건유형_ko": evidence_row["사건유형_ko"],
                    "최종고장양상_ko": evidence_row["최종고장양상_ko"],
                    "라벨된 fault": evidence_row["라벨된 fault"],
                    "후보순위": rank_idx,
                    "후보canonical_ko": candidate,
                    "후보표시명_ko": display_heuristic_name(candidate),
                    "후보점수": score,
                    "top1_flag": rank_idx == 1,
                    "공동상위_flag": bool(score == top_score and top_score > 0),
                    "원인후보_경합상태_ko": normalize_text(heuristic_row.get("원인후보_경합상태_ko")),
                    "원인후보_신뢰도_ko": normalize_text(heuristic_row.get("원인후보_신뢰도_ko")),
                    "커널로그 기존 알고리즘": evidence_row["커널로그 기존 알고리즘"],
                    "critical_sources_csv": evidence_row["critical_sources_csv"],
                    "anom_subtypes_csv": evidence_row["anom_subtypes_csv"],
                    "점수근거메모_ko": ", ".join(score_notes),
                    "후보해석메모_ko": normalize_text(heuristic_row.get("원인후보_해석메모_ko")),
                }
            )

        core_cols = [
            "date",
            "recon_error",
            "dtw_dist",
            "hs_score",
            "mid_ratio",
            "mid_peer",
            "mid_v_ratio",
            "mid_i_ratio",
            "last_ratio",
            "last_peer",
            "event_A",
            "v_drop",
            "critical_fault",
            "critical_suspect",
            "critical_confirmed",
            "group_off_like",
            "fault_like_day",
            "final_fault",
            "critical_source",
            "anom_level",
            "anom_subtype",
        ]
        gate_cols = [
            "date",
            "ews_warning",
            "pre_alarm",
            "pre_ews",
            "site_event_soft",
            "site_event_hard",
            "group_off_date",
            "prefault_B",
            "prefault_B_effective",
            "prefault_B_common_cause_overlap",
            "prefault_cond_mid",
            "prefault_cond_ae",
            "prefault_cond_dtw",
            "prefault_cond_ews",
        ]
        merged = panel_core.loc[:, [c for c in core_cols if c in panel_core.columns]].merge(
            panel_gate.loc[:, [c for c in gate_cols if c in panel_gate.columns]],
            on="date",
            how="outer",
        )
        signal_cols = [
            "event_A",
            "critical_fault",
            "critical_suspect",
            "critical_confirmed",
            "group_off_like",
            "fault_like_day",
            "final_fault",
            "ews_warning",
            "pre_alarm",
            "pre_ews",
            "site_event_soft",
            "site_event_hard",
            "group_off_date",
            "prefault_B",
            "prefault_B_effective",
            "prefault_B_common_cause_overlap",
            "prefault_cond_mid",
            "prefault_cond_ae",
            "prefault_cond_dtw",
            "prefault_cond_ews",
        ]
        available_signal_cols = [column for column in signal_cols if column in merged.columns]
        signal_mask = merged[available_signal_cols].fillna(False).astype(bool).any(axis=1)
        subtype_mask = merged.get("anom_subtype", pd.Series(dtype=object)).astype(str).str.contains(
            "degradation|fault_like|shadow_like|critical",
            case=False,
            na=False,
        )
        merged = merged.sort_values("date").reset_index(drop=True)
        merged["신호있는날_flag"] = (signal_mask | subtype_mask).reset_index(drop=True)
        for record in merged.to_dict(orient="records"):
            all_date_rows.append(
                {
                    "site": site,
                    "panel_id": panel_id,
                    "base": base,
                    "운영해석등급_ko": evidence_row["운영해석등급_ko"],
                    "1순위_의심원인_ko": evidence_row["1순위_의심원인_ko"],
                    "date": pd.to_datetime(record.get("date"), errors="coerce"),
                    "신호있는날_flag": bool(record.get("신호있는날_flag")),
                    "관찰포인트_csv": signal_label_text(record),
                    "recon_error": record.get("recon_error"),
                    "dtw_dist": record.get("dtw_dist"),
                    "hs_score": record.get("hs_score"),
                    "mid_ratio": record.get("mid_ratio"),
                    "mid_peer": record.get("mid_peer"),
                    "mid_v_ratio": record.get("mid_v_ratio"),
                    "mid_i_ratio": record.get("mid_i_ratio"),
                    "last_ratio": record.get("last_ratio"),
                    "last_peer": record.get("last_peer"),
                    "event_A": record.get("event_A"),
                    "v_drop": record.get("v_drop"),
                    "critical_fault": record.get("critical_fault"),
                    "critical_suspect": record.get("critical_suspect"),
                    "critical_confirmed": record.get("critical_confirmed"),
                    "fault_like_day": record.get("fault_like_day"),
                    "final_fault": record.get("final_fault"),
                    "ews_warning": record.get("ews_warning"),
                    "pre_alarm": record.get("pre_alarm"),
                    "pre_ews": record.get("pre_ews"),
                    "site_event_soft": record.get("site_event_soft"),
                    "site_event_hard": record.get("site_event_hard"),
                    "group_off_date": record.get("group_off_date"),
                    "prefault_B": record.get("prefault_B"),
                    "prefault_B_effective": record.get("prefault_B_effective"),
                    "prefault_B_common_cause_overlap": record.get("prefault_B_common_cause_overlap"),
                    "prefault_cond_mid": record.get("prefault_cond_mid"),
                    "prefault_cond_ae": record.get("prefault_cond_ae"),
                    "prefault_cond_dtw": record.get("prefault_cond_dtw"),
                    "prefault_cond_ews": record.get("prefault_cond_ews"),
                    "critical_source": normalize_text(record.get("critical_source")),
                    "anom_level": normalize_text(record.get("anom_level")),
                    "anom_subtype": normalize_text(record.get("anom_subtype")),
                }
            )
        for record in merged.loc[merged["신호있는날_flag"]].to_dict(orient="records"):
            timeline_rows.append(
                {
                    "site": site,
                    "panel_id": panel_id,
                    "base": base,
                    "운영해석등급_ko": evidence_row["운영해석등급_ko"],
                    "1순위_의심원인_ko": evidence_row["1순위_의심원인_ko"],
                    "date": pd.to_datetime(record.get("date"), errors="coerce"),
                    "recon_error": record.get("recon_error"),
                    "dtw_dist": record.get("dtw_dist"),
                    "hs_score": record.get("hs_score"),
                    "mid_ratio": record.get("mid_ratio"),
                    "mid_peer": record.get("mid_peer"),
                    "mid_v_ratio": record.get("mid_v_ratio"),
                    "mid_i_ratio": record.get("mid_i_ratio"),
                    "last_ratio": record.get("last_ratio"),
                    "last_peer": record.get("last_peer"),
                    "event_A": record.get("event_A"),
                    "v_drop": record.get("v_drop"),
                    "critical_fault": record.get("critical_fault"),
                    "critical_suspect": record.get("critical_suspect"),
                    "critical_confirmed": record.get("critical_confirmed"),
                    "fault_like_day": record.get("fault_like_day"),
                    "final_fault": record.get("final_fault"),
                    "ews_warning": record.get("ews_warning"),
                    "pre_alarm": record.get("pre_alarm"),
                    "pre_ews": record.get("pre_ews"),
                    "site_event_soft": record.get("site_event_soft"),
                    "site_event_hard": record.get("site_event_hard"),
                    "group_off_date": record.get("group_off_date"),
                    "prefault_B": record.get("prefault_B"),
                    "prefault_B_effective": record.get("prefault_B_effective"),
                    "prefault_B_common_cause_overlap": record.get("prefault_B_common_cause_overlap"),
                    "prefault_cond_mid": record.get("prefault_cond_mid"),
                    "prefault_cond_ae": record.get("prefault_cond_ae"),
                    "prefault_cond_dtw": record.get("prefault_cond_dtw"),
                    "prefault_cond_ews": record.get("prefault_cond_ews"),
                    "critical_source": normalize_text(record.get("critical_source")),
                    "anom_level": normalize_text(record.get("anom_level")),
                    "anom_subtype": normalize_text(record.get("anom_subtype")),
                }
            )

    evidence_df = pd.DataFrame(evidence_rows).sort_values(["site", "base", "panel_id"]).reset_index(drop=True)
    cluster_df = (
        evidence_df.groupby(["site", "base"], dropna=False)
        .agg(
            candidate_panels=("panel_id", "nunique"),
            확정_panel_count=("운영해석등급_ko", lambda s: int((s == "확정").sum())),
            고위험관찰_panel_count=("운영해석등급_ko", lambda s: int((s == "고위험 관찰").sum())),
            관찰_panel_count=("운영해석등급_ko", lambda s: int((s == "관찰").sum())),
            final_days_total=("final_days", "sum"),
            critical_days_total=("critical_days", "sum"),
            fault_like_days_total=("fault_like_days", "sum"),
            event_A_days_total=("event_A_days", "sum"),
            ews_warning_total=("ews_warning_days", "sum"),
            pre_ews_total=("pre_ews_days", "sum"),
            top1_candidates_csv=("1순위_의심원인_ko", lambda s: ",".join(sorted({normalize_text(v) for v in s if normalize_text(v)}))),
            labeled_fault_csv=("라벨된 fault", lambda s: ",".join(sorted({normalize_text(v) for v in s if normalize_text(v)}))),
        )
        .reset_index()
    )
    if not cluster_df.empty:
        cluster_df["군집해석_ko"] = cluster_df.apply(
            lambda row: (
                "군집 내 hard fault 포함"
                if int(row["확정_panel_count"]) > 0
                else "여러 패널이 함께 흔들려 공통 원인 가능성"
                if int(row["candidate_panels"]) >= 3
                else "소수 패널 관찰"
            ),
            axis=1,
        )

    heuristic_definition_rows = [
        {
            "항목": "1/2/3순위_의심원인_ko",
            "설명": "한국어 표시용 heuristic candidate 라벨이며, internal code를 대신하지 않는다. 라벨은 엔지니어 친화적으로 유지하고 쉬운 설명은 definitions에서 별도로 붙인다",
        },
        *[
            {
                "항목": display_name,
                "설명": display_heuristic_note(display_name),
            }
            for display_name in DISPLAY_HEURISTIC_NAME_MAP.values()
        ],
    ]

    definitions_df = pd.DataFrame(
        [
            {
                "항목": "definitions 시트",
                "설명": "상세 리포트 안에서 artifact 역할과 주요 컬럼 뜻을 짧게 설명하는 analyst/support glossary로, 읽기 순서나 auto-open 정책을 대신하지 않는다",
            },
            {
                "항목": "detailed report",
                "설명": "여러 row universe와 lineage를 함께 담는 analyst primary 문서로, current/master report를 대체하지 않는다",
            },
            {
                "항목": "official current",
                "설명": "frozen-support live chain 기준의 운영 공식 결과 묶음으로, detailed definitions에서는 역할과 공식성 차이만 짧게 설명한다",
            },
            {
                "항목": "raw_only current",
                "설명": "raw-only candidate 우주에서 strict current subset만 따로 보여주는 analyst/support 추가 자료로, official current를 대체하지 않는다",
            },
            {
                "항목": "운영해석등급_ko",
                "설명": "상세 리포트용 보조 등급으로 core verdict를 바꾸지 않고 사람이 읽기 쉽게 정리한 값",
            },
            {
                "항목": "확정",
                "설명": "최종 고장 신호 또는 강한 고장 신호가 존재하는 패널",
            },
            {
                "항목": "고위험 관찰",
                "설명": "즉시 확정에 쓰는 신호는 없지만 EWS, prefault_cond_ae/dtw/ews, fault_like 누적이 강한 패널",
            },
            {
                "항목": "관찰",
                "설명": "약한 이상 또는 간헐 이상으로 추가 추적이 필요한 패널",
            },
            {
                "항목": "precursor_report",
                "설명": "고장 신호가 아직 없는 precursor candidate만 따로 정리한 watchlist 성격의 보조표로, current artifact를 대체하지 않는다",
            },
            {
                "항목": "fault_signal_report",
                "설명": "raw-only candidate 우주에서 고장 신호가 이미 관측된 패널만 따로 정리한 analyst/support 보조표로, operator 기본 읽기 순서에는 직접 포함되지 않는다",
            },
            {
                "항목": "전조 축",
                "설명": "EWS/AE/DTW/규칙징후 중 어떤 축이 precursor candidate를 만들었는지 보여주는 묶음",
            },
            {
                "항목": "규칙징후",
                "설명": "pre_alarm, fault_like, 상대 전압 이탈 같은 규칙 기반 이상 징후를 완곡하게 묶은 표현",
            },
            {
                "항목": "Option B 유효 일수",
                "설명": "prefault_B 중 site_event/group_off 공통원인 겹침을 제외하고 실제 precursor 승격 설명에 반영한 일수",
            },
            {
                "항목": "공통원인 겹침 일수",
                "설명": "prefault_B가 켜졌지만 site_event/group_off와 직접 겹쳐 operator-facing precursor 승격에서는 별도 분리한 일수",
            },
            {
                "항목": "대표 전조 신호",
                "설명": "전조 표에서 누적된 핵심 신호를 짧게 요약한 값",
            },
            {
                "항목": "모니터링 권고",
                "설명": "precursor candidate에 대해 다음 수집 주기에서 무엇을 먼저 볼지 안내하는 운영 메모",
            },
            {
                "항목": "공통원인 위험",
                "설명": "site_event/group_off 겹침과 동일 subgroup 동시 흔들림을 바탕으로 panel-local precursor 해석을 얼마나 보수적으로 볼지 적은 보조 라벨",
            },
            {
                "항목": "권고 검토 레인",
                "설명": "일반 모니터링, 단일 패널 우선 추적, 공통원인 검토 중 다음 확인 방향을 짧게 정리한 값",
            },
            {
                "항목": "근접 공통원인",
                "설명": "raw-only 고장 신호 표에서 strict_trigger 기준 ±3일 안에 common-cause 이력이 같이 있으면 채우는 analyst/support 보조 값",
            },
            {
                "항목": "group root",
                "설명": "panel_id에서 마지막 두 서브인덱스를 제외한 넓은 family root로, 같은 상위 군집인지 보기 위한 값",
            },
            {
                "항목": "subgroup base",
                "설명": "panel_id에서 마지막 서브인덱스 하나만 제외한 하위 묶음으로, runtime common-cause 검토 단위에 더 가까운 값",
            },
            {
                "항목": "동일 subgroup row 수",
                "설명": "같은 raw-only current/fault-signal 우주에서 동일 subgroup base 아래 함께 잡힌 panel row 수로, row 수와 독립 사건 수를 혼동하지 않도록 돕는 값",
            },
            {
                "항목": "subgroup cluster",
                "설명": f"같은 subgroup base 안에서 신호 기준일 간격이 {FAULT_SIGNAL_CLUSTER_GAP_DAYS}일 이하인 row를 하나의 보조 cluster로 묶어 읽기 쉽게 만든 값",
            },
            {
                "항목": "동일 cluster row 수",
                "설명": "같은 subgroup cluster 안에 함께 들어간 panel row 수로, 대략적인 사건 뭉치를 읽기 쉽게 보조하는 값",
            },
            {
                "항목": "확정 경로",
                "설명": "raw-only 고장 신호 표에서 주된 고장 신호 경로 하나만 표시한 값",
            },
            {
                "항목": "고장 신호 요약",
                "설명": "고장 신호 지속 일수와 vdrop 같은 보조 근거를 함께 적은 요약",
            },
            {
                "항목": "현장 점검 권고",
                "설명": "raw-only 고장 신호 표에서 첫 현장 액션 우선순위를 짧게 적은 값",
            },
            {
                "항목": "strict_trigger_proximal_common_cause_flag",
                "설명": "raw-only audit에서 strict_trigger 기준 ±3일 안의 common-cause 이력을 잡는 내부 analyst flag",
            },
            {
                "항목": "warning_proximal_common_cause_flag",
                "설명": "raw-only audit에서 earliest_warning 기준 ±3일 안의 common-cause 이력을 잡는 내부 analyst flag로, 현재는 audit 전용",
            },
            {
                "항목": "raw_only_chain 주의",
                "설명": "raw-only candidate chain은 current/frozen 공식 결과보다 넓은 후보 우주를 보여주며, official current를 대체하지 않는다",
            },
            *heuristic_definition_rows,
        ]
    )

    frames["raw_only_evidence"] = nonempty_sheet_df(
        evidence_df,
        "raw-only evidence rows unavailable",
    )
    frames["raw_only_candidate_scores"] = nonempty_sheet_df(
        pd.DataFrame(candidate_score_rows).sort_values(["site", "base", "panel_id", "후보순위"]).reset_index(drop=True),
        "raw-only candidate score matrix unavailable",
    )
    frames["raw_only_timeline"] = nonempty_sheet_df(
        pd.DataFrame(timeline_rows).sort_values(["site", "panel_id", "date"]).reset_index(drop=True),
        "raw-only timeline rows unavailable",
    )
    frames["raw_only_daily_log"] = nonempty_sheet_df(
        pd.DataFrame(all_date_rows).sort_values(["site", "panel_id", "date"]).reset_index(drop=True),
        "raw-only all-date log unavailable",
    )
    frames["raw_only_cluster"] = nonempty_sheet_df(
        cluster_df,
        "raw-only cluster summary unavailable",
    )
    frames["precursor_report"] = build_precursor_report_df(evidence_df)
    frames["fault_signal_report"] = build_fault_signal_report_df(evidence_df)
    frames["definitions"] = definitions_df
    return frames


def write_detailed_report_xlsx(path: Path, frames: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    auto_fit_workbook_columns(path)


def stage_live_chain_workspace(output_root: Path, sites: list[str]) -> Path:
    workspace_root = output_root / "live_chain_workspace"
    if workspace_root.exists():
        shutil.rmtree(workspace_root)
    copy_tree(packaged_share_root(), workspace_root / "_share")
    for site in sites:
        copy_tree(output_root / "sites" / site / "output", workspace_root / "data" / site / "out")
    return workspace_root


def stage_raw_only_chain_workspace(output_root: Path, sites: list[str]) -> Path:
    workspace_root = output_root / "raw_only_chain_workspace"
    if workspace_root.exists():
        shutil.rmtree(workspace_root)
    for site in sites:
        copy_tree(output_root / "sites" / site / "output", workspace_root / "data" / site / "out")
    return workspace_root


def run_live_chain(output_root: Path, sites: list[str], baseline_comparison: dict[str, object]) -> dict[str, object]:
    support = packaged_live_chain_support()
    result_dir = output_root / "result" / "live_chain"
    result_dir.mkdir(parents=True, exist_ok=True)
    payload: dict[str, object] = {
        "requested": True,
        "supported": bool(support["supported"]),
        "support": support,
        "workspace_root": "",
        "result_dir": str(result_dir),
        "status_ko": "",
        "generated_outputs": {},
        "fixed_fault_reference_compare": {},
        "note_ko": (
            "live chain은 package 내부의 bootstrap verdict -> fault_event_audit -> final verdict -> gpvs evidence -> heuristic "
            "경로를 workspace-only로 수행한다."
        ),
    }
    if not support["supported"]:
        payload["status_ko"] = "packaged live chain assets missing"
        return payload
    if sorted(sites) != sorted(DEFAULT_SITES):
        payload["status_ko"] = "current live chain supports baseline tri-site universe only"
        return payload

    workspace_root = stage_live_chain_workspace(output_root, sites)
    payload["workspace_root"] = str(workspace_root)
    commands = [
        [sys.executable, str(packaged_script_path("build_panel_day_engine_bootstrap_verdict_v1.py")), "--root", str(workspace_root), "--write-panel-verdict-alias"],
        [sys.executable, str(packaged_script_path("build_panel_day_engine_fault_panel_event_audit_v1.py")), "--root", str(workspace_root)],
        [sys.executable, str(packaged_script_path("build_panel_day_engine_panel_multiaxis_verdict_v1.py")), "--root", str(workspace_root)],
        [sys.executable, str(packaged_script_path("build_panel_day_engine_gpvs_evidence_pack_v1.py")), "--root", str(workspace_root)],
        [sys.executable, str(packaged_script_path("build_panel_day_engine_cause_candidate_heuristics_v1.py")), "--root", str(workspace_root)],
    ]
    for cmd in commands:
        subprocess.run(cmd, cwd=package_root(), check=True)

    live_fault_df = build_live_fault_table(workspace_root)
    live_preview_df = build_live_fault_preview(workspace_root, live_fault_df)
    live_fault_path = result_dir / "fault_panel_result_live_v1.csv"
    live_preview_path = result_dir / "fault_panel_result_live_preview_v1.csv"
    live_fault_df.to_csv(live_fault_path, index=False, encoding="utf-8-sig")
    live_preview_df.to_csv(live_preview_path, index=False, encoding="utf-8-sig")

    generated = {
        "bootstrap_verdict": str(workspace_root / "_share" / "panel_day_engine_bootstrap_verdict_v1.csv"),
        "fault_event_audit": str(workspace_root / "_share" / "panel_day_engine_fault_panel_event_audit_v1.csv"),
        "final_verdict": str(workspace_root / "_share" / "panel_day_engine_panel_multiaxis_verdict_v1.csv"),
        "gpvs_evidence": str(workspace_root / "_share" / "panel_day_engine_gpvs_evidence_pack_v1.csv"),
        "heuristic": str(workspace_root / "_share" / "panel_day_engine_cause_candidate_heuristics_v1.csv"),
        "fault_panel_result_live_v1": str(live_fault_path),
        "fault_panel_result_live_preview_v1": str(live_preview_path),
    }
    for name, source in [
        ("panel_day_engine_bootstrap_verdict_v1.csv", workspace_root / "_share" / "panel_day_engine_bootstrap_verdict_v1.csv"),
        ("panel_day_engine_fault_panel_event_audit_v1.csv", workspace_root / "_share" / "panel_day_engine_fault_panel_event_audit_v1.csv"),
        ("panel_day_engine_panel_multiaxis_verdict_v1.csv", workspace_root / "_share" / "panel_day_engine_panel_multiaxis_verdict_v1.csv"),
        ("panel_day_engine_gpvs_evidence_pack_v1.csv", workspace_root / "_share" / "panel_day_engine_gpvs_evidence_pack_v1.csv"),
        ("panel_day_engine_cause_candidate_heuristics_v1.csv", workspace_root / "_share" / "panel_day_engine_cause_candidate_heuristics_v1.csv"),
    ]:
        target = result_dir / name
        shutil.copy2(source, target)
        generated[name] = str(target)

    compare = compare_live_fault_to_fixed(live_fault_df)
    compare["baseline_input_all_sites_match"] = bool(baseline_comparison.get("all_sites_match", False))
    payload["generated_outputs"] = generated
    payload["fixed_fault_reference_compare"] = compare
    payload["status_ko"] = "completed"
    summary_path = result_dir / "live_chain_summary_v1.json"
    write_json(summary_path, payload)
    payload["summary_path"] = str(summary_path)
    payload["published_outputs"] = publish_live_chain_outputs(output_root, result_dir, summary_path)
    write_json(summary_path, payload)
    root_summary_path = output_root / "result" / ROOT_LIVE_SUMMARY_NAME
    shutil.copy2(summary_path, root_summary_path)
    payload["published_outputs"][ROOT_LIVE_SUMMARY_NAME] = str(root_summary_path)
    root_report_path = output_root / "result" / ROOT_LIVE_REPORT_NAME
    write_text(
        root_report_path,
        build_live_report_markdown(
            sites=sites,
            baseline_comparison=baseline_comparison,
            compare=compare,
            published_outputs=payload["published_outputs"],
            live_preview_df=live_preview_df,
        ),
    )
    payload["published_outputs"][ROOT_LIVE_REPORT_NAME] = str(root_report_path)
    write_json(summary_path, payload)
    shutil.copy2(summary_path, root_summary_path)
    return payload


def run_raw_only_chain(output_root: Path, sites: list[str]) -> dict[str, object]:
    support = packaged_raw_only_chain_support()
    result_dir = output_root / "result" / "raw_only_chain"
    result_dir.mkdir(parents=True, exist_ok=True)
    payload: dict[str, object] = {
        "requested": True,
        "supported": bool(support["supported"]),
        "support": support,
        "workspace_root": "",
        "result_dir": str(result_dir),
        "status_ko": "",
        "generated_outputs": {},
        "fixed_fault_reference_compare": {},
        "note_ko": (
            "raw-only chain은 panel_day_core와 precursor gate만 사용해 audit -> final verdict -> heuristic를 다시 계산한다. "
            "커널로그_원인군_ko는 algorithm-derived family 의미로 해석해야 한다."
        ),
    }
    if not support["supported"]:
        payload["status_ko"] = "packaged raw-only chain assets missing"
        return payload

    workspace_root = stage_raw_only_chain_workspace(output_root, sites)
    payload["workspace_root"] = str(workspace_root)
    commands = [
        [sys.executable, str(packaged_script_path("build_panel_day_engine_runtime_fault_event_audit_v1.py")), "--root", str(workspace_root)],
        [sys.executable, str(packaged_script_path("build_panel_day_engine_runtime_final_verdict_v1.py")), "--root", str(workspace_root)],
        [sys.executable, str(packaged_script_path("build_panel_day_engine_runtime_heuristic_v1.py")), "--root", str(workspace_root)],
    ]
    for cmd in commands:
        subprocess.run(cmd, cwd=package_root(), check=True)

    raw_only_common = load_raw_only_common_module()
    raw_only_fault_df = raw_only_common.build_fault_table_from_outputs(
        workspace_root=workspace_root,
        verdict_name=raw_only_common.RUNTIME_VERDICT_OUTPUT_NAME,
        heuristic_name=raw_only_common.RUNTIME_HEURISTIC_OUTPUT_NAME,
    )
    raw_only_preview_df = raw_only_common.build_fault_preview(workspace_root, raw_only_fault_df)
    raw_only_fault_path = result_dir / "fault_panel_result_raw_only_v1.csv"
    raw_only_preview_path = result_dir / "fault_panel_result_raw_only_preview_v1.csv"
    raw_only_fault_df.to_csv(raw_only_fault_path, index=False, encoding="utf-8-sig")
    raw_only_preview_df.to_csv(raw_only_preview_path, index=False, encoding="utf-8-sig")

    generated = {
        "runtime_audit": str(workspace_root / "_share" / raw_only_common.RUNTIME_AUDIT_OUTPUT_NAME),
        "runtime_verdict": str(workspace_root / "_share" / raw_only_common.RUNTIME_VERDICT_OUTPUT_NAME),
        "runtime_heuristic": str(workspace_root / "_share" / raw_only_common.RUNTIME_HEURISTIC_OUTPUT_NAME),
        "fault_panel_result_raw_only_v1": str(raw_only_fault_path),
        "fault_panel_result_raw_only_preview_v1": str(raw_only_preview_path),
    }
    for name, source in [
        (raw_only_common.RUNTIME_AUDIT_OUTPUT_NAME, workspace_root / "_share" / raw_only_common.RUNTIME_AUDIT_OUTPUT_NAME),
        (raw_only_common.RUNTIME_VERDICT_OUTPUT_NAME, workspace_root / "_share" / raw_only_common.RUNTIME_VERDICT_OUTPUT_NAME),
        (raw_only_common.RUNTIME_HEURISTIC_OUTPUT_NAME, workspace_root / "_share" / raw_only_common.RUNTIME_HEURISTIC_OUTPUT_NAME),
    ]:
        target = result_dir / name
        shutil.copy2(source, target)
        generated[name] = str(target)

    compare = raw_only_common.compare_fault_table_to_reference(raw_only_fault_df, fixed_fault6_table_path())
    payload["generated_outputs"] = generated
    payload["fixed_fault_reference_compare"] = compare
    payload["status_ko"] = "completed"
    summary_path = result_dir / "raw_only_chain_summary_v1.json"
    write_json(summary_path, payload)
    payload["summary_path"] = str(summary_path)
    return payload


def build_shadow_compare_report(
    output_root: Path,
    site_plans: list[dict[str, object]],
    baseline_comparison: dict[str, object],
) -> dict[str, object]:
    reference = load_core_baseline_digest()
    report: dict[str, object] = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "reference_path": str(baseline_core_digest_path()),
        "note_ko": (
            "이 shadow compare는 동일 baseline raw corpus로 runtime pack을 다시 실행했을 때 "
            "panel_day_core.csv가 reference digest와 같은지 점검한다. "
            "현재는 full-chain verdict/evidence/heuristic live compare가 아니라 engine core compare만 수행한다."
        ),
        "sites": {},
        "compared_site_count": 0,
        "matched_site_count": 0,
        "all_compared_sites_match": True,
    }
    reference_sites = reference.get("sites", {})

    for plan in site_plans:
        site = str(plan["site"])
        site_entry: dict[str, object] = {
            "baseline_input_match": bool(baseline_comparison["sites"].get(site, {}).get("match", False)),
            "compared": False,
            "match": None,
            "skipped_reason": "",
            "expected": {},
            "actual": {},
            "diffs": [],
        }
        expected = reference_sites.get(site)
        if not expected:
            site_entry["skipped_reason"] = "missing_packaged_reference_digest"
            report["sites"][site] = site_entry
            continue
        if not site_entry["baseline_input_match"]:
            site_entry["skipped_reason"] = "input_manifest_mismatch"
            site_entry["expected"] = expected
            report["sites"][site] = site_entry
            continue

        core_path = output_root / "sites" / site / "output" / "panel_day_core.csv"
        if not core_path.exists():
            site_entry["skipped_reason"] = "missing_generated_panel_day_core"
            site_entry["expected"] = expected
            report["sites"][site] = site_entry
            continue

        actual_df = pd.read_csv(core_path, low_memory=False)
        actual_digest = build_core_digest_payload(actual_df, core_path.name)
        diffs = compare_single_site_digest(expected, actual_digest)
        site_entry.update(
            {
                "compared": True,
                "match": not diffs,
                "expected": expected,
                "actual": actual_digest,
                "diffs": diffs,
            }
        )
        report["sites"][site] = site_entry
        report["compared_site_count"] += 1
        if not diffs:
            report["matched_site_count"] += 1
        else:
            report["all_compared_sites_match"] = False

    if report["compared_site_count"] == 0:
        report["all_compared_sites_match"] = False
    return report


def main() -> None:
    args = parse_args()
    if not engine_path().exists():
        raise SystemExit(f"missing packaged engine: {engine_path()}")

    emit_progress(1, "실행 준비를 시작합니다.")
    data_root = args.data_root.expanduser().resolve()
    output_root = args.output_root.expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    reuse_existing_site_outs_root = (
        args.reuse_existing_site_outs_root.expanduser().resolve()
        if args.reuse_existing_site_outs_root is not None
        else None
    )

    sites = normalize_sites(args.sites)
    effective_reuse_existing_site_outs_root, reuse_decision, reuse_freshness = resolve_reuse_existing_site_outs_root(
        data_root=data_root,
        explicit_reuse_root=reuse_existing_site_outs_root,
        prefer_existing_site_outs=args.prefer_existing_site_outs,
        sites=sites,
    )
    site_plans: list[dict[str, object]] = []
    commands: list[list[str]] = []
    for site in sites:
        plan, cmd = build_site_plan(args, site)
        site_plans.append(plan)
        commands.append(cmd)

    emit_progress(8, "입력 CSV 구조와 실행 계획을 점검했습니다.")
    fixed_outputs = copy_fixed_results(output_root)
    baseline_comparison = compare_to_baseline(site_plans)
    live_chain_support = packaged_live_chain_support()
    raw_only_chain_support = packaged_raw_only_chain_support()
    live_chain_plan = {
        "requested": args.run_live_chain == "on",
        "supported": bool(live_chain_support["supported"]),
        "support": live_chain_support,
        "status_ko": "",
    }
    if not live_chain_plan["requested"]:
        live_chain_plan["status_ko"] = "disabled by option"
    elif sorted(sites) != sorted(DEFAULT_SITES):
        live_chain_plan["status_ko"] = "current live chain supports baseline tri-site universe only"
    elif not live_chain_plan["supported"]:
        live_chain_plan["status_ko"] = "packaged live chain assets missing"
    else:
        live_chain_plan["status_ko"] = (
            "will run after precomputed out reuse"
            if effective_reuse_existing_site_outs_root is not None
            else "will run after engine execution"
        )
    raw_only_chain_plan = {
        "requested": args.run_raw_only_chain == "on",
        "supported": bool(raw_only_chain_support["supported"]),
        "support": raw_only_chain_support,
        "status_ko": "",
    }
    if not raw_only_chain_plan["requested"]:
        raw_only_chain_plan["status_ko"] = "disabled by option"
    elif not raw_only_chain_plan["supported"]:
        raw_only_chain_plan["status_ko"] = "packaged raw-only chain assets missing"
    else:
        raw_only_chain_plan["status_ko"] = (
            "will run after precomputed out reuse"
            if effective_reuse_existing_site_outs_root is not None
            else "will run after engine execution"
        )

    plan = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "package_root": str(package_root()),
        "engine_path": str(engine_path()),
        "data_root": str(data_root),
        "output_root": str(output_root),
        "reuse_existing_site_outs_root": str(reuse_existing_site_outs_root) if reuse_existing_site_outs_root else "",
        "effective_reuse_existing_site_outs_root": (
            str(effective_reuse_existing_site_outs_root)
            if effective_reuse_existing_site_outs_root
            else ""
        ),
        "prefer_existing_site_outs": args.prefer_existing_site_outs,
        "reuse_decision_ko": reuse_decision,
        "reuse_freshness": reuse_freshness,
        "execution_mode_ko": (
            "auto_reuse_existing_site_outs"
            if effective_reuse_existing_site_outs_root is not None and reuse_decision == "auto_fresh"
            else "reuse_precomputed_site_outs"
            if effective_reuse_existing_site_outs_root is not None
            else "run_engine_then_live_chain"
        ),
        "sites": sites,
        "site_plans": site_plans,
        "fixed_outputs": fixed_outputs,
        "baseline_comparison": baseline_comparison,
        "live_chain": live_chain_plan,
        "raw_only_chain": raw_only_chain_plan,
        "workspace_retention": args.workspace_retention,
        "shadow_compare_reference_path": str(baseline_core_digest_path()),
        "fault6_provenance_path": str(fault6_provenance_path()),
        "dependency_audit_json_path": str(dependency_audit_json_path()),
        "dependency_audit_md_path": str(dependency_audit_md_path()),
        "shadow_compare_report_path": str(output_root / "shadow_compare_v1.json"),
        "workspace_cleanup_report_path": str(output_root / "workspace_cleanup_v1.json"),
        "dry_run": bool(args.dry_run),
        "note_ko": (
            "이 pack은 conalog/gangui/ktc_ess baseline sites에 대해 실제 panel_day_engine.py 를 실행하고, "
            "현재 frozen fault 결과표도 함께 export 한다. "
            "fault6 결과표 provenance도 함께 남겨, 이 표가 frozen verdict+heuristic direct assembly인지 확인할 수 있다. "
            "추가로 baseline core output shadow compare 경로를 남겨, same baseline 입력일 때 engine core output이 유지되는지도 점검한다. "
            "workspace_retention=result-only를 사용하면 재생성 가능한 대용량 site/workspace data copy를 실행 후 제거한다."
        ),
    }

    if args.dry_run:
        write_json(output_root / "run_plan_v1.json", plan)
        emit_progress(100, "dry-run 계획 파일 생성을 완료했습니다.")
        print(f"[OK] dry-run plan written: {output_root / 'run_plan_v1.json'}")
        return

    reused_site_outs: dict[str, str] = {}
    if effective_reuse_existing_site_outs_root is not None:
        emit_progress(15, "기존 site out 산출물을 재사용합니다.")
        reused_site_outs = copy_existing_site_outs(effective_reuse_existing_site_outs_root, output_root, sites)
    else:
        site_count = max(1, len(commands))
        for idx, cmd in enumerate(commands, start=1):
            site_name = str(site_plans[idx - 1]["site"])
            start_pct = 15 + int((idx - 1) * 45 / site_count)
            done_pct = 15 + int(idx * 45 / site_count)
            emit_progress(start_pct, f"메인 엔진 실행 시작: {site_name}")
            subprocess.run(cmd, check=True)
            emit_progress(done_pct, f"메인 엔진 실행 완료: {site_name}")

    emit_progress(65, "engine core 결과를 shadow compare 기준으로 점검합니다.")
    shadow_compare = build_shadow_compare_report(output_root, site_plans, baseline_comparison)
    write_json(output_root / "shadow_compare_v1.json", shadow_compare)
    live_chain_result = {"requested": False, "status_ko": "not requested"}
    if args.run_live_chain == "on":
        emit_progress(75, "live chain 결과표를 생성합니다.")
        live_chain_result = run_live_chain(output_root, sites, baseline_comparison)
    raw_only_chain_result = {"requested": False, "status_ko": "not requested"}
    if args.run_raw_only_chain == "on":
        emit_progress(88, "raw-only candidate chain 결과를 생성합니다.")
        raw_only_chain_result = run_raw_only_chain(output_root, sites)

    master_report_path = output_root / "result" / ROOT_MASTER_REPORT_NAME
    detailed_report_path = output_root / "result" / ROOT_DETAILED_REPORT_NAME
    precursor_report_path = output_root / "result" / ROOT_PRECURSOR_REPORT_NAME
    fault_signal_report_path = output_root / "result" / ROOT_FAULT_SIGNAL_REPORT_NAME
    live_preview_path = output_root / "result" / ROOT_LIVE_PREVIEW_NAME
    live_preview_df = pd.read_csv(live_preview_path, encoding="utf-8-sig", low_memory=False) if live_preview_path.exists() else pd.DataFrame()
    raw_only_candidate_preview_path = output_root / "result" / "raw_only_chain" / "fault_panel_result_raw_only_preview_v1.csv"
    raw_only_candidate_preview_df = (
        pd.read_csv(raw_only_candidate_preview_path, encoding="utf-8-sig", low_memory=False)
        if raw_only_candidate_preview_path.exists()
        else pd.DataFrame()
    )
    detailed_frames = build_detailed_report_frames(
        output_root=output_root,
        sites=sites,
        baseline_comparison=baseline_comparison,
        live_chain_result=live_chain_result,
        raw_only_chain_result=raw_only_chain_result,
        live_preview_df=live_preview_df,
        raw_only_preview_df=raw_only_candidate_preview_df,
    )
    raw_only_current_preview_df = raw_only_candidate_preview_df.copy()
    if raw_only_chain_result.get("requested") and normalize_text(raw_only_chain_result.get("status_ko")) == "completed":
        strict_fault_df, strict_preview_df, publish_meta = build_strict_raw_only_current_outputs(
            raw_only_chain_result=raw_only_chain_result,
            evidence_df=detailed_frames["raw_only_evidence"],
        )
        raw_only_chain_result["publish_meta"] = publish_meta
        raw_only_chain_result["published_outputs"] = publish_raw_only_current_outputs(
            output_root,
            strict_fault_df,
            strict_preview_df,
        )
        root_summary_path = output_root / "result" / ROOT_RAWONLY_SUMMARY_NAME
        raw_only_chain_result["published_outputs"][ROOT_RAWONLY_SUMMARY_NAME] = str(root_summary_path)
        root_report_path = output_root / "result" / ROOT_RAWONLY_REPORT_NAME
        write_text(
            root_report_path,
            build_raw_only_report_markdown(
                sites=sites,
                compare=raw_only_chain_result.get("fixed_fault_reference_compare", {}),
                published_outputs=raw_only_chain_result["published_outputs"],
                live_preview_df=to_user_preview_schema(strict_preview_df),
                publish_meta=publish_meta,
            ),
        )
        raw_only_chain_result["published_outputs"][ROOT_RAWONLY_REPORT_NAME] = str(root_report_path)
        raw_only_current_preview_df = strict_preview_df.copy()
        summary_path = Path(str(raw_only_chain_result.get("summary_path", "")))
        if summary_path.exists():
            write_json(summary_path, raw_only_chain_result)
            shutil.copy2(summary_path, root_summary_path)
    live_preview_display_df = to_user_preview_schema(live_preview_df)
    raw_only_current_preview_display_df = to_user_preview_schema(raw_only_current_preview_df)
    detailed_frames["current_preview"] = nonempty_sheet_df(
        live_preview_display_df.copy(),
        "live current preview not available",
    )
    detailed_frames["raw_only_preview"] = nonempty_sheet_df(
        raw_only_current_preview_display_df.copy(),
        "raw-only preview not available",
    )
    precursor_report_df = detailed_frames.get(
        "precursor_report",
        pd.DataFrame(columns=PRECURSOR_REPORT_OUTPUT_COLS),
    )
    fault_signal_report_df = detailed_frames.get(
        "fault_signal_report",
        pd.DataFrame(columns=FAULT_SIGNAL_REPORT_OUTPUT_COLS),
    )
    precursor_report_df.to_csv(precursor_report_path, index=False, encoding="utf-8-sig")
    fault_signal_report_df.to_csv(fault_signal_report_path, index=False, encoding="utf-8-sig")
    write_detailed_report_xlsx(
        detailed_report_path,
        detailed_frames,
    )
    write_text(
        master_report_path,
        build_master_report_markdown(
            sites=sites,
            baseline_comparison=baseline_comparison,
            live_chain_result=live_chain_result,
            raw_only_chain_result=raw_only_chain_result,
            live_preview_df=live_preview_display_df,
            raw_only_preview_df=raw_only_current_preview_display_df,
            precursor_report_df=precursor_report_df,
            fault_signal_report_df=fault_signal_report_df,
            detailed_report_path=detailed_report_path,
            precursor_report_path=precursor_report_path,
            fault_signal_report_path=fault_signal_report_path,
        ),
    )
    if live_preview_path.exists():
        live_preview_display_df.to_csv(live_preview_path, index=False, encoding="utf-8-sig")
    raw_only_current_preview_path = output_root / "result" / ROOT_RAWONLY_PREVIEW_NAME
    if raw_only_current_preview_path.exists():
        raw_only_current_preview_display_df.to_csv(
            raw_only_current_preview_path,
            index=False,
            encoding="utf-8-sig",
        )

    workspace_cleanup = apply_workspace_retention(output_root, args.workspace_retention)
    workspace_cleanup_path = output_root / "workspace_cleanup_v1.json"
    write_json(workspace_cleanup_path, workspace_cleanup)
    metadata = {
        **plan,
        "dry_run": False,
        "reused_site_outs": reused_site_outs,
        "shadow_compare": shadow_compare,
        "live_chain": live_chain_result,
        "raw_only_chain": raw_only_chain_result,
        "workspace_cleanup": workspace_cleanup,
        "detailed_report_path": str(detailed_report_path),
        "precursor_report_path": str(precursor_report_path),
        "fault_signal_report_path": str(fault_signal_report_path),
        "master_report_path": str(master_report_path),
    }
    write_json(output_root / "run_metadata_v1.json", metadata)
    emit_progress(100, "실행 완료. 결과 리포트를 열 수 있습니다.")
    print(f"[OK] result dir: {output_root / 'result'}")
    print(f"[OK] shadow compare: {output_root / 'shadow_compare_v1.json'}")
    print(f"[OK] detailed report: {detailed_report_path}")
    print(f"[OK] precursor report: {precursor_report_path}")
    print(f"[OK] raw-only fault signal report: {fault_signal_report_path}")
    print(f"[OK] master report: {master_report_path}")
    print(f"[OK] workspace cleanup: {workspace_cleanup_path}")
    if live_chain_result.get("requested"):
        print(f"[OK] live chain status: {live_chain_result.get('status_ko')}")
    if raw_only_chain_result.get("requested"):
        print(f"[OK] raw-only chain status: {raw_only_chain_result.get('status_ko')}")
    if args.workspace_retention == "full":
        for site in sites:
            print(f"[OK] site output: {output_root / 'sites' / site / 'output' / 'panel_day_core.csv'}")
    else:
        removed_bytes = int(workspace_cleanup.get("bytes_removed_estimate", 0))
        print(f"[OK] workspace retention: result-only removed approximately {removed_bytes} bytes")


if __name__ == "__main__":
    main()
