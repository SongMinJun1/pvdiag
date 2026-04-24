#!/usr/bin/env python3
from __future__ import annotations

import atexit
from contextlib import ExitStack
import json
import py_compile
import subprocess
import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

if __package__ in {None, ""}:
    if str(Path(__file__).resolve().parents[2]) not in sys.path:
        sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from research.prognostics.smoke_frozen_share_fixture_v1 import (
        stage_missing_repo_data_link,
        stage_missing_share_fixtures,
    )
else:
    from .smoke_frozen_share_fixture_v1 import (
        stage_missing_repo_data_link,
        stage_missing_share_fixtures,
    )


REPO_ROOT = Path(__file__).resolve().parents[2]
if __package__ in {None, ""}:
    if str(REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(REPO_ROOT))
    from research.prognostics.build_conalog_full_runtime_pack_v1 import (
        OPTIONAL_PACKAGED_SHARE_INPUTS,
        PACKAGED_RUNTIME_CHAIN_SHARE_INPUTS,
    )
    from research.prognostics.heuristic_display_registry_v1 import (
        DISPLAY_HEURISTIC_NAME_MAP,
        HEURISTIC_DISPLAY_NOTE_MAP,
        LEGACY_HEURISTIC_DISPLAY_NAMES,
        contains_legacy_heuristic_display_name,
    )
else:
    from .build_conalog_full_runtime_pack_v1 import (
        OPTIONAL_PACKAGED_SHARE_INPUTS,
        PACKAGED_RUNTIME_CHAIN_SHARE_INPUTS,
    )
    from .heuristic_display_registry_v1 import (
        DISPLAY_HEURISTIC_NAME_MAP,
        HEURISTIC_DISPLAY_NOTE_MAP,
        LEGACY_HEURISTIC_DISPLAY_NAMES,
        contains_legacy_heuristic_display_name,
    )

BUILD_SCRIPT = REPO_ROOT / "research" / "prognostics" / "build_conalog_full_runtime_pack_v1.py"
RELEASE_ROOT = REPO_ROOT / "release" / "conalog_full_runtime_v1"
PACKAGE_ROOT = RELEASE_ROOT / "package"
RUNNER = PACKAGE_ROOT / "app" / "run_full_algorithm_pack.py"
IMPORT_HELPER = PACKAGE_ROOT / "app" / "import_any_csv_root.py"
ENGINE = PACKAGE_ROOT / "pv_ae" / "panel_day_engine.py"
FAULT6_TABLE = PACKAGE_ROOT / "artifacts" / "fault6_fixed_result_table_v1.csv"
FAULT6_PREVIEW = PACKAGE_ROOT / "artifacts" / "fault6_label_and_algorithm_preview_v1.csv"
FAULT6_PROVENANCE = PACKAGE_ROOT / "artifacts" / "fault6_fixed_result_provenance_v1.json"
BASELINE_MANIFEST = PACKAGE_ROOT / "artifacts" / "input_baseline_manifest_v1.json"
CORE_BASELINE_DIGEST = PACKAGE_ROOT / "artifacts" / "panel_day_core_baseline_digest_v1.json"
DEPENDENCY_AUDIT_JSON = PACKAGE_ROOT / "artifacts" / "runtime_chain_dependency_audit_v1.json"
DEPENDENCY_AUDIT_MD = PACKAGE_ROOT / "artifacts" / "runtime_chain_dependency_audit_v1.md"
PACKAGE_BOOTSTRAP_VERDICT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_bootstrap_verdict_v1.py"
PACKAGE_AUDIT_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_fault_panel_event_audit_v1.py"
PACKAGE_VERDICT_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_panel_multiaxis_verdict_v1.py"
PACKAGE_GPVS_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_gpvs_evidence_pack_v1.py"
PACKAGE_HEURISTIC_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_cause_candidate_heuristics_v1.py"
PACKAGE_RAWONLY_COMMON = PACKAGE_ROOT / "research" / "prognostics" / "runtime_rawonly_chain_common_v1.py"
PACKAGE_RAWONLY_AUDIT_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_runtime_fault_event_audit_v1.py"
PACKAGE_RAWONLY_VERDICT_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_runtime_final_verdict_v1.py"
PACKAGE_RAWONLY_HEURISTIC_SCRIPT = PACKAGE_ROOT / "research" / "prognostics" / "build_panel_day_engine_runtime_heuristic_v1.py"
PACKAGE_SHARE_ROOT = PACKAGE_ROOT / "_share"
OUT_WATCH = REPO_ROOT / "_share" / "panel_day_engine_integrated_result_table_v1.csv"
FROZEN_VERDICT = REPO_ROOT / "_share" / "panel_day_engine_panel_multiaxis_verdict_v1.csv"
FROZEN_HEURISTIC = REPO_ROOT / "_share" / "panel_day_engine_cause_candidate_heuristics_v1.csv"
SUMMARY = RELEASE_ROOT / "pack_summary_v1.json"
STAGING_PS1 = PACKAGE_ROOT / "bin" / "stage_recent_120d.ps1"
SNAPSHOT_COPY_PS1 = PACKAGE_ROOT / "bin" / "snapshot_copy.ps1"
DAILY_RUN_BAT = PACKAGE_ROOT / "bin" / "daily_run.bat"
INCREMENTAL_RUN_BAT = PACKAGE_ROOT / "bin" / "incremental_run.bat"
RUN_DEMO_BAT = PACKAGE_ROOT / "bin" / "run_demo.bat"
RUN_GUIDED_REAL_BAT = PACKAGE_ROOT / "bin" / "run_guided_real.bat"
RUN_IMPORTED_REAL_BAT = PACKAGE_ROOT / "bin" / "run_imported_real.bat"
RESOLVE_PYTHON_BAT = PACKAGE_ROOT / "bin" / "resolve_python.bat"
WINDOWS_RUNTIME_MANIFEST = PACKAGE_ROOT / "runtime" / "windows_x64" / "runtime_manifest_v1.json"
WINDOWS_RUNTIME_PYTHON_EXE = PACKAGE_ROOT / "runtime" / "windows_x64" / "python" / "python.exe"
README_PATH = RELEASE_ROOT / "README.md"

EXPECTED_FAULT6_COLS = [
    "site",
    "panel_id",
    "패널고장여부_ko",
    "사건유형_ko",
    "최종고장양상_ko",
    "커널로그_원인군_ko",
    "1순위_의심원인_ko",
    "2순위_의심원인_ko",
    "3순위_의심원인_ko",
]
EXPECTED_PREVIEW_COLS = [
    "site",
    "panel_id",
    "전조날짜",
    "고장 기준일",
    "운영 판정",
    "급락 종결 관측",
    "점진 저하 누적",
    "사건 종결 요약",
    "상위 해석 후보",
    "기존 알고리즘 source",
]
EXPECTED_PREVIEW_DATE_ROWS = {
    ("conalog", "7f7dd654-2760-4eb2-a197-3ebb72b85cda.2.0"): ("2024-11-06", "2024-11-26"),
    ("conalog", "c42997a6-5881-47e7-9035-7de8a2673b54.1.1"): ("2025-01-20", "2025-03-21"),
    ("gangui", "bf1a912f-6cf0-4f12-8e97-9d9d86576511.0.7"): ("전조없음", "2025-06-08"),
    ("gangui", "bf1a912f-6cf0-4f12-8e97-9d9d86576511.2.16"): ("전조없음", "2025-06-08"),
    ("ktc_ess", "10305b40-b67e-40d1-9cd1-271b6642a3d9.2.12"): ("전조없음", "2025-08-16"),
    ("ktc_ess", "70ad2d87-cdb6-4842-81b7-71c7599bbf05.1.4"): ("2025-01-25", "2025-02-02"),
}
REQUIRED_SMOKE_SHARE_FIXTURES = sorted(
    (set(PACKAGED_RUNTIME_CHAIN_SHARE_INPUTS) - set(OPTIONAL_PACKAGED_SHARE_INPUTS))
    | {
        "panel_day_engine_cause_candidate_heuristics_v1.csv",
        "panel_day_engine_fault_panel_event_audit_v1.csv",
        "panel_day_engine_integrated_result_table_v1.csv",
    }
)

def make_dummy_site(tmp_root: Path, site: str) -> None:
    raw_dir = tmp_root / site / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    for day in range(1, 7):
        (raw_dir / f"sample_2024-01-0{day}.csv").write_text("dummy\n", encoding="utf-8")


def main() -> None:
    py_compile.compile(str(REPO_ROOT / "pv_ae" / "panel_day_engine.py"), doraise=True)
    py_compile.compile(str(BUILD_SCRIPT), doraise=True)
    with stage_missing_repo_data_link(REPO_ROOT):
        with stage_missing_share_fixtures(
            REPO_ROOT,
            REQUIRED_SMOKE_SHARE_FIXTURES,
        ):
            before_watch = OUT_WATCH.read_bytes()

            missing_display_notes = [
                label
                for label in DISPLAY_HEURISTIC_NAME_MAP.values()
                if not str(HEURISTIC_DISPLAY_NOTE_MAP.get(label, "")).strip()
            ]
            if missing_display_notes:
                raise SystemExit(
                    "shared heuristic display registry is missing short notes for: "
                    f"{missing_display_notes}"
                )

            subprocess.run([sys.executable, str(BUILD_SCRIPT)], cwd=REPO_ROOT, check=True)
            py_compile.compile(str(RUNNER), doraise=True)

            for path in [
                RUNNER,
                IMPORT_HELPER,
                ENGINE,
                FAULT6_TABLE,
                FAULT6_PREVIEW,
                FAULT6_PROVENANCE,
                BASELINE_MANIFEST,
                CORE_BASELINE_DIGEST,
                DEPENDENCY_AUDIT_JSON,
                DEPENDENCY_AUDIT_MD,
                PACKAGE_BOOTSTRAP_VERDICT,
                PACKAGE_AUDIT_SCRIPT,
                PACKAGE_VERDICT_SCRIPT,
                PACKAGE_GPVS_SCRIPT,
                PACKAGE_HEURISTIC_SCRIPT,
                PACKAGE_RAWONLY_COMMON,
                PACKAGE_RAWONLY_AUDIT_SCRIPT,
                PACKAGE_RAWONLY_VERDICT_SCRIPT,
                PACKAGE_RAWONLY_HEURISTIC_SCRIPT,
                SUMMARY,
                PACKAGE_ROOT / "bin" / "run_real.bat",
                RUN_DEMO_BAT,
                RUN_GUIDED_REAL_BAT,
                PACKAGE_ROOT / "requirements.txt",
                STAGING_PS1,
                SNAPSHOT_COPY_PS1,
                DAILY_RUN_BAT,
                INCREMENTAL_RUN_BAT,
                RUN_IMPORTED_REAL_BAT,
                RESOLVE_PYTHON_BAT,
                WINDOWS_RUNTIME_MANIFEST,
                WINDOWS_RUNTIME_PYTHON_EXE,
                README_PATH,
            ]:
                if not path.exists():
                    raise SystemExit(f"missing package file: {path}")

            subprocess.run([sys.executable, str(RUNNER), "--help"], cwd=REPO_ROOT, check=True)
            subprocess.run([sys.executable, str(IMPORT_HELPER), "--help"], cwd=REPO_ROOT, check=True)

            with tempfile.TemporaryDirectory() as tmp_dir:
                tmp_root = Path(tmp_dir)
                data_root = tmp_root / "data"
                output_root = tmp_root / "output"
                for site in ["conalog", "gangui", "ktc_ess"]:
                    make_dummy_site(data_root, site)

                subprocess.run(
                    [
                        sys.executable,
                        str(RUNNER),
                        "--data-root",
                        str(data_root),
                        "--output-root",
                        str(output_root),
                        "--dry-run",
                        "--epochs",
                        "1",
                    ],
                    cwd=REPO_ROOT,
                    check=True,
                )

                run_plan = output_root / "run_plan_v1.json"
                if not run_plan.exists():
                    raise SystemExit("dry-run must create run_plan_v1.json")
                if not (output_root / "result" / "fault6_fixed_result_table_v1.csv").exists():
                    raise SystemExit("dry-run must export fixed fault6 result table")
                if not (output_root / "result" / "fault6_label_and_algorithm_preview_v1.csv").exists():
                    raise SystemExit("dry-run must export fault6 preview artifact")
                if (output_root / "result" / "integrated_result_table_fixed_v1.csv").exists():
                    raise SystemExit("integrated result table snapshot should no longer be exported by the runtime pack")

                plan = json.loads(run_plan.read_text(encoding="utf-8"))
                if plan.get("dry_run") is not True:
                    raise SystemExit("dry-run plan must mark dry_run=true")
                if plan.get("sites") != ["conalog", "gangui", "ktc_ess"]:
                    raise SystemExit(f"unexpected site list: {plan.get('sites')}")
                if len(plan.get("site_plans", [])) != 3:
                    raise SystemExit("dry-run plan must describe 3 site plans")
                if "shadow_compare_reference_path" not in plan:
                    raise SystemExit("dry-run plan must expose shadow compare reference path")
                if "fault6_provenance_path" not in plan:
                    raise SystemExit("dry-run plan must expose fault6 provenance path")
                if "dependency_audit_json_path" not in plan:
                    raise SystemExit("dry-run plan must expose dependency audit path")
                if "live_chain" not in plan:
                    raise SystemExit("dry-run plan must describe live chain support")
                if plan["live_chain"].get("requested") is not True:
                    raise SystemExit("live chain should be requested by default in dry-run plan")
                if plan["live_chain"].get("supported") is not True:
                    raise SystemExit("dry-run plan must mark packaged live chain as supported")
                if "raw_only_chain" not in plan:
                    raise SystemExit("dry-run plan must describe raw-only chain support")
                if plan["raw_only_chain"].get("requested") is not True:
                    raise SystemExit("raw-only chain should be requested by default in dry-run plan")
                if plan["raw_only_chain"].get("supported") is not True:
                    raise SystemExit("dry-run plan must mark packaged raw-only chain as supported")
                if plan.get("prefer_existing_site_outs") != "auto":
                    raise SystemExit("dry-run plan must expose prefer-existing-site-outs default=auto")

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_root = Path(tmp_dir)
        arbitrary_root = tmp_root / "customer_drop"
        (arbitrary_root / "west_plant" / "nested").mkdir(parents=True, exist_ok=True)
        (arbitrary_root / "east_plant" / "raw").mkdir(parents=True, exist_ok=True)
        (arbitrary_root / "west_plant" / "nested" / "2026-04-01-west.csv").write_text(
            "date_time,map_type,map_id\n2026-04-01 00:00,panel,a\n",
            encoding="utf-8",
        )
        (arbitrary_root / "east_plant" / "raw" / "2026-04-01-east.csv").write_text(
            "date_time,map_type,map_id\n2026-04-01 00:00,panel,b\n",
            encoding="utf-8",
        )

        staged_root = tmp_root / "staged_data"
        manifest_path = tmp_root / "import_manifest.json"
        env_bat_path = tmp_root / "import_env.bat"
        subprocess.run(
            [
                sys.executable,
                str(IMPORT_HELPER),
                "--input-root",
                str(arbitrary_root),
                "--output-root",
                str(staged_root),
                "--clear-output",
                "--manifest-path",
                str(manifest_path),
                "--env-bat-path",
                str(env_bat_path),
            ],
            cwd=REPO_ROOT,
            check=True,
        )
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        if sorted(manifest.get("sites", [])) != ["east_plant", "west_plant"]:
            raise SystemExit(f"import helper inferred unexpected sites: {manifest.get('sites')}")
        if not (staged_root / "east_plant" / "raw" / "2026-04-01-east.csv").exists():
            raise SystemExit("import helper must stage east_plant csv into site/raw")
        if not (staged_root / "west_plant" / "raw" / "2026-04-01-west.csv").exists():
            raise SystemExit("import helper must stage west_plant csv into site/raw")
        if "IMPORTED_DATA_ROOT" not in env_bat_path.read_text(encoding="utf-8"):
            raise SystemExit("import helper must emit env bat with IMPORTED_DATA_ROOT")

    post_build_fixture_stack = ExitStack()
    atexit.register(post_build_fixture_stack.close)
    post_build_fixture_stack.enter_context(stage_missing_repo_data_link(REPO_ROOT))
    post_build_fixture_stack.enter_context(
        stage_missing_share_fixtures(REPO_ROOT, REQUIRED_SMOKE_SHARE_FIXTURES)
    )

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_root = Path(tmp_dir) / "reuse_out_run"
        subprocess.run(
            [
                sys.executable,
                str(RUNNER),
                "--data-root",
                str(REPO_ROOT / "data"),
                "--output-root",
                str(output_root),
                "--sites",
                "conalog,gangui,ktc_ess",
                "--reuse-existing-site-outs-root",
                str(REPO_ROOT / "data"),
            ],
            cwd=REPO_ROOT,
            check=True,
        )

        live_chain_dir = output_root / "result" / "live_chain"
        raw_only_chain_dir = output_root / "result" / "raw_only_chain"
        summary_path = live_chain_dir / "live_chain_summary_v1.json"
        live_fault_path = live_chain_dir / "fault_panel_result_live_v1.csv"
        live_preview_path = live_chain_dir / "fault_panel_result_live_preview_v1.csv"
        raw_only_summary_path = raw_only_chain_dir / "raw_only_chain_summary_v1.json"
        raw_only_fault_path = raw_only_chain_dir / "fault_panel_result_raw_only_v1.csv"
        raw_only_preview_path = raw_only_chain_dir / "fault_panel_result_raw_only_preview_v1.csv"
        root_live_fault_path = output_root / "result" / "fault_panel_result_current_v1.csv"
        root_live_preview_path = output_root / "result" / "fault_panel_result_current_preview_v1.csv"
        root_live_summary_path = output_root / "result" / "live_chain_summary_v1.json"
        root_live_report_path = output_root / "result" / "fault_panel_result_current_report_v1.md"
        root_master_report_path = output_root / "result" / "fault_panel_result_master_report_v1.md"
        root_detailed_report_path = output_root / "result" / "fault_panel_result_detailed_report_v1.xlsx"
        root_precursor_report_path = output_root / "result" / "fault_panel_result_precursor_report_v1.csv"
        root_fault_signal_report_path = output_root / "result" / "fault_panel_result_raw_only_fault_signal_report_v1.csv"
        root_raw_only_fault_path = output_root / "result" / "fault_panel_result_raw_only_current_v1.csv"
        root_raw_only_preview_path = output_root / "result" / "fault_panel_result_raw_only_current_preview_v1.csv"
        root_raw_only_summary_path = output_root / "result" / "raw_only_chain_summary_v1.json"
        root_raw_only_report_path = output_root / "result" / "fault_panel_result_raw_only_current_report_v1.md"
        verdict_copy_path = live_chain_dir / "panel_day_engine_panel_multiaxis_verdict_v1.csv"
        gpvs_copy_path = live_chain_dir / "panel_day_engine_gpvs_evidence_pack_v1.csv"
        heuristic_copy_path = live_chain_dir / "panel_day_engine_cause_candidate_heuristics_v1.csv"
        for path in [
            summary_path,
            live_fault_path,
            live_preview_path,
            raw_only_summary_path,
            raw_only_fault_path,
            raw_only_preview_path,
            root_live_fault_path,
            root_live_preview_path,
            root_live_summary_path,
            root_live_report_path,
            root_master_report_path,
            root_detailed_report_path,
            root_precursor_report_path,
            root_fault_signal_report_path,
            root_raw_only_fault_path,
            root_raw_only_preview_path,
            root_raw_only_summary_path,
            root_raw_only_report_path,
            verdict_copy_path,
            gpvs_copy_path,
            heuristic_copy_path,
        ]:
            if not path.exists():
                raise SystemExit(f"packaged live chain run missing output: {path}")
        detailed_book = load_workbook(root_detailed_report_path, read_only=True)
        expected_sheets = {
            "overview",
            "current_preview",
            "raw_only_preview",
            "raw_only_evidence",
            "precursor_report",
            "fault_signal_report",
            "raw_only_candidate_scores",
            "raw_only_timeline",
            "raw_only_daily_log",
            "raw_only_cluster",
            "definitions",
        }
        if not expected_sheets.issubset(set(detailed_book.sheetnames)):
            raise SystemExit(f"detailed report sheets mismatch: {detailed_book.sheetnames}")
        detailed_book.close()

        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        if summary.get("status_ko") != "completed":
            raise SystemExit(f"packaged live chain run did not complete: {summary.get('status_ko')}")
        if summary.get("fixed_fault_reference_compare", {}).get("exact_match") is not True:
            raise SystemExit("packaged live chain fault result must exactly match fixed fault reference on baseline reuse run")
        if not summary.get("generated_outputs", {}).get("fault_panel_result_live_v1"):
            raise SystemExit("packaged live chain summary must record live fault result path")
        published = summary.get("published_outputs", {})
        if not published.get("fault_panel_result_current_v1.csv"):
            raise SystemExit("packaged live chain summary must record published root live fault table")
        if not published.get("fault_panel_result_current_preview_v1.csv"):
            raise SystemExit("packaged live chain summary must record published root live preview table")
        if not published.get("live_chain_summary_v1.json"):
            raise SystemExit("packaged live chain summary must record published root live summary")
        if not published.get("fault_panel_result_current_report_v1.md"):
            raise SystemExit("packaged live chain summary must record published root live report")
        current_fault_df = pd.read_csv(root_live_fault_path, encoding="utf-8-sig")
        if current_fault_df.astype(str).applymap(contains_legacy_heuristic_display_name).any().any():
            raise SystemExit("packaged live chain current fault result must not contain legacy softened heuristic labels")

        raw_only_summary = json.loads(raw_only_summary_path.read_text(encoding="utf-8"))
        if raw_only_summary.get("status_ko") != "completed":
            raise SystemExit(f"packaged raw-only chain run did not complete: {raw_only_summary.get('status_ko')}")
        raw_only_compare = raw_only_summary.get("fixed_fault_reference_compare", {})
        if raw_only_compare.get("reference_available") is not True:
            raise SystemExit("packaged raw-only chain must compare against fixed fault reference")
        if raw_only_compare.get("candidate_row_count", 0) < raw_only_compare.get("reference_row_count", 0):
            raise SystemExit("packaged raw-only chain should not shrink below fixed reference row count on baseline reuse run")
        if raw_only_compare.get("matched_row_key_count", 0) <= 0:
            raise SystemExit("packaged raw-only chain must overlap fixed reference keys on baseline reuse run")
        if raw_only_compare.get("overlap_decision_columns_match") is not True:
            raise SystemExit("packaged raw-only chain must preserve status/event/terminal on overlapping fixed reference keys")
        raw_only_published = raw_only_summary.get("published_outputs", {})
        for name in [
            "fault_panel_result_raw_only_current_v1.csv",
            "fault_panel_result_raw_only_current_preview_v1.csv",
            "raw_only_chain_summary_v1.json",
            "fault_panel_result_raw_only_current_report_v1.md",
        ]:
            if not raw_only_published.get(name):
                raise SystemExit(f"packaged raw-only chain summary must record published output: {name}")
        root_raw_only_fault_df = pd.read_csv(root_raw_only_fault_path, encoding="utf-8-sig")
        candidate_raw_only_fault_df = pd.read_csv(raw_only_fault_path, encoding="utf-8-sig")
        for df, name in [
            (pd.read_csv(root_live_preview_path, encoding="utf-8-sig"), root_live_preview_path.name),
            (pd.read_csv(root_raw_only_preview_path, encoding="utf-8-sig"), root_raw_only_preview_path.name),
        ]:
            for column in ["전조날짜", "고장 기준일"]:
                if column not in df.columns:
                    raise SystemExit(f"{name} must include date column: {column}")
        for column in ["전조날짜", "고장날짜"]:
            if column not in root_raw_only_fault_df.columns:
                raise SystemExit(f"{root_raw_only_fault_path.name} must include date column: {column}")
        if len(root_raw_only_fault_df) > len(candidate_raw_only_fault_df):
            raise SystemExit("strict raw-only current output must not exceed candidate raw-only row count")
        publish_meta = raw_only_summary.get("publish_meta", {})
        if publish_meta.get("publish_policy_ko") != "raw_only current는 운영해석등급_ko=확정 strict subset만 노출":
            raise SystemExit("raw-only summary must document the strict current publish policy")
        if publish_meta.get("published_current_row_count") != len(root_raw_only_fault_df):
            raise SystemExit("raw-only summary published_current_row_count must match root current rows")
        precursor_report_df = pd.read_csv(root_precursor_report_path, encoding="utf-8-sig")
        fault_signal_report_df = pd.read_csv(root_fault_signal_report_path, encoding="utf-8-sig")
        for column in [
            "site",
            "panel_id",
            "운영 판정",
            "판정 근거",
            "전조날짜",
            "전조 축",
            "대표 전조 신호",
            "전조 요약",
            "상위 해석 후보",
            "기존 알고리즘 source",
            "패턴 설명",
            "모니터링 권고",
        ]:
            if column not in precursor_report_df.columns:
                raise SystemExit(f"precursor report must include column: {column}")
        if precursor_report_df.empty:
            raise SystemExit("precursor report should not be empty on baseline reuse run")
        for column in [
            "site",
            "panel_id",
            "운영 판정",
            "확정 경로",
            "고장 신호 요약",
            "전조 시작일",
            "신호 기준일",
            "사건유형",
            "사건 종결 요약",
            "상위 해석 후보",
            "기존 알고리즘 source",
            "패턴 설명",
            "현장 점검 권고",
        ]:
            if column not in fault_signal_report_df.columns:
                raise SystemExit(f"raw-only fault signal report must include column: {column}")
        if fault_signal_report_df.empty:
            raise SystemExit("raw-only fault signal report should not be empty on baseline reuse run")
        precursor_keys = set(
            zip(
                precursor_report_df["site"].astype(str),
                precursor_report_df["panel_id"].astype(str),
            )
        )
        fault_signal_keys = set(
            zip(
                fault_signal_report_df["site"].astype(str),
                fault_signal_report_df["panel_id"].astype(str),
            )
        )
        if precursor_keys & fault_signal_keys:
            raise SystemExit("precursor report and raw-only fault signal report must not overlap on row keys")

    with tempfile.TemporaryDirectory() as tmp_dir:
        output_root = Path(tmp_dir) / "auto_reuse_run"
        subprocess.run(
            [
                sys.executable,
                str(RUNNER),
                "--data-root",
                str(REPO_ROOT / "data"),
                "--output-root",
                str(output_root),
                "--sites",
                "conalog,gangui,ktc_ess",
            ],
            cwd=REPO_ROOT,
            check=True,
        )
        metadata_path = output_root / "run_metadata_v1.json"
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        if metadata.get("execution_mode_ko") != "auto_reuse_existing_site_outs":
            raise SystemExit("runner must auto-reuse data-root/<site>/out when available")
        if metadata.get("reuse_decision_ko") != "auto_fresh":
            raise SystemExit("runner must report auto_fresh reuse decision when data-root out is newer than raw")
        freshness = metadata.get("reuse_freshness", {}).get("sites", {})
        if not freshness:
            raise SystemExit("runner must record reuse freshness details")
        if not all(entry.get("fresh_enough") is True for entry in freshness.values()):
            raise SystemExit("auto-reuse smoke expects all packaged baseline site outs to be fresh enough")
        if metadata.get("live_chain", {}).get("status_ko") != "completed":
            raise SystemExit("auto-reuse run must still complete live chain")
        if metadata.get("raw_only_chain", {}).get("status_ko") != "completed":
            raise SystemExit("auto-reuse run must still complete raw-only chain")
        if not (output_root / "result" / "fault_panel_result_current_report_v1.md").exists():
            raise SystemExit("auto-reuse run must publish current root report")
        if not (output_root / "result" / "fault_panel_result_master_report_v1.md").exists():
            raise SystemExit("auto-reuse run must publish master root report")
        if not (output_root / "result" / "fault_panel_result_detailed_report_v1.xlsx").exists():
            raise SystemExit("auto-reuse run must publish detailed xlsx report")
        if not (output_root / "result" / "fault_panel_result_precursor_report_v1.csv").exists():
            raise SystemExit("auto-reuse run must publish precursor report")
        if not (output_root / "result" / "fault_panel_result_raw_only_fault_signal_report_v1.csv").exists():
            raise SystemExit("auto-reuse run must publish raw-only fault signal report")

    fault_df = pd.read_csv(FAULT6_TABLE, encoding="utf-8-sig")
    if fault_df.columns.tolist() != EXPECTED_FAULT6_COLS:
        raise SystemExit(f"fault6 artifact schema mismatch: {fault_df.columns.tolist()}")
    if len(fault_df) != 6:
        raise SystemExit(f"fault6 artifact must have 6 rows, found {len(fault_df)}")

    frozen_verdict = pd.read_csv(FROZEN_VERDICT, encoding="utf-8-sig", low_memory=False)
    frozen_heuristic = pd.read_csv(FROZEN_HEURISTIC, encoding="utf-8-sig", low_memory=False)
    expected_fault = frozen_verdict.loc[
        frozen_verdict["패널고장여부_ko"].astype(str).eq("고장"),
        ["site", "panel_id", "패널고장여부_ko", "사건유형_ko", "최종고장양상_ko", "커널로그_원인군_ko"],
    ].copy()
    heuristic_cols = ["site", "panel_id", "원인후보_top1_ko", "원인후보_top2_ko", "원인후보_top3_ko"]
    expected_fault = expected_fault.merge(
        frozen_heuristic[heuristic_cols],
        on=["site", "panel_id"],
        how="left",
    )
    expected_fault["1순위_의심원인_ko"] = expected_fault["원인후보_top1_ko"].map(
        lambda value: DISPLAY_HEURISTIC_NAME_MAP.get(str(value), str(value))
    )
    expected_fault["2순위_의심원인_ko"] = expected_fault["원인후보_top2_ko"].map(
        lambda value: DISPLAY_HEURISTIC_NAME_MAP.get(str(value), str(value))
    )
    expected_fault["3순위_의심원인_ko"] = expected_fault["원인후보_top3_ko"].map(
        lambda value: DISPLAY_HEURISTIC_NAME_MAP.get(str(value), str(value))
    )
    expected_fault = (
        expected_fault[
            ["site", "panel_id", "패널고장여부_ko", "사건유형_ko", "최종고장양상_ko", "커널로그_원인군_ko", "1순위_의심원인_ko", "2순위_의심원인_ko", "3순위_의심원인_ko"]
        ]
        .sort_values(["site", "panel_id"], ascending=[True, True])
        .reset_index(drop=True)
    )
    if not fault_df.fillna("").astype(str).equals(expected_fault.fillna("").astype(str)):
        raise SystemExit("fault6 artifact must exactly match frozen verdict+heuristic-derived fault table")

    preview_df = pd.read_csv(FAULT6_PREVIEW, encoding="utf-8-sig")
    if preview_df.columns.tolist() != EXPECTED_PREVIEW_COLS:
        raise SystemExit(f"preview artifact schema mismatch: {preview_df.columns.tolist()}")
    if len(preview_df) != 6:
        raise SystemExit(f"preview artifact must have 6 rows, found {len(preview_df)}")
    if set(preview_df["운영 판정"].astype(str)) != {"확정"}:
        raise SystemExit("preview artifact should mark all fixed fault6 rows as 확정")
    preview_date_map = {
        (str(row["site"]), str(row["panel_id"])): (str(row["전조날짜"]), str(row["고장 기준일"]))
        for row in preview_df.to_dict(orient="records")
    }
    for key, expected_dates in EXPECTED_PREVIEW_DATE_ROWS.items():
        if preview_date_map.get(key) != expected_dates:
            raise SystemExit(
                f"preview artifact date labels mismatch for {key}: expected={expected_dates} actual={preview_date_map.get(key)}"
            )
    provenance = json.loads(FAULT6_PROVENANCE.read_text(encoding="utf-8"))
    if provenance.get("legacy_integrated_exact_match") is not True:
        raise SystemExit("fault6 provenance must confirm exact match to legacy integrated 6-row slice")
    if provenance.get("source_chain_ko") != "frozen verdict plus frozen heuristic with integrated display-name mapping":
        raise SystemExit("fault6 provenance must document the new direct source chain")
    if provenance.get("display_name_map") != DISPLAY_HEURISTIC_NAME_MAP:
        raise SystemExit("fault6 provenance display_name_map must exactly match shared registry")
    if any(legacy_name in preview_df.to_csv(index=False) for legacy_name in LEGACY_HEURISTIC_DISPLAY_NAMES):
        raise SystemExit("preview artifact must not contain legacy softened heuristic labels")
    if any(legacy_name in fault_df.to_csv(index=False) for legacy_name in LEGACY_HEURISTIC_DISPLAY_NAMES):
        raise SystemExit("fault6 artifact must not contain legacy softened heuristic labels")

    baseline = json.loads(BASELINE_MANIFEST.read_text(encoding="utf-8"))
    for site in ["conalog", "gangui", "ktc_ess"]:
        if site not in baseline.get("sites", {}):
            raise SystemExit(f"baseline manifest missing site: {site}")

    core_digest = json.loads(CORE_BASELINE_DIGEST.read_text(encoding="utf-8"))
    for site in ["conalog", "gangui", "ktc_ess"]:
        if site not in core_digest.get("sites", {}):
            raise SystemExit(f"core baseline digest missing site: {site}")
        digest_entry = core_digest["sites"][site]
        for key in ["row_count", "digest_sha256", "critical_source_counts", "final_fault_true_count"]:
            if key not in digest_entry:
                raise SystemExit(f"core baseline digest missing key={key} for site={site}")

    dependency_audit = json.loads(DEPENDENCY_AUDIT_JSON.read_text(encoding="utf-8"))
    if dependency_audit.get("runtime_live_full_chain_ready_flag") is not False:
        raise SystemExit("dependency audit must mark live full chain as not ready")
    if dependency_audit.get("hard_cycle", {}).get("verified_flag") is not True:
        raise SystemExit("dependency audit must document the verified hard cycle")
    runtime_manifest = json.loads(WINDOWS_RUNTIME_MANIFEST.read_text(encoding="utf-8"))
    if runtime_manifest.get("python_version") != "3.11.9":
        raise SystemExit(f"unexpected portable runtime python version: {runtime_manifest.get('python_version')}")
    if "torch==2.9.1" not in runtime_manifest.get("primary_packages", []):
        raise SystemExit("portable runtime manifest must include torch==2.9.1")
    if "openpyxl==3.1.5" not in runtime_manifest.get("primary_packages", []):
        raise SystemExit("portable runtime manifest must include openpyxl==3.1.5")
    if runtime_manifest.get("wheel_count", 0) < 4:
        raise SystemExit("portable runtime manifest must report downloaded wheels")

    for filename in [
        "panel_day_engine_operator_workflow_default_v1.csv",
        "panel_day_engine_fault_panel_event_audit_v1.csv",
        "panel_day_engine_panel_multiaxis_verdict_v1.csv",
        "panel_day_engine_gpvs_evidence_pack_v1.csv",
        "panel_date_reaudit_working.csv",
    ]:
        if not (PACKAGE_SHARE_ROOT / filename).exists():
            raise SystemExit(f"packaged live chain share input missing: {filename}")

    daily_run_text = DAILY_RUN_BAT.read_text(encoding="utf-8", errors="ignore")
    if "stage_recent_120d.ps1" not in daily_run_text or "run_full_algorithm_pack.py" not in daily_run_text:
        raise SystemExit("daily_run.bat must stage recent data and invoke run_full_algorithm_pack.py")
    if "resolve_python.bat" not in daily_run_text:
        raise SystemExit("daily_run.bat must resolve embedded/system python through resolve_python.bat")
    if "fault_panel_result_master_report_v1.md" not in daily_run_text:
        raise SystemExit("daily_run.bat must prefer opening the master report")
    if "fault_panel_result_raw_only_current_preview_v1.csv" in daily_run_text:
        raise SystemExit("daily_run.bat must not auto-open raw-only preview in the operator default path")
    snapshot_text = SNAPSHOT_COPY_PS1.read_text(encoding="utf-8", errors="ignore")
    if "StableMinutes" not in snapshot_text or "Move-Item" not in snapshot_text:
        raise SystemExit("snapshot_copy.ps1 must implement stable-minute gating and atomic rename")
    incremental_text = INCREMENTAL_RUN_BAT.read_text(encoding="utf-8", errors="ignore")
    if "import_any_csv_root.py" not in incremental_text or "run_full_algorithm_pack.py" not in incremental_text:
        raise SystemExit("incremental_run.bat must invoke import_any_csv_root.py then run_full_algorithm_pack.py")
    if "--stable-minutes" not in incremental_text:
        raise SystemExit("incremental_run.bat must pass stable-minutes to import_any_csv_root.py")
    if "[040%%] 학습/추론 및 결과표 생성을 시작합니다." not in incremental_text:
        raise SystemExit("incremental_run.bat must show a visible stage progress message before runner execution")
    if "resolve_python.bat" not in incremental_text:
        raise SystemExit("incremental_run.bat must resolve embedded/system python through resolve_python.bat")
    if "fault_panel_result_master_report_v1.md" not in incremental_text:
        raise SystemExit("incremental_run.bat must prefer opening the master report")
    if "fault_panel_result_raw_only_current_preview_v1.csv" in incremental_text:
        raise SystemExit("incremental_run.bat must not auto-open raw-only preview in the operator default path")
    run_real_text = (PACKAGE_ROOT / "bin" / "run_real.bat").read_text(encoding="utf-8", errors="ignore")
    if "FolderBrowserDialog" not in run_real_text or "run_full_algorithm_pack.py" not in run_real_text:
        raise SystemExit("run_real.bat must use folder picker and invoke run_full_algorithm_pack.py")
    if "import_any_csv_root.py" not in run_real_text:
        raise SystemExit("run_real.bat must fallback to import_any_csv_root.py for arbitrary folder structures")
    if "[040%%] 학습/추론 및 결과표 생성을 시작합니다." not in run_real_text:
        raise SystemExit("run_real.bat must show a visible stage progress message before runner execution")
    if "resolve_python.bat" not in run_real_text:
        raise SystemExit("run_real.bat must resolve embedded/system python through resolve_python.bat")
    if "run_oneclick.py" in run_real_text or "run_realtime.py" in run_real_text:
        raise SystemExit("run_real.bat must stay on the stable full-runtime path")
    if "fault_panel_result_master_report_v1.md" not in run_real_text:
        raise SystemExit("run_real.bat must prefer opening the master report")
    if "fault_panel_result_raw_only_current_preview_v1.csv" in run_real_text:
        raise SystemExit("run_real.bat must not auto-open raw-only preview in the operator default path")
    imported_real_text = RUN_IMPORTED_REAL_BAT.read_text(encoding="utf-8", errors="ignore")
    if "import_any_csv_root.py" not in imported_real_text or "run_full_algorithm_pack.py" not in imported_real_text:
        raise SystemExit("run_imported_real.bat must import arbitrary folder roots then invoke run_full_algorithm_pack.py")
    if "[040%%] 학습/추론 및 결과표 생성을 시작합니다." not in imported_real_text:
        raise SystemExit("run_imported_real.bat must show a visible stage progress message before runner execution")
    run_demo_text = RUN_DEMO_BAT.read_text(encoding="utf-8", errors="ignore")
    if "fault6_label_and_algorithm_preview_v1.csv" not in run_demo_text:
        raise SystemExit("run_demo.bat must open the packaged fault preview artifact")
    if "integrated_result_table_fixed_v1.csv" in run_demo_text:
        raise SystemExit("run_demo.bat must no longer reference integrated result table snapshot")
    guided_real_text = RUN_GUIDED_REAL_BAT.read_text(encoding="utf-8", errors="ignore")
    for phrase in [
        "showcase_runs",
        "import_any_csv_root.py",
        "run_full_algorithm_pack.py",
        "fault_panel_result_master_report_v1.md",
        "[040%%] 학습/추론 및 결과표 생성을 시작합니다.",
        "pause",
    ]:
        if phrase not in guided_real_text:
            raise SystemExit(f"run_guided_real.bat missing phrase: {phrase}")
    if "fault_panel_result_raw_only_current_preview_v1.csv" in guided_real_text:
        raise SystemExit("run_guided_real.bat must not auto-open raw-only preview in the operator default path")
    if "fault_panel_result_raw_only_current_preview_v1.csv" in imported_real_text:
        raise SystemExit("run_imported_real.bat must not auto-open raw-only preview in the operator default path")
    resolve_python_text = RESOLVE_PYTHON_BAT.read_text(encoding="utf-8", errors="ignore")
    if "runtime\\windows_x64\\python\\python.exe" not in resolve_python_text:
        raise SystemExit("resolve_python.bat must prefer the packaged embedded python runtime")

    readme_text = README_PATH.read_text(encoding="utf-8", errors="ignore")
    for phrase in [
        "최근 120일",
        "상위 해석 후보",
        "운영 판정",
        "판정 근거",
        "고장 기준일",
        "기존 알고리즘 source",
        "급락 종결 관측",
        "점진 저하 누적",
        "사건 종결 요약",
        "패턴 설명",
        "fault_panel_result_detailed_report_v1.xlsx",
        "openpyxl",
        "legacy integrated 6행과 exact match",
        "shadow compare",
        "runtime_chain_dependency_audit_v1",
        "frozen verdict와 heuristic",
        "bootstrap verdict",
        "reuse-existing-site-outs-root",
        "prefer-existing-site-outs",
        "data-root/<site>/out",
        "snapshot_copy.ps1",
        "incremental_run.bat",
        "run_imported_real.bat",
        "run_guided_real.bat",
        "import_any_csv_root.py",
        "embedded Python 3.11.9",
        "runtime\\windows_x64\\python\\python.exe",
        "StableMinutes",
        "showcase_runs",
        "fault_panel_result_current_v1.csv",
        "fault_panel_result_current_preview_v1.csv",
        "fault_panel_result_current_report_v1.md",
        "fault_panel_result_master_report_v1.md",
    ]:
        if phrase not in readme_text:
            raise SystemExit(f"README missing phrase: {phrase}")

    after_watch = OUT_WATCH.read_bytes()
    if before_watch != after_watch:
        raise SystemExit("frozen integrated result table was modified by runtime pack build/test")
    post_build_fixture_stack.close()

    print("[OK] conalog_full_runtime_v1 smoke test passed")


if __name__ == "__main__":
    main()
