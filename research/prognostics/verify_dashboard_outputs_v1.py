#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - target environment reports this.
    load_workbook = None


PRIMARY_PREVIEW_NAME = "fault_panel_result_current_preview_v1.csv"
CURRENT_RESULT_NAME = "fault_panel_result_current_v1.csv"
PRECURSOR_REPORT_NAME = "fault_panel_result_precursor_report_v1.csv"
MASTER_REPORT_NAME = "fault_panel_result_master_report_v1.md"
DETAILED_REPORT_NAME = "fault_panel_result_detailed_report_v1.xlsx"
RAW_ONLY_SIGNAL_REPORT_NAME = "fault_panel_result_raw_only_fault_signal_report_v1.csv"

PRIMARY_PREVIEW_COLUMNS = [
    "site",
    "panel_id",
    "전조날짜",
    "고장 기준일",
    "운영 판정",
    "급락 종결 관측",
    "점진 저하 누적",
    "사건 종결 요약",
    "상위 해석 후보",
    "기존 알고리즘 source",
]

CURRENT_RESULT_REQUIRED_COLUMNS = [
    "site",
    "panel_id",
    "패널고장여부_ko",
    "사건유형_ko",
    "최종고장양상_ko",
]

PRECURSOR_REQUIRED_COLUMNS = [
    "site",
    "panel_id",
    "운영 판정",
    "판정 근거",
    "전조날짜",
    "상위 해석 후보",
    "패턴 설명",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Verify dashboard-facing pvdiag runtime outputs after run_full_algorithm_pack.py."
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        required=True,
        help="Output root passed to run_full_algorithm_pack.py.",
    )
    parser.add_argument(
        "--json-out",
        type=Path,
        default=None,
        help="Optional JSON summary path. Defaults to <output-root>/dashboard_output_check_v1.json.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise SystemExit(f"missing required output: {path}")
    return pd.read_csv(path, encoding="utf-8-sig", low_memory=False)


def require_columns(df: pd.DataFrame, columns: list[str], path: Path) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise SystemExit(f"{path.name} missing required columns: {missing}")


def require_file(path: Path) -> None:
    if not path.exists():
        raise SystemExit(f"missing required output: {path}")


def result_dir(output_root: Path) -> Path:
    candidate = output_root / "result"
    if candidate.exists():
        return candidate
    if output_root.name == "result":
        return output_root
    raise SystemExit(f"missing result directory under output root: {candidate}")


def workbook_sheet_names(path: Path) -> list[str]:
    if load_workbook is None:
        raise SystemExit("openpyxl is required to verify detailed xlsx outputs")
    workbook = load_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()
    root = args.output_root.resolve()
    res_dir = result_dir(root)

    primary_path = res_dir / PRIMARY_PREVIEW_NAME
    current_path = res_dir / CURRENT_RESULT_NAME
    precursor_path = res_dir / PRECURSOR_REPORT_NAME
    master_report_path = res_dir / MASTER_REPORT_NAME
    detailed_report_path = res_dir / DETAILED_REPORT_NAME
    raw_only_signal_path = res_dir / RAW_ONLY_SIGNAL_REPORT_NAME

    primary_df = read_csv(primary_path)
    require_columns(primary_df, PRIMARY_PREVIEW_COLUMNS, primary_path)

    current_df = read_csv(current_path)
    require_columns(current_df, CURRENT_RESULT_REQUIRED_COLUMNS, current_path)

    precursor_df = read_csv(precursor_path)
    require_columns(precursor_df, PRECURSOR_REQUIRED_COLUMNS, precursor_path)

    require_file(master_report_path)
    require_file(detailed_report_path)
    require_file(raw_only_signal_path)
    sheet_names = workbook_sheet_names(detailed_report_path)

    payload = {
        "generated_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "status": "pass",
        "output_root": str(root),
        "result_dir": str(res_dir),
        "primary_preview": {
            "path": str(primary_path),
            "row_count": int(len(primary_df)),
            "columns": primary_df.columns.tolist(),
        },
        "current_result": {
            "path": str(current_path),
            "row_count": int(len(current_df)),
        },
        "precursor_report": {
            "path": str(precursor_path),
            "row_count": int(len(precursor_df)),
        },
        "master_report_path": str(master_report_path),
        "detailed_report": {
            "path": str(detailed_report_path),
            "sheet_names": sheet_names,
        },
        "raw_only_fault_signal_report_path": str(raw_only_signal_path),
        "note_ko": "dashboard primary CSV와 주요 support artifact의 존재/schema를 확인했다.",
    }

    json_out = args.json_out or (root / "dashboard_output_check_v1.json")
    json_out.parent.mkdir(parents=True, exist_ok=True)
    json_out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"[OK] dashboard output contract verified: {json_out}")


if __name__ == "__main__":
    main()
